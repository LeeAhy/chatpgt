from __future__ import annotations

import math
import re
from pathlib import Path

from runtime_bootstrap import ensure_bundled_python_path

ensure_bundled_python_path()

from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import range_boundaries, get_column_letter


PATH = Path("/Users/chandelar/Documents/销售排单/outputs/sales_fill_0602/2026年第一事业部销售排单-0602_已回填.xlsx")


def is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool) and v == v


def to_number(v):
    if v is None or v == "":
        return 0
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        if v.startswith("#"):
            return 0
        try:
            return float(v)
        except Exception:
            return 0
    return 0


def to_text(v):
    if v is None:
        return ""
    return str(v)


class ExcelArray:
    def __init__(self, values):
        self.values = list(values)

    def __iter__(self):
        return iter(self.values)

    def __len__(self):
        return len(self.values)

    def _numeric_op(self, other, op):
        if isinstance(other, ExcelArray):
            return ExcelArray([op(to_number(a), to_number(b)) for a, b in zip(self.values, other.values)])
        return ExcelArray([op(to_number(a), to_number(other)) for a in self.values])

    def _compare_op(self, other, op):
        if isinstance(other, ExcelArray):
            return ExcelArray([op(a, b) for a, b in zip(self.values, other.values)])
        return ExcelArray([op(a, other) for a in self.values])

    def __add__(self, other):
        return self._numeric_op(other, lambda a, b: a + b)

    def __radd__(self, other):
        return self.__add__(other)

    def __sub__(self, other):
        return self._numeric_op(other, lambda a, b: a - b)

    def __rsub__(self, other):
        if isinstance(other, ExcelArray):
            return other.__sub__(self)
        return ExcelArray([to_number(other) - to_number(a) for a in self.values])

    def __mul__(self, other):
        return self._numeric_op(other, lambda a, b: a * b)

    def __rmul__(self, other):
        return self.__mul__(other)

    def __truediv__(self, other):
        def div(a, b):
            try:
                return a / b
            except ZeroDivisionError:
                return 0

        return self._numeric_op(other, div)

    def __rtruediv__(self, other):
        def div(a, b):
            try:
                return a / b
            except ZeroDivisionError:
                return 0

        if isinstance(other, ExcelArray):
            return other.__truediv__(self)
        return ExcelArray([div(to_number(other), to_number(a)) for a in self.values])

    def __pow__(self, other):
        return self._numeric_op(other, lambda a, b: a**b)

    def __neg__(self):
        return ExcelArray([-to_number(a) for a in self.values])

    def __pos__(self):
        return ExcelArray([+to_number(a) for a in self.values])

    def __eq__(self, other):
        return self._compare_op(other, lambda a, b: a == b)

    def __ne__(self, other):
        return self._compare_op(other, lambda a, b: a != b)

    def __lt__(self, other):
        return self._compare_op(other, lambda a, b: a < b)

    def __le__(self, other):
        return self._compare_op(other, lambda a, b: a <= b)

    def __gt__(self, other):
        return self._compare_op(other, lambda a, b: a > b)

    def __ge__(self, other):
        return self._compare_op(other, lambda a, b: a >= b)


def flatten(arg):
    if isinstance(arg, ExcelArray):
        return arg.values
    return [arg]


def sum_helper(*args):
    total = 0
    for arg in args:
        for v in flatten(arg):
            total += to_number(v)
    return total


def subtotal_helper(function_num, arg):
    # Only the SUM variant is used in this workbook.
    if int(to_number(function_num)) != 9:
        return sum_helper(arg)
    return sum_helper(arg)


def sumproduct_helper(*args):
    arrays = [flatten(a) for a in args]
    if not arrays:
        return 0
    length = max(len(a) for a in arrays)
    total = 0
    for i in range(length):
        prod = 1
        for arr in arrays:
            value = arr[i] if len(arr) > 1 else arr[0]
            prod *= to_number(value)
        total += prod
    return total


def criterion_matches(value, criteria):
    if isinstance(criteria, ExcelArray):
        criteria = criteria.values[0] if criteria.values else ""
    if isinstance(criteria, str):
        m = re.match(r"^(<=|>=|<>|=|<|>)(.*)$", criteria)
        if m:
            op, rhs = m.groups()
            rhs = rhs.strip()
        else:
            op, rhs = "=", criteria
        val_num = None
        rhs_num = None
        try:
            val_num = float(value)
        except Exception:
            pass
        try:
            rhs_num = float(rhs)
        except Exception:
            pass
        left = val_num if val_num is not None and rhs_num is not None else to_text(value)
        right = rhs_num if val_num is not None and rhs_num is not None else rhs
        if op == "=":
            return left == right
        if op == "<>":
            return left != right
        if op == ">":
            return left > right
        if op == "<":
            return left < right
        if op == ">=":
            return left >= right
        if op == "<=":
            return left <= right
        return False
    else:
        return value == criteria


def sumifs_helper(sum_range, *criteria_pairs):
    values = flatten(sum_range)
    if len(criteria_pairs) % 2 != 0:
        raise ValueError("SUMIFS criteria must be in pairs")
    masks = None
    for i in range(0, len(criteria_pairs), 2):
        crange = flatten(criteria_pairs[i])
        crit = criteria_pairs[i + 1]
        current = [criterion_matches(v, crit) for v in crange]
        masks = current if masks is None else [a and b for a, b in zip(masks, current)]
    total = 0
    if masks is None:
        masks = [True] * len(values)
    for v, keep in zip(values, masks):
        if keep:
            total += to_number(v)
    return total


def split_ref(ref, current_sheet):
    ref = ref.replace("$", "")
    if "!" in ref:
        sheet_part, addr = ref.split("!", 1)
        sheet_name = sheet_part.strip("'")
    else:
        sheet_name = current_sheet
        addr = ref
    return sheet_name, addr


def ref_value(wb, cache, current_sheet, ref):
    sheet_name, addr = split_ref(ref, current_sheet)
    key = (sheet_name, addr)
    if key in cache:
        return cache[key]
    ws = wb[sheet_name]
    value = ws[addr].value
    if isinstance(value, str) and value.startswith("="):
        value = evaluate_formula(wb, cache, sheet_name, ws[addr].coordinate, value)
    cache[key] = value
    return value


def range_values(wb, cache, current_sheet, ref):
    sheet_name, addr = split_ref(ref, current_sheet)
    min_col, min_row, max_col, max_row = range_boundaries(addr)
    ws = wb[sheet_name]
    vals = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                value = evaluate_formula(wb, cache, sheet_name, cell.coordinate, value)
            vals.append(value)
    return ExcelArray(vals)


def translate_formula(formula, current_sheet):
    tok = Tokenizer(formula)
    pieces = []
    for t in tok.items:
        if t.type == "OPERAND" and t.subtype == "RANGE":
            ref = t.value
            if ":" in ref:
                pieces.append(f"RANGE({ref!r}, current_sheet)")
            else:
                pieces.append(f"REF({ref!r}, current_sheet)")
        elif t.type == "OPERAND" and t.subtype == "NUMBER":
            pieces.append(t.value)
        elif t.type == "OPERAND" and t.subtype == "TEXT":
            pieces.append(repr(t.value))
        elif t.type == "FUNC" and t.subtype == "OPEN":
            pieces.append(t.value[:-1].upper() + "(")
        elif t.type == "FUNC" and t.subtype == "CLOSE":
            pieces.append(")")
        elif t.type == "SEP" and t.subtype == "ARG":
            pieces.append(",")
        elif t.type == "OPERATOR-INFIX":
            pieces.append({"=": "==", "<>": "!=", "^": "**"}.get(t.value, t.value))
        elif t.type == "OPERATOR-PREFIX":
            pieces.append(t.value)
        elif t.type == "PAREN" and t.subtype == "OPEN":
            pieces.append("(")
        elif t.type == "PAREN" and t.subtype == "CLOSE":
            pieces.append(")")
        else:
            pieces.append(t.value)
    return "".join(pieces)


def evaluate_formula(wb, cache, sheet_name, coord, formula):
    expr = translate_formula(formula, sheet_name)
    row = int(re.search(r"(\d+)$", coord).group(1))

    def REF(ref, current_sheet=sheet_name):
        return ref_value(wb, cache, current_sheet, ref)

    def RANGE(ref, current_sheet=sheet_name):
        return range_values(wb, cache, current_sheet, ref)

    def ROW(*args):
        if not args:
            return row
        arg = args[0]
        if isinstance(arg, ExcelArray):
            return row
        if isinstance(arg, str):
            return ref_value(wb, cache, sheet_name, arg)
        return row

    env = {
        "REF": REF,
        "RANGE": RANGE,
        "ROW": ROW,
        "SUM": sum_helper,
        "SUBTOTAL": subtotal_helper,
        "SUMIFS": sumifs_helper,
        "SUMPRODUCT": sumproduct_helper,
        "ABS": abs,
        "MIN": min,
        "MAX": max,
        "ROUND": round,
        "INT": int,
        "FLOAT": float,
        "AND": lambda *args: all(bool(a) for a in args),
        "OR": lambda *args: any(bool(a) for a in args),
        "NOT": lambda x: not bool(x),
        "current_sheet": sheet_name,
    }

    try:
        result = eval(expr, {"__builtins__": {}}, env)
        if isinstance(result, ExcelArray):
            # Array formulas are not expected as final values in this workbook.
            result = result.values[0] if result.values else 0
        if isinstance(result, bool):
            return int(result)
        if result is None:
            return 0
        if isinstance(result, float) and (math.isnan(result) or math.isinf(result)):
            return 0
        return result
    except Exception:
        return 0


def freeze_workbook_inplace(path: Path | str) -> int:
    wb = load_workbook(path, read_only=False, data_only=False)
    cache = {}
    formulas = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append((ws.title, cell.coordinate, cell.value))

    print(f"formula cells: {len(formulas)}")
    for i, (sheet_name, coord, formula) in enumerate(formulas, start=1):
        value = evaluate_formula(wb, cache, sheet_name, coord, formula)
        wb[sheet_name][coord] = value
        if i % 10000 == 0:
            print(f"  processed {i}/{len(formulas)}")

    wb.save(path)
    print(f"saved recalculated workbook: {path}")
    return len(formulas)


def main():
    freeze_workbook_inplace(PATH)


if __name__ == "__main__":
    main()
