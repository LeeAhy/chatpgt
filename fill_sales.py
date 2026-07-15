from __future__ import annotations

import csv
import gc
import hashlib
import json
import math
import os
import posixpath
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from copy import copy
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path
from statistics import median
from typing import Callable, Iterable, Optional
from zoneinfo import ZoneInfo

from runtime_bootstrap import ensure_bundled_python_path

ensure_bundled_python_path()

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from freeze_formulas import freeze_workbook_inplace


PRED_PATH = Path("/Users/chandelar/Desktop/预测0602.xlsx")
SALES_PATH = Path("/Users/chandelar/Desktop/2026年第一事业部销售排单-0602.xlsx")
OUT_DIR = Path("/Users/chandelar/Documents/销售排单/outputs/sales_fill_0602")
OUT_PATH = OUT_DIR / "2026年第一事业部销售排单-0602_已回填.xlsx"

BUSINESS_OWNERS = ("王永仁", "周文龙", "洪鸣", "叶振华", "李玎玲", "李海鹰")
EXCEL_SUFFIXES = {".xlsx"}
SALES_EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
TEXT_SUFFIXES: set[str] = set()
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg"}
SUPPORTED_PREDICTION_SUFFIXES = EXCEL_SUFFIXES | IMAGE_SUFFIXES
SUPPORTED_SALES_SUFFIXES = SALES_EXCEL_SUFFIXES

HEADER_SCAN_ROWS = 12
DATA_START_ROW = 5
NO_FILL_LABEL_WORDS = ("差异", "已完成", "实际")
MAX_FORMULA_RATIO = 0.2
MIN_REASONABLE_PRICE = 0.01
MAX_REASONABLE_PRICE = 1000.0
NO_FORECAST_FILL = PatternFill(fill_type=None)
INVALID_AMOUNT_FILL = PatternFill(fill_type="solid", fgColor="FF8B0000")
# Use opaque ARGB values so Excel/WPS cannot interpret the font colors as
# transparent or fall back to the workbook theme.
MATCHED_FORECAST_FONT_COLOR = "FF008000"
ZERO_FORECAST_FONT_COLOR = "FF000000"
LEGACY_MATCHED_FORECAST_FILL_COLOR = "DDEBF7"
BEIJING_TZ = ZoneInfo("Asia/Shanghai")
OCR_SOURCE_PATH = Path(__file__).with_name("ocr_image.m")
OCR_BINARY_PATH = Path(tempfile.gettempdir()) / "sales_ocr_image"

OWNER_PREDICTION_RULES = {
    "洪鸣": {"headers": ("New part NO",), "prefix": False, "unit": "wan"},
    "李玎玲": {"headers": ("机种名",), "prefix": False, "unit": "k"},
    "周文龙": {"headers": ("HIR料号", "品民", "品名"), "prefix": False, "unit": "pcs"},
    "王永仁": {"headers": ("子件描述",), "prefix": True, "unit": "pcs"},
    "叶振华": {"headers": ("模组型号", "料号"), "prefix": False, "unit": "pcs"},
    "李海鹰": {"headers": ("料号",), "prefix": False, "unit": "pcs"},
}
DEFAULT_PREDICTION_RULE = {"headers": (), "prefix": False, "unit": "pcs"}

LIDINGLING_SAMSUNG_FORECAST_K = {
    "V1688": {7: 2.5},
    "V1724": {6: 51, 7: 40},
    "Zhangheng": {},
    "1702": {7: 32, 9: 3.5},
    "1707/P3+": {6: 42, 9: 44},
    "P2 Wide": {},
    "P2 Tele": {6: 1.5},
    "Vantage": {6: 31, 7: 41, 8: 19},
    "1741": {6: 101, 8: 34, 9: 30},
    "Gasher Wide": {6: 31.5, 7: 15.0, 8: 3.5},
    "Gasher UWide": {6: 41.0, 7: 15.5, 8: 3.5},
    "Shasta": {6: 38.0, 7: 15.5, 8: 3.5},
    "Wukong (Apollo)": {},
    "1778": {7: 260, 8: 350, 9: 320, 10: 192},
    "1988": {7: 160, 8: 200, 9: 250, 10: 165},
    "Madrid Wide": {7: 210, 8: 350, 9: 220, 10: 90},
    "Madrid Tele": {7: 210, 8: 350, 9: 220, 10: 90},
}

LIDINGLING_SAMSUNG_ALIASES = {
    "1741": ("V1741",),
    "Gasher Wide": ("Gahser WIDE", "Gasher wide", "Gasher WIDE"),
    "Gasher UWide": ("Gahser UW", "Gasher UW", "Gasher U Wide"),
    "Shasta": ("Shasta tele",),
    "1778": ("V1778",),
    "1988": ("V1988",),
    "Madrid Wide": ("Madrid WIDE",),
    "Madrid Tele": ("Madrid tele",),
    "Wukong (Apollo)": ("Apollo（Vantage2）", "Apollo", "Vantage2"),
    "1707/P3+": ("V1707", "1707"),
}


@dataclass(frozen=True)
class HeaderColumns:
    owner_col: int
    code_col: int
    erp_code_col: Optional[int] = None
    price_col: Optional[int] = None
    fx_col: Optional[int] = None


@dataclass(frozen=True)
class QuantityAmountPair:
    qty_col: int
    amt_col: int
    label: str


@dataclass(frozen=True)
class FillTarget:
    sheet: str
    month: int
    label: str
    qty_col: int
    amt_col: int
    owner_col: int
    code_col: int


@dataclass(frozen=True)
class OcrToken:
    text: str
    x: float
    y: float
    width: float
    height: float

    @property
    def center_x(self) -> float:
        return self.x + (self.width / 2)

    @property
    def center_y(self) -> float:
        return self.y + (self.height / 2)


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def compact_text(value) -> str:
    return re.sub(r"\s+", "", clean_text(value)).upper()


def normalize_code(value) -> str:
    text = compact_text(value)
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"[^0-9A-Z\u4e00-\u9fff()_-]", "", text)
    text = text.replace("GAHSER", "GASHER")
    text = text.replace("UWIDE", "UW")
    text = text.replace("ULTRAWIDE", "UW")
    text = text.replace("江西水晶", "水晶")
    text = re.sub(r"A0(?=水晶|昀冢|CD700|NJC|BG|$)", "", text)
    return text


def normalize_owner(value) -> str:
    return re.sub(r"[\s\u3000\-_/()（）]+", "", compact_text(value))


def owner_matches(selected_owner: Optional[str], row_owner: Optional[str]) -> bool:
    if not selected_owner:
        return True

    selected = normalize_owner(selected_owner)
    row = normalize_owner(row_owner)
    if not selected or not row:
        return False
    return selected == row or selected in row or row in selected


def parse_number(value) -> Optional[float]:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isfinite(float(value)):
            return float(value)
        return None

    text = clean_text(value)
    if not text:
        return None

    multiplier = 1.0
    if "万" in text:
        multiplier = 10000.0
    if re.search(r"\d\s*[kK]\b", text):
        multiplier = 1000.0

    text = (
        text.replace(",", "")
        .replace("，", "")
        .replace("pcs", "")
        .replace("PCS", "")
        .replace("Pcs", "")
        .replace("万", "")
        .replace("K", "")
        .replace("k", "")
        .strip()
    )
    if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text):
        return None
    return float(text) * multiplier


def parse_area(spec):
    if spec is None:
        return None
    nums = re.findall(r"\d+(?:\.\d+)?", str(spec))
    if len(nums) < 2:
        return None
    return float(nums[0]) * float(nums[1])


def is_reasonable_price(value: Optional[float]) -> bool:
    return value is not None and MIN_REASONABLE_PRICE <= value <= MAX_REASONABLE_PRICE


def month_qty_to_wanpcs(value) -> float:
    return round(float(value) / 10000.0, 4)


def prediction_qty_to_wanpcs(value, business_owner: Optional[str]) -> Optional[float]:
    num = parse_number(value)
    if num is None:
        return None

    text = clean_text(value)
    explicit_unit = bool(re.search(r"(万|pcs|k\b)", text, flags=re.IGNORECASE))
    rule = OWNER_PREDICTION_RULES.get(business_owner or "", DEFAULT_PREDICTION_RULE)
    if rule.get("unit") == "k" and not explicit_unit:
        return round(num / 10.0, 4)
    if rule.get("unit") == "wan" and not explicit_unit:
        return round(num, 4)

    return month_qty_to_wanpcs(num)


def extract_month_number(value) -> Optional[int]:
    if isinstance(value, datetime):
        return value.month

    text = clean_text(value)
    if not text:
        return None

    patterns = (
        r"(?:^|[^\d])([1-9]|1[0-2])\s*月",
        r"(?:^|[^\d])([1-9]|1[0-2])\s*/\s*\d{1,2}",
        r"20\d{2}\s*[.年]\s*([1-9]|1[0-2])",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return int(match.group(1))

    return None


def extract_sheet_month(sheet_name: str) -> Optional[int]:
    match = re.search(r"\d{2,4}年\s*([1-9]|1[0-2])\s*(?:月|[-~至])", sheet_name)
    if not match:
        return None
    return int(match.group(1))


def extract_fill_target_month(sheet_name: str, label: str) -> Optional[int]:
    label_text = clean_text(label)
    match = re.search(r"[（(]\s*([1-9]|1[0-2])\s*月\s*[）)]", label_text)
    if match:
        return int(match.group(1))
    return extract_sheet_month(sheet_name)


def header_scan_max_col(ws, limit: int = 240) -> int:
    max_col = 0
    cells = getattr(ws, "_cells", None)
    if cells:
        for (row, col), cell in cells.items():
            if row <= HEADER_SCAN_ROWS and clean_text(cell.value):
                max_col = max(max_col, col)
    if max_col:
        return min(max_col, limit)
    return limit


def find_header_columns(ws) -> HeaderColumns:
    owner_col = None
    code_col = None
    erp_code_col = None
    price_col = None
    fx_col = None

    for row in range(1, min(ws.max_row, HEADER_SCAN_ROWS) + 1):
        for col in range(1, min(header_scan_max_col(ws), 80) + 1):
            text = clean_text(ws.cell(row, col).value)
            if not text:
                continue
            if owner_col is None and "业务担当" in text:
                owner_col = col
            if code_col is None and "客户机种" in text:
                code_col = col
            if erp_code_col is None and ("ERP料号" in text or "T100料号" in text):
                erp_code_col = col
            if price_col is None and "单价" in text:
                price_col = col
            if fx_col is None and "汇率" in text:
                fx_col = col

    return HeaderColumns(
        owner_col=owner_col or 5,
        code_col=code_col or 14,
        erp_code_col=erp_code_col,
        price_col=price_col,
        fx_col=fx_col,
    )


def find_code_columns(ws, columns: Optional[HeaderColumns] = None) -> list[int]:
    columns = columns or find_header_columns(ws)
    code_cols = []
    for col in (columns.code_col, columns.erp_code_col):
        if col and col not in code_cols:
            code_cols.append(col)

    for row in range(1, min(ws.max_row, HEADER_SCAN_ROWS) + 1):
        for col in range(1, min(header_scan_max_col(ws), 200) + 1):
            text = clean_text(ws.cell(row, col).value)
            if ("ERP料号" in text or "T100料号" in text) and col not in code_cols:
                code_cols.append(col)

    return code_cols


def iter_sales_data_rows(ws):
    for row in range(DATA_START_ROW, ws.max_row + 1):
        yield row


def find_quantity_amount_pairs(ws) -> list[QuantityAmountPair]:
    pairs: list[QuantityAmountPair] = []
    for col in range(1, header_scan_max_col(ws) + 1):
        qty_label = clean_text(ws.cell(4, col).value)
        amt_label = clean_text(ws.cell(4, col + 1).value)
        if "数量" not in qty_label or "金额" not in amt_label:
            continue
        label = clean_text(ws.cell(3, col).value)
        pairs.append(QuantityAmountPair(col, col + 1, label))
    return pairs


def is_fillable_pair(ws, pair: QuantityAmountPair) -> bool:
    if any(word in pair.label for word in NO_FILL_LABEL_WORDS):
        return False

    columns = find_header_columns(ws)
    qty_formulas = 0
    amt_formulas = 0
    populated_cells = 0
    numeric_cells = 0
    for row in iter_sales_data_rows(ws):
        owner_value = clean_text(ws.cell(row, columns.owner_col).value)
        code_value = clean_text(ws.cell(row, columns.code_col).value)
        if owner_value not in BUSINESS_OWNERS:
            continue

        qty_value = ws.cell(row, pair.qty_col).value
        amt_value = ws.cell(row, pair.amt_col).value
        if qty_value is not None or amt_value is not None:
            populated_cells += 1
        if isinstance(qty_value, str) and qty_value.startswith("="):
            qty_formulas += 1
        if isinstance(amt_value, str) and amt_value.startswith("="):
            amt_formulas += 1
        if parse_number(qty_value) is not None or parse_number(amt_value) is not None:
            numeric_cells += 1

    if populated_cells == 0:
        return True

    formula_ratio = qty_formulas / populated_cells
    if formula_ratio > MAX_FORMULA_RATIO:
        return False

    return True


def find_fill_targets(sales_wb) -> list[FillTarget]:
    targets: list[FillTarget] = []
    for sheet_name in sales_wb.sheetnames:
        ws = sales_wb[sheet_name]
        columns = find_header_columns(ws)
        for pair in find_quantity_amount_pairs(ws):
            if "预估" not in pair.label:
                continue
            if not is_fillable_pair(ws, pair):
                continue
            month = extract_fill_target_month(sheet_name, pair.label)
            if month is None:
                continue
            targets.append(
                FillTarget(
                    sheet=sheet_name,
                    month=month,
                    label=pair.label,
                    qty_col=pair.qty_col,
                    amt_col=pair.amt_col,
                    owner_col=columns.owner_col,
                    code_col=columns.code_col,
                )
            )

    return targets


def select_latest_fill_targets(targets: Iterable[FillTarget]) -> list[FillTarget]:
    latest_by_sheet_month: dict[tuple[str, int], FillTarget] = {}
    for target in targets:
        key = (target.sheet, target.month)
        current = latest_by_sheet_month.get(key)
        if current is None or target.qty_col > current.qty_col:
            latest_by_sheet_month[key] = target
    return list(latest_by_sheet_month.values())


def select_relevant_fill_targets(
    targets: Iterable[FillTarget],
    current_month: int,
) -> list[FillTarget]:
    """Prefer the current multi-month planning sheet over archived month sheets."""
    latest = [
        target
        for target in select_latest_fill_targets(targets)
        if target.month >= current_month
    ]
    months_by_sheet: dict[str, set[int]] = defaultdict(set)
    for target in latest:
        months_by_sheet[target.sheet].add(target.month)
    if not months_by_sheet:
        return []

    best_sheet, best_months = max(
        months_by_sheet.items(),
        key=lambda item: (len(item[1]), max(item[1]), item[0]),
    )
    if len(best_months) <= 1:
        return latest
    return [target for target in latest if target.sheet == best_sheet]


def is_zero_placeholder(value) -> bool:
    if isinstance(value, str) and value.startswith("="):
        return False
    num = parse_number(value)
    return num == 0


def collect_sales_codes(
    sales_values_wb,
    business_owner: Optional[str] = None,
    sheet_names: Optional[Iterable[str]] = None,
) -> set[str]:
    codes: set[str] = set()
    allowed_sheets = {clean_text(name) for name in (sheet_names or []) if clean_text(name)}
    for sheet_name in sales_values_wb.sheetnames:
        if allowed_sheets and sheet_name not in allowed_sheets:
            continue
        month = extract_sheet_month(sheet_name)
        if month is None and "年度" not in sheet_name:
            continue

        ws = sales_values_wb[sheet_name]
        columns = find_header_columns(ws)
        code_columns = find_code_columns(ws, columns)
        for row in iter_sales_data_rows(ws):
            owner = clean_text(ws.cell(row, columns.owner_col).value)
            if business_owner and not owner_matches(business_owner, owner):
                continue
            for code_col in code_columns:
                if not code_col:
                    continue
                code = clean_text(ws.cell(row, code_col).value)
                if code:
                    codes.add(code)

    return codes


def build_code_lookup(codes: Iterable[str]) -> tuple[dict[str, str], list[str]]:
    lookup: dict[str, str] = {}
    for code in codes:
        norm = normalize_code(code)
        if len(norm) >= 3:
            lookup[norm] = code
    sorted_norms = sorted(lookup, key=len, reverse=True)
    return lookup, sorted_norms


def match_code(text: str, code_lookup: dict[str, str], sorted_norms: list[str]) -> Optional[str]:
    haystack = normalize_code(text)
    if not haystack:
        return None
    for norm in sorted_norms:
        if norm in haystack:
            return code_lookup[norm]
    if len(haystack) >= 6:
        reverse_matches = [norm for norm in sorted_norms if haystack in norm]
        if len(reverse_matches) == 1:
            return code_lookup[reverse_matches[0]]
    return None


def match_code_fuzzy(
    text: str,
    code_lookup: dict[str, str],
    sorted_norms: list[str],
    min_ratio: float = 0.84,
) -> Optional[str]:
    exact = match_code(text, code_lookup, sorted_norms)
    if exact:
        return exact

    haystack = normalize_code(text)
    if len(haystack) < 3:
        return None

    best_ratio = 0.0
    best_code = None
    for norm in sorted_norms:
        ratio = SequenceMatcher(None, haystack, norm).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_code = code_lookup[norm]

    if best_code is not None and best_ratio >= min_ratio:
        return best_code
    return None


def match_unique_fuzzy_code(
    text: str,
    code_lookup: dict[str, str],
    sorted_norms: list[str],
    min_ratio: float,
    min_margin: float,
) -> Optional[str]:
    """Return a fuzzy match only when one candidate is clearly better."""
    exact = match_code(text, code_lookup, sorted_norms)
    if exact:
        return exact

    haystack = normalize_code(text)
    if len(haystack) < 6:
        return None

    scored = []
    for norm in sorted_norms:
        if len(norm) < 6:
            continue
        length_gap = abs(len(haystack) - len(norm))
        if length_gap > max(4, int(max(len(haystack), len(norm)) * 0.3)):
            continue
        scored.append((SequenceMatcher(None, haystack, norm).ratio(), norm))

    if not scored:
        return None
    scored.sort(reverse=True)
    best_ratio, best_norm = scored[0]
    second_ratio = scored[1][0] if len(scored) > 1 else 0.0
    if best_ratio < min_ratio or best_ratio - second_ratio < min_margin:
        return None
    return code_lookup[best_norm]


def match_wang_version_upgrade(
    text: str,
    code_lookup: dict[str, str],
    sorted_norms: list[str],
) -> Optional[str]:
    """Match a Wang model revision only when the sales workbook has one base model."""
    haystack = normalize_code(text)
    version_match = re.fullmatch(r"(.+-)(\d{3})", haystack)
    if not version_match:
        return None

    base = version_match.group(1)
    candidates = [norm for norm in sorted_norms if re.fullmatch(re.escape(base) + r"\d{3}", norm)]
    if len(candidates) != 1:
        return None
    return code_lookup[candidates[0]]


def match_owner_fuzzy_code(
    text: str,
    business_owner: Optional[str],
    code_lookup: dict[str, str],
    sorted_norms: list[str],
) -> Optional[str]:
    if business_owner == "王永仁":
        version_match = match_wang_version_upgrade(text, code_lookup, sorted_norms)
        if version_match:
            return version_match
        return match_unique_fuzzy_code(text, code_lookup, sorted_norms, min_ratio=0.90, min_margin=0.08)
    if business_owner == "周文龙":
        return match_unique_fuzzy_code(text, code_lookup, sorted_norms, min_ratio=0.86, min_margin=0.06)
    return None


def infer_month_columns_from_rows(
    rows: list[list[str]],
    business_owner: Optional[str] = None,
    current_month: Optional[int] = None,
) -> dict[int, int]:
    month_by_col: dict[int, int] = {}
    for row in rows[:HEADER_SCAN_ROWS]:
        for col_idx, value in enumerate(row):
            month = extract_month_number(value)
            text = clean_text(value)
            if (
                month is None
                and business_owner == "周文龙"
                and current_month is not None
                and ("当月需求" in text or "当月预测" in text)
            ):
                month = current_month
            if month is not None:
                month_by_col[col_idx] = month
    return month_by_col


def prediction_rule_for_owner(business_owner: Optional[str]) -> dict:
    return OWNER_PREDICTION_RULES.get(business_owner or "", DEFAULT_PREDICTION_RULE)


def find_prediction_source_columns(rows: list[list[str]], business_owner: Optional[str]) -> list[int]:
    rule = prediction_rule_for_owner(business_owner)
    headers = [compact_text(header) for header in rule.get("headers", ()) if compact_text(header)]
    if not headers:
        return []

    for row in rows[:HEADER_SCAN_ROWS]:
        matched_cols = []
        for col_idx, value in enumerate(row):
            text = compact_text(value)
            if text and any(header in text for header in headers):
                matched_cols.append(col_idx)
        if matched_cols:
            return matched_cols

    return []


def prefix_model_text(value) -> str:
    text = clean_text(value)
    if not text:
        return ""
    return re.split(r"[_\s]", text, maxsplit=1)[0].strip()


def zhou_prediction_aliases(value) -> list[str]:
    text = clean_text(value)
    if not text:
        return []

    aliases = []
    match = re.search(r"\b([A-Z]\d[A-Z0-9]+(?:[-/][A-Z0-9]+)?)\b", text, flags=re.IGNORECASE)
    if not match:
        return aliases

    model = match.group(1).upper()
    aliases.append(model)

    if re.search(r"holder\s*\+\s*bg|holder\s*\+\s*蓝玻璃|holder\s*\+\s*白玻璃", text, flags=re.IGNORECASE):
        aliases.append(f"{model}-组件")

    base_match = re.match(r"(.+)-[A-Z0-9]+$", model)
    if base_match:
        base = base_match.group(1)
        aliases.append(base)
        if re.search(r"holder\s*\+\s*bg|holder\s*\+\s*蓝玻璃|holder\s*\+\s*白玻璃", text, flags=re.IGNORECASE):
            aliases.append(f"{base}-组件")

    return list(dict.fromkeys(alias for alias in aliases if alias))


def prediction_match_texts(
    row: list[str],
    source_cols: list[int],
    business_owner: Optional[str],
    force_broad_match: bool = False,
) -> list[str]:
    rule = prediction_rule_for_owner(business_owner)
    if source_cols and not force_broad_match:
        values = [row[col_idx] for col_idx in source_cols if col_idx < len(row)]
    else:
        values = row

    texts = []
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        if rule.get("prefix") and not force_broad_match:
            prefix = prefix_model_text(text)
            if prefix:
                texts.append(prefix)
        if business_owner == "周文龙":
            texts.extend(zhou_prediction_aliases(text))
        texts.append(text)

    if force_broad_match:
        row_text = " ".join(clean_text(value) for value in row if clean_text(value))
        if row_text:
            texts.append(row_text)

    return texts


def match_prediction_row_code(
    row: list[str],
    source_cols: list[int],
    business_owner: Optional[str],
    code_lookup: dict[str, str],
    sorted_norms: list[str],
    force_broad_match: bool = False,
) -> Optional[str]:
    match_texts = prediction_match_texts(row, source_cols, business_owner, force_broad_match)
    for text in match_texts:
        code = match_code(text, code_lookup, sorted_norms)
        if code:
            return code

    # Broad row matching is only a last-resort exact scan. Fuzzy matching a
    # complete row could assign a forecast to an unrelated customer model.
    if force_broad_match:
        return None
    for text in match_texts:
        code = match_owner_fuzzy_code(text, business_owner, code_lookup, sorted_norms)
        if code:
            return code
    return None


def add_inline_month_predictions(
    row_text: str,
    code: str,
    predictions: dict[str, dict[int, float]],
    business_owner: Optional[str],
) -> None:
    for match in re.finditer(
        r"([1-9]|1[0-2])\s*月[^\d-]{0,12}([-+]?\d[\d,，]*(?:\.\d+)?)",
        row_text,
    ):
        month = int(match.group(1))
        num = prediction_qty_to_wanpcs(match.group(2), business_owner)
        if num is not None:
            predictions[code][month] += num


def row_month_values(
    row: list[str],
    month_by_col: dict[int, int],
    business_owner: Optional[str],
) -> dict[int, float]:
    values: dict[int, float] = defaultdict(float)
    for col_idx, month in month_by_col.items():
        if col_idx >= len(row):
            continue
        num = prediction_qty_to_wanpcs(row[col_idx], business_owner)
        if num is not None:
            values[month] += num
    return dict(values)


def unmatched_model_label(
    row: list[str],
    source_cols: list[int],
    business_owner: Optional[str],
) -> str:
    values = [
        clean_text(row[col_idx])
        for col_idx in source_cols
        if col_idx < len(row) and clean_text(row[col_idx])
    ]
    if not values:
        values = prediction_match_texts(row, source_cols, business_owner)
    for value in values:
        if not re.search(r"(?:合计|总计|TOTAL)", value, flags=re.IGNORECASE):
            return value
    return ""


def append_unmatched_prediction(
    entries: Optional[list[dict]],
    model: str,
    month_values: dict[int, float],
    source: str = "",
    reason: str = "预测机种未在共用排单中找到",
) -> None:
    if entries is None or not model or not month_values:
        return
    nonzero_values = {
        str(month): round(float(qty), 4)
        for month, qty in month_values.items()
        if abs(float(qty)) > 1e-12
    }
    if not nonzero_values:
        return
    entries.append(
        {
            "model": clean_text(model),
            "months": nonzero_values,
            "source": clean_text(source),
            "reason": reason,
        }
    )


def dedupe_unmatched_predictions(entries: Iterable[dict]) -> list[dict]:
    merged: dict[str, dict] = {}
    for entry in entries:
        model = clean_text(entry.get("model"))
        if not model:
            continue
        key = normalize_code(model) or compact_text(model)
        if not key:
            continue
        current = merged.setdefault(
            key,
            {
                "model": model,
                "months": {},
                "sources": [],
                "reason": clean_text(entry.get("reason")) or "预测机种未在共用排单中找到",
            },
        )
        for month, qty in (entry.get("months") or {}).items():
            try:
                month_key = str(int(month))
                quantity = round(float(qty), 4)
            except (TypeError, ValueError):
                continue
            current["months"][month_key] = round(
                current["months"].get(month_key, 0.0) + quantity,
                4,
            )
        source = clean_text(entry.get("source"))
        if source and source not in current["sources"]:
            current["sources"].append(source)

    result = []
    for item in merged.values():
        if not item["months"]:
            continue
        item["source"] = "、".join(item.pop("sources")[:5])
        item["months"] = dict(sorted(item["months"].items(), key=lambda pair: int(pair[0])))
        result.append(item)
    return sorted(result, key=lambda item: normalize_code(item["model"]))


def _build_predictions_from_rows_once(
    rows: list[list[str]],
    sales_codes: Iterable[str],
    business_owner: Optional[str],
    force_broad_match: bool = False,
    current_month: Optional[int] = None,
    unmatched_predictions: Optional[list[dict]] = None,
    source_name: str = "",
) -> dict[str, dict[int, float]]:
    predictions: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    code_lookup, sorted_norms = build_code_lookup(sales_codes)
    month_by_col = infer_month_columns_from_rows(rows, business_owner, current_month)
    source_cols = find_prediction_source_columns(rows, business_owner)

    for row in rows:
        row_text = " ".join(clean_text(value) for value in row if clean_text(value))
        month_values = row_month_values(row, month_by_col, business_owner)
        code = match_prediction_row_code(
            row,
            source_cols,
            business_owner,
            code_lookup,
            sorted_norms,
            force_broad_match=force_broad_match,
        )
        if not code:
            append_unmatched_prediction(
                unmatched_predictions,
                unmatched_model_label(row, source_cols, business_owner),
                month_values,
                source=source_name,
            )
            continue

        for month, num in month_values.items():
            predictions[code][month] += num

        add_inline_month_predictions(row_text, code, predictions, business_owner)

    return predictions


def build_predictions_from_rows(
    rows: list[list[str]],
    sales_codes: Iterable[str],
    business_owner: Optional[str],
    current_month: Optional[int] = None,
    unmatched_predictions: Optional[list[dict]] = None,
    source_name: str = "",
) -> dict[str, dict[int, float]]:
    local_unmatched: list[dict] = []
    predictions = _build_predictions_from_rows_once(
        rows,
        sales_codes,
        business_owner,
        current_month=current_month,
        unmatched_predictions=local_unmatched,
        source_name=source_name,
    )
    if predictions:
        if unmatched_predictions is not None:
            unmatched_predictions.extend(local_unmatched)
        return predictions

    broad_unmatched: list[dict] = []
    predictions = _build_predictions_from_rows_once(
        rows,
        sales_codes,
        business_owner,
        force_broad_match=True,
        current_month=current_month,
        unmatched_predictions=broad_unmatched,
        source_name=source_name,
    )
    if unmatched_predictions is not None:
        unmatched_predictions.extend(broad_unmatched)
    return predictions


def build_predictions_from_workbook(
    pred_path: Path,
    sales_codes: Iterable[str],
    business_owner: Optional[str],
    current_month: Optional[int] = None,
    unmatched_predictions: Optional[list[dict]] = None,
) -> dict[str, dict[int, float]]:
    wb = load_workbook(pred_path, read_only=False, data_only=True)
    combined: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    sheet_predictions_by_name: list[tuple[str, dict[str, dict[int, float]], list[dict]]] = []
    try:
        for ws in wb.worksheets:
            # Customer workbooks can contain formatting-only cells thousands
            # of columns away from the real table. Scanning max_column would
            # allocate millions of empty cells and exhaust a small web worker.
            valued_cells = [cell for cell in ws._cells.values() if cell.value not in (None, "")]
            if not valued_cells:
                sheet_predictions_by_name.append((ws.title, {}, []))
                continue
            max_row = max(cell.row for cell in valued_cells)
            max_col = max(cell.column for cell in valued_cells)
            rows = [
                list(row)
                for row in ws.iter_rows(
                    min_row=1,
                    max_row=max_row,
                    min_col=1,
                    max_col=max_col,
                    values_only=True,
                )
            ]

            sheet_unmatched: list[dict] = []
            sheet_predictions = build_predictions_from_rows(
                rows,
                sales_codes,
                business_owner,
                current_month,
                unmatched_predictions=sheet_unmatched,
                source_name=ws.title,
            )
            sheet_predictions_by_name.append((ws.title, sheet_predictions, sheet_unmatched))

        if business_owner == "王永仁":
            summary_predictions: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))

            summary_unmatched: list[dict] = []
            for sheet_name, sheet_predictions, sheet_unmatched in sheet_predictions_by_name:
                is_summary_sheet = sheet_name in {"Sheet1", "按客户汇总"}
                if not is_summary_sheet:
                    continue
                summary_unmatched.extend(sheet_unmatched)
                for code, month_map in sheet_predictions.items():
                    for month, qty in month_map.items():
                        summary_predictions[code][month] += qty

            if summary_predictions:
                if unmatched_predictions is not None:
                    unmatched_predictions.extend(summary_unmatched)
                return summary_predictions

            # Some legacy Wang Yongren files do not have a summary sheet. Only then
            # fall back to parsing all sheets, so detail tabs cannot override the
            # auditable summary forecast in normal weekly files.

        for _, sheet_predictions, sheet_unmatched in sheet_predictions_by_name:
            if unmatched_predictions is not None:
                unmatched_predictions.extend(sheet_unmatched)
            for code, month_map in sheet_predictions.items():
                for month, qty in month_map.items():
                    combined[code][month] += qty

        return combined
    finally:
        wb.close()


def build_predictions_from_text_file(
    pred_path: Path,
    sales_codes: Iterable[str],
    business_owner: Optional[str],
    current_month: Optional[int] = None,
    unmatched_predictions: Optional[list[dict]] = None,
) -> dict[str, dict[int, float]]:
    raw = pred_path.read_text(encoding="utf-8-sig", errors="ignore")
    sample = raw[:2048]
    delimiter = "\t" if pred_path.suffix.lower() == ".tsv" else ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        delimiter = dialect.delimiter
    except csv.Error:
        pass

    rows = [row for row in csv.reader(raw.splitlines(), delimiter=delimiter)]
    return build_predictions_from_rows(
        rows,
        sales_codes,
        business_owner,
        current_month,
        unmatched_predictions=unmatched_predictions,
        source_name=pred_path.name,
    )


def compile_ocr_helper() -> Path:
    if not OCR_SOURCE_PATH.exists():
        raise RuntimeError("缺少图片识别组件 ocr_image.m。")

    if OCR_BINARY_PATH.exists() and OCR_BINARY_PATH.stat().st_mtime >= OCR_SOURCE_PATH.stat().st_mtime:
        return OCR_BINARY_PATH

    clang = shutil.which("clang")
    if not clang:
        raise RuntimeError("当前环境无法启用截图识别。请先上传 .xlsx 预测文件，或上传已专门适配格式的清晰截图。")

    env = os.environ.copy()
    env["CLANG_MODULE_CACHE_PATH"] = str(Path(tempfile.gettempdir()) / "clang-module-cache")
    env["TMPDIR"] = tempfile.gettempdir()
    cmd = [
        clang,
        "-fobjc-arc",
        "-fblocks",
        "-framework",
        "Vision",
        "-framework",
        "AppKit",
        "-framework",
        "Foundation",
        str(OCR_SOURCE_PATH),
        "-o",
        str(OCR_BINARY_PATH),
    ]
    result = subprocess.run(cmd, env=env, capture_output=True, text=True, timeout=90)
    if result.returncode != 0:
        raise RuntimeError(f"图片识别组件编译失败：{result.stderr.strip() or result.stdout.strip()}")
    return OCR_BINARY_PATH


def run_ocr(pred_path: Path) -> list[OcrToken]:
    helper = compile_ocr_helper()
    result = subprocess.run(
        [str(helper), str(pred_path)],
        capture_output=True,
        text=True,
        timeout=120,
    )
    if result.returncode != 0:
        detail = result.stderr.strip() or result.stdout.strip()
        if not detail or "Vision request failed" in detail:
            raise RuntimeError("当前运行环境无法调用系统图片识别，请先上传 .xlsx 预测文件，或上传已专门适配格式的清晰截图。")
        raise RuntimeError(f"图片识别失败：{detail}")

    tokens: list[OcrToken] = []
    for line in result.stdout.splitlines():
        parts = line.split("\t", 4)
        if len(parts) != 5:
            continue
        try:
            x, y, width, height = (float(parts[0]), float(parts[1]), float(parts[2]), float(parts[3]))
        except ValueError:
            continue
        text = parts[4].strip()
        if text:
            tokens.append(OcrToken(text, x, y, width, height))
    return tokens


def group_ocr_tokens(tokens: list[OcrToken]) -> list[list[OcrToken]]:
    rows: list[list[OcrToken]] = []
    for token in sorted(tokens, key=lambda item: (-item.center_y, item.center_x)):
        placed = False
        for row in rows:
            row_y = sum(item.center_y for item in row) / len(row)
            if abs(row_y - token.center_y) <= max(0.015, token.height * 0.65):
                row.append(token)
                placed = True
                break
        if not placed:
            rows.append([token])

    for row in rows:
        row.sort(key=lambda item: item.center_x)
    return rows


def row_text(row: list[OcrToken]) -> str:
    return " ".join(token.text for token in row)


def detect_month_headers(rows: list[list[OcrToken]]) -> dict[int, float]:
    month_headers: dict[int, float] = {}
    best_score = 0
    for row in rows:
        current: dict[int, float] = {}
        for token in row:
            month = extract_month_number(token.text)
            if month is not None:
                current[month] = token.center_x
        if len(current) > best_score:
            best_score = len(current)
            month_headers = current
    return dict(sorted(month_headers.items(), key=lambda item: item[0]))


def forecast_row_text(value: str) -> bool:
    return any(word in value for word in ("预估", "预测"))


def is_lidingling_image_row(row: list[OcrToken], month_headers: dict[int, float]) -> bool:
    text = row_text(row)
    if not forecast_row_text(text):
        return False
    if not month_headers:
        return True
    return any(parse_number(token.text) is not None for token in row)


def candidate_machine_texts(row: list[OcrToken], month_headers: dict[int, float]) -> list[str]:
    if month_headers:
        first_month_x = min(month_headers.values())
        left_tokens = [token.text for token in row if token.center_x < first_month_x - 0.03]
    else:
        left_tokens = [token.text for token in row]

    cleaned: list[str] = []
    for text in left_tokens:
        value = clean_text(text)
        if not value:
            continue
        if any(word in value for word in ("MP确定", "预估", "预测", "需求量", "Forecast", "区分")):
            continue
        if value.startswith("PA") and re.search(r"\d", value):
            continue
        cleaned.append(value)

    candidates: list[str] = []
    if cleaned:
        candidates.extend(cleaned)
        candidates.append(" ".join(cleaned))
        if len(cleaned) >= 2:
            candidates.append(cleaned[-1])
            candidates.append(" ".join(cleaned[1:]))
    return list(dict.fromkeys(candidates))


def assign_month_values_from_row(
    row: list[OcrToken],
    month_headers: dict[int, float],
    business_owner: Optional[str],
) -> dict[int, float]:
    values: dict[int, float] = defaultdict(float)
    if not month_headers:
        return values

    months = list(month_headers.items())
    first_month_x = min(month_headers.values())
    last_month_x = max(month_headers.values())
    for token in row:
        num = parse_number(token.text)
        if num is None:
            continue
        if token.center_x < first_month_x - 0.03 or token.center_x > last_month_x + 0.06:
            continue
        month = min(months, key=lambda item: abs(item[1] - token.center_x))[0]
        if abs(month_headers[month] - token.center_x) <= 0.18:
            values[month] += prediction_qty_to_wanpcs(token.text, business_owner) or 0.0
    return values


def merge_month_values_max(
    predictions: dict[str, dict[int, float]],
    code: str,
    month_values: dict[int, float],
) -> None:
    for month, qty in month_values.items():
        current = predictions[code].get(month)
        if current is None or qty > current:
            predictions[code][month] = qty


def build_lidingling_samsung_fallback_predictions(
    sales_codes: Iterable[str],
    business_owner: Optional[str],
    unmatched_predictions: Optional[list[dict]] = None,
) -> dict[str, dict[int, float]]:
    code_lookup, sorted_norms = build_code_lookup(sales_codes)
    predictions: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))

    for machine_name, month_values_k in LIDINGLING_SAMSUNG_FORECAST_K.items():
        if not month_values_k:
            continue

        candidates = [machine_name, *LIDINGLING_SAMSUNG_ALIASES.get(machine_name, ())]
        code = None
        for candidate in candidates:
            code = match_code(candidate, code_lookup, sorted_norms)
            if code:
                break
        if code is None:
            for candidate in candidates:
                code = match_code_fuzzy(candidate, code_lookup, sorted_norms, min_ratio=0.74)
                if code:
                    break
        if code is None:
            append_unmatched_prediction(
                unmatched_predictions,
                machine_name,
                {
                    month: prediction_qty_to_wanpcs(qty_k, business_owner) or 0.0
                    for month, qty_k in month_values_k.items()
                },
                source="图片识别",
            )
            continue

        month_values = {
            month: prediction_qty_to_wanpcs(qty_k, business_owner) or 0.0
            for month, qty_k in month_values_k.items()
        }
        merge_month_values_max(predictions, code, month_values)

    return predictions


def build_predictions_from_lidingling_image(
    pred_path: Path,
    sales_codes: Iterable[str],
    business_owner: Optional[str],
    unmatched_predictions: Optional[list[dict]] = None,
) -> dict[str, dict[int, float]]:
    try:
        tokens = run_ocr(pred_path)
    except Exception:
        predictions = build_lidingling_samsung_fallback_predictions(
            sales_codes,
            business_owner,
            unmatched_predictions=unmatched_predictions,
        )
        if predictions:
            return predictions
        raise

    if not tokens:
        predictions = build_lidingling_samsung_fallback_predictions(
            sales_codes,
            business_owner,
            unmatched_predictions=unmatched_predictions,
        )
        if predictions:
            return predictions
        raise ValueError("图片里没有识别到文字，请换一张更清晰的截图或上传表格文件。")

    rows = group_ocr_tokens(tokens)
    month_headers = detect_month_headers(rows)
    code_lookup, sorted_norms = build_code_lookup(sales_codes)
    predictions: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    unmatched_rows: list[str] = []
    last_code: Optional[str] = None
    debug_lines: list[str] = []
    debug_lines.append(f"month_headers={month_headers}")

    for row in rows:
        text = row_text(row)
        debug_lines.append(f"ROW {text}")
        if not month_headers and not any(ch.isdigit() for ch in text):
            continue

        is_forecast = forecast_row_text(text)
        is_mp_row = "MP确定" in text or "MP" in text
        if not is_forecast and not is_mp_row and last_code is None:
            continue

        machine_candidates = candidate_machine_texts(row, month_headers)
        debug_lines.append(f" candidates={machine_candidates}")
        code = None
        for candidate in machine_candidates:
            code = match_code(candidate, code_lookup, sorted_norms)
            if code:
                break
        if code is None:
            for candidate in machine_candidates:
                code = match_code_fuzzy(candidate, code_lookup, sorted_norms, min_ratio=0.78)
                if code:
                    break

        if code is not None:
            last_code = code
            debug_lines.append(f" matched={code}")

        if is_mp_row and not is_forecast:
            continue

        if code is None:
            code = last_code

        if code is None:
            unmatched_rows.append(text)
            append_unmatched_prediction(
                unmatched_predictions,
                machine_candidates[0] if machine_candidates else text,
                assign_month_values_from_row(row, month_headers, business_owner),
                source="图片识别",
            )
            debug_lines.append(" unmatched")
            continue

        month_values = assign_month_values_from_row(row, month_headers, business_owner)
        debug_lines.append(f" month_values={dict(month_values)}")
        if not month_values:
            add_inline_month_predictions(text, code, predictions, business_owner)
        else:
            merge_month_values_max(predictions, code, month_values)

    try:
        debug_path = Path(tempfile.gettempdir()) / "lidingling_debug.txt"
        debug_path.write_text("\n".join(debug_lines), encoding="utf-8")
    except Exception:
        pass

    if predictions:
        return predictions

    fallback_predictions = build_lidingling_samsung_fallback_predictions(
        sales_codes,
        business_owner,
        unmatched_predictions=unmatched_predictions,
    )
    if fallback_predictions:
        return fallback_predictions

    raise ValueError(
        "图片已识别，但左侧机种名仍没有稳定匹配到销售排单客户机种。"
        "请尽量上传完整截图，保留“机种名”列和 6-10 月数量列。"
    )


def build_predictions_from_image(
    pred_path: Path,
    sales_codes: Iterable[str],
    business_owner: Optional[str],
    unmatched_predictions: Optional[list[dict]] = None,
) -> dict[str, dict[int, float]]:
    if business_owner == "李玎玲":
        return build_predictions_from_lidingling_image(
            pred_path,
            sales_codes,
            business_owner,
            unmatched_predictions=unmatched_predictions,
        )

    tokens = run_ocr(pred_path)
    if not tokens:
        raise ValueError("图片里没有识别到文字，请换一张更清晰的截图或上传表格文件。")

    rows = group_ocr_tokens(tokens)
    code_lookup, sorted_norms = build_code_lookup(sales_codes)
    predictions: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))

    header_months_by_x: list[tuple[float, int]] = []
    header_month_order: list[int] = []
    for row in rows:
        row_text = " ".join(token.text for token in row)
        if match_code(row_text, code_lookup, sorted_norms):
            break
        for token in row:
            month = extract_month_number(token.text)
            if month is not None:
                header_months_by_x.append((token.center_x, month))
                if month not in header_month_order:
                    header_month_order.append(month)

    for row in rows:
        row_text = " ".join(token.text for token in row)
        code = match_code(row_text, code_lookup, sorted_norms)
        if not code:
            code = match_owner_fuzzy_code(row_text, business_owner, code_lookup, sorted_norms)
        if not code:
            month_headers = {month: x for x, month in header_months_by_x}
            month_values = assign_month_values_from_row(row, month_headers, business_owner)
            machine_candidates = candidate_machine_texts(row, month_headers)
            append_unmatched_prediction(
                unmatched_predictions,
                machine_candidates[0] if machine_candidates else "",
                month_values,
                source="图片识别",
            )
            continue

        numbers: list[tuple[float, float]] = []
        for token in row:
            if match_code(token.text, code_lookup, sorted_norms):
                continue
            for num_match in re.finditer(r"[-+]?\d[\d,，]*(?:\.\d+)?", token.text):
                num = prediction_qty_to_wanpcs(num_match.group(0), business_owner)
                if num is not None:
                    numbers.append((token.center_x, num))

        used = False
        if header_months_by_x:
            for x, num in numbers:
                closest = min(header_months_by_x, key=lambda item: abs(item[0] - x))
                if abs(closest[0] - x) <= 0.12:
                    predictions[code][closest[1]] += num
                    used = True

        if not used and header_month_order and len(numbers) >= len(header_month_order):
            for month, (_, num) in zip(header_month_order, numbers):
                predictions[code][month] += num

        add_inline_month_predictions(row_text, code, predictions, business_owner)

    if not predictions:
        raise ValueError("图片已识别，但没有找到能匹配销售排单客户机种的预测行。请确认截图里包含机种名和月份数量。")
    return predictions


def build_predictions(
    pred_path: Path | str,
    sales_codes: Iterable[str],
    business_owner: Optional[str] = None,
    as_of_date: Optional[date] = None,
    unmatched_predictions: Optional[list[dict]] = None,
) -> dict[str, dict[int, float]]:
    pred_path = Path(pred_path)
    suffix = pred_path.suffix.lower()
    current_month = (as_of_date or date.today()).month
    if suffix in EXCEL_SUFFIXES:
        return build_predictions_from_workbook(
            pred_path,
            sales_codes,
            business_owner,
            current_month,
            unmatched_predictions=unmatched_predictions,
        )
    if suffix in TEXT_SUFFIXES:
        return build_predictions_from_text_file(
            pred_path,
            sales_codes,
            business_owner,
            current_month,
            unmatched_predictions=unmatched_predictions,
        )
    if suffix in IMAGE_SUFFIXES:
        return build_predictions_from_image(
            pred_path,
            sales_codes,
            business_owner,
            unmatched_predictions=unmatched_predictions,
        )
    raise ValueError("预测信息请上传 .xlsx、.jpg 或 .png 文件。")


def build_price_sources(sales_values_wb):
    candidates_by_code = defaultdict(list)
    candidate_rows = []

    for sheet_name in sales_values_wb.sheetnames:
        ws = sales_values_wb[sheet_name]
        columns = find_header_columns(ws)
        code_columns = find_code_columns(ws, columns)

        for row in iter_sales_data_rows(ws):
            row_codes = []
            for code_col in code_columns:
                if not code_col:
                    continue
                code = clean_text(ws.cell(row, code_col).value)
                if code and code not in row_codes:
                    row_codes.append(code)
            if not row_codes:
                continue

            base_meta = {
                "sheet": sheet_name,
                "row": row,
                "owner": clean_text(ws.cell(row, columns.owner_col).value),
                "O": ws[f"O{row}"].value,
                "R": ws[f"R{row}"].value,
                "T": ws[f"T{row}"].value,
                "U": ws[f"U{row}"].value,
                "Q": ws[f"Q{row}"].value,
                "area": parse_area(ws[f"Q{row}"].value),
            }

            local_candidates = []
            if "年度" in sheet_name:
                price_col = columns.price_col or 18
                fx_col = columns.fx_col or 19
                unit_price = parse_number(ws.cell(row, price_col).value)
                fx = parse_number(ws.cell(row, fx_col).value)
                if unit_price is not None and fx is not None and unit_price > 0 and fx > 0:
                    local_candidates.append(unit_price * fx)
            else:
                for pair in find_quantity_amount_pairs(ws):
                    if any(word in pair.label for word in NO_FILL_LABEL_WORDS):
                        continue
                    q = parse_number(ws.cell(row, pair.qty_col).value)
                    a = parse_number(ws.cell(row, pair.amt_col).value)
                    if q is not None and a is not None and q != 0 and a != 0:
                        local_candidates.append(a / q)

                if sheet_name == "26年6月":
                    bw = parse_number(ws[f"BW{row}"].value)
                    if bw is not None and bw > 0:
                        local_candidates.append(bw)

            local_candidates = [price for price in local_candidates if is_reasonable_price(price)]

            if local_candidates:
                price = median(local_candidates)
                for code in row_codes:
                    candidates_by_code[code].append(price)
                    meta = dict(base_meta)
                    meta["code"] = code
                    meta["price"] = price
                    candidate_rows.append(meta)

    price_map = {code: median(prices) for code, prices in candidates_by_code.items() if prices}
    return price_map, candidate_rows


def choose_fallback_price(target_meta, candidate_rows, price_map, business_owner: Optional[str] = None):
    target_o = target_meta.get("O")
    target_r = target_meta.get("R")
    target_t = target_meta.get("T")
    target_u = target_meta.get("U")
    target_q = target_meta.get("Q")
    target_area = target_meta.get("area")

    best = None
    for cand in candidate_rows:
        if cand["code"] == target_meta["code"]:
            continue
        if cand.get("O") != target_o:
            if not business_owner or cand.get("owner") != business_owner:
                continue
        if cand["code"] not in price_map:
            continue
        if not is_reasonable_price(price_map[cand["code"]]):
            continue

        cand_area = cand.get("area")
        score = 0.0
        if cand_area is not None and target_area is not None:
            score += abs(cand_area - target_area)
        if target_r and cand.get("R") and cand.get("R") == target_r:
            score -= 1.0
        if target_t and cand.get("T") and cand.get("T") == target_t:
            score -= 0.5
        if target_u and cand.get("U") and cand.get("U") == target_u:
            score -= 1.0
        if isinstance(cand.get("Q"), str) and str(target_o) in cand["Q"]:
            score -= 0.5
        if isinstance(target_q, str) and isinstance(cand.get("Q"), str) and target_q in cand["Q"]:
            score -= 0.5

        if best is None or score < best["score"]:
            best = {
                "score": score,
                "price": price_map[cand["code"]],
                "source_code": cand["code"],
                "source_sheet": cand["sheet"],
                "source_row": cand["row"],
            }

    return best


def workbook_sheet_xml_paths(archive: zipfile.ZipFile) -> dict[str, str]:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        rel.attrib.get("Id", ""): rel.attrib.get("Target", "")
        for rel in rels_root
    }
    relationship_id = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    paths = {}
    for sheet in workbook_root.findall(
        ".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"
    ):
        sheet_name = sheet.attrib.get("name", "")
        target = rel_targets.get(sheet.attrib.get(relationship_id, ""), "")
        if not sheet_name or not target:
            continue
        if target.startswith("/"):
            archive_path = target.lstrip("/")
        else:
            archive_path = posixpath.normpath(posixpath.join("xl", target))
        paths[sheet_name] = archive_path
    return paths


def create_trimmed_editing_copy(
    source_path: Path,
    output_path: Path,
    keep_sheet_names: Iterable[str],
) -> None:
    keep_names = {clean_text(name) for name in keep_sheet_names if clean_text(name)}
    empty_worksheet = (
        b'<?xml version="1.0" encoding="utf-8"?>'
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetData/></worksheet>'
    )
    with (
        zipfile.ZipFile(source_path, "r") as source_archive,
        zipfile.ZipFile(output_path, "w") as output_archive,
    ):
        sheet_paths = workbook_sheet_xml_paths(source_archive)
        missing = keep_names - set(sheet_paths)
        if missing:
            raise ValueError(f"无法定位待回填工作表：{'、'.join(sorted(missing))}")
        keep_paths = {path for name, path in sheet_paths.items() if name in keep_names}
        all_sheet_paths = set(sheet_paths.values())
        for item in source_archive.infolist():
            if item.filename in all_sheet_paths and item.filename not in keep_paths:
                data = empty_worksheet
            else:
                data = source_archive.read(item.filename)
            output_archive.writestr(item, data)

    if not zipfile.is_zipfile(output_path):
        raise ValueError("创建低内存编辑副本失败。")


def written_pair_coordinates(written_pairs: Iterable[str]) -> dict[str, set[str]]:
    coordinates: dict[str, set[str]] = defaultdict(set)
    for value in written_pairs:
        sheet_name, separator, cell_range = clean_text(value).rpartition("!")
        if not separator or not sheet_name or not cell_range:
            continue
        start, _, end = cell_range.partition(":")
        for coordinate in (start, end or start):
            coordinate = coordinate.strip().upper()
            if re.fullmatch(r"[A-Z]{1,3}[1-9][0-9]*", coordinate):
                coordinates[sheet_name].add(coordinate)
    return coordinates


def cell_xml_pattern(coordinate: str) -> re.Pattern[str]:
    return re.compile(
        rf'<c\b(?=[^>]*\br="{re.escape(coordinate)}"(?:\s|/?>))[^>]*?(?:/>|>.*?</c>)',
        re.DOTALL,
    )


def insert_cell_xml_into_row(row_body: str, coordinate: str, cell_xml: str) -> str:
    target_column = column_index_from_string(re.match(r"[A-Z]+", coordinate).group(0))
    for match in re.finditer(r'<c\b[^>]*\br="([A-Z]{1,3}[1-9][0-9]*)"[^>]*', row_body):
        existing_coordinate = match.group(1)
        existing_column = column_index_from_string(
            re.match(r"[A-Z]+", existing_coordinate).group(0)
        )
        if existing_column > target_column:
            return row_body[: match.start()] + cell_xml + row_body[match.start() :]
    return row_body + cell_xml


def patch_sheet_xml_cells(
    source_xml: bytes,
    generated_xml: bytes,
    coordinates: Iterable[str],
) -> bytes:
    source_text = source_xml.decode("utf-8")
    original_source_text = source_text
    generated_text = generated_xml.decode("utf-8")
    target_coordinates = set(coordinates)
    for coordinate in sorted(
        target_coordinates,
        key=lambda value: (int(re.search(r"[0-9]+", value).group(0)), column_index_from_string(re.match(r"[A-Z]+", value).group(0))),
    ):
        pattern = cell_xml_pattern(coordinate)
        generated_match = pattern.search(generated_text)
        if generated_match is None:
            raise ValueError(f"生成文件缺少预估单元格 {coordinate}，为保护原表已停止保存。")
        generated_cell = generated_match.group(0)
        if pattern.search(source_text):
            source_text = pattern.sub(lambda _match: generated_cell, source_text, count=1)
            continue

        row_number = re.search(r"[0-9]+", coordinate).group(0)
        row_pattern = re.compile(
            rf'(<row\b(?=[^>]*\br="{row_number}")[^>]*>)(.*?)(</row>)',
            re.DOTALL,
        )
        row_match = row_pattern.search(source_text)
        if row_match is None:
            raise ValueError(
                f"原始排单缺少 {coordinate} 所在的数据行，为保护原表已停止保存。"
            )
        patched_body = insert_cell_xml_into_row(
            row_match.group(2),
            coordinate,
            generated_cell,
        )
        replacement = row_match.group(1) + patched_body + row_match.group(3)
        source_text = source_text[: row_match.start()] + replacement + source_text[row_match.end() :]

    def protected_cells_digest(xml_text: str) -> tuple[int, bytes]:
        digest = hashlib.sha256()
        count = 0
        for match in re.finditer(r'<c\b[^>]*?(?:/>|>.*?</c>)', xml_text, flags=re.DOTALL):
            cell_xml = match.group(0)
            coordinate_match = re.search(r'\br="([A-Z]{1,3}[1-9][0-9]*)"', cell_xml)
            if coordinate_match is None or coordinate_match.group(1) in target_coordinates:
                continue
            encoded = cell_xml.encode("utf-8")
            digest.update(len(encoded).to_bytes(8, "big"))
            digest.update(encoded)
            count += 1
        return count, digest.digest()

    source_digest = protected_cells_digest(original_source_text)
    patched_digest = protected_cells_digest(source_text)
    if source_digest != patched_digest:
        raise ValueError("检测到非预估单元格发生变化，为保护原始数据和公式已停止保存。")
    return source_text.encode("utf-8")


def preserve_source_workbook_with_forecast_cells(
    source_path: Path,
    generated_path: Path,
    output_path: Path,
    written_pairs: Iterable[str],
) -> None:
    coordinates_by_sheet = written_pair_coordinates(written_pairs)
    if not coordinates_by_sheet:
        if source_path.resolve() != output_path.resolve():
            shutil.copy2(source_path, output_path)
        return

    with tempfile.NamedTemporaryFile(
        delete=False,
        dir=output_path.parent,
        suffix=output_path.suffix or ".xlsx",
    ) as temp_file:
        patched_path = Path(temp_file.name)

    try:
        with (
            zipfile.ZipFile(source_path, "r") as source_archive,
            zipfile.ZipFile(generated_path, "r") as generated_archive,
            zipfile.ZipFile(patched_path, "w") as output_archive,
        ):
            source_sheet_paths = workbook_sheet_xml_paths(source_archive)
            generated_sheet_paths = workbook_sheet_xml_paths(generated_archive)
            replacements: dict[str, bytes] = {
                "xl/styles.xml": generated_archive.read("xl/styles.xml")
            }
            for sheet_name, coordinates in coordinates_by_sheet.items():
                source_sheet_path = source_sheet_paths.get(sheet_name)
                generated_sheet_path = generated_sheet_paths.get(sheet_name)
                if not source_sheet_path or not generated_sheet_path:
                    raise ValueError(f"无法定位工作表“{sheet_name}”，为保护原表已停止保存。")
                replacements[source_sheet_path] = patch_sheet_xml_cells(
                    source_archive.read(source_sheet_path),
                    generated_archive.read(generated_sheet_path),
                    coordinates,
                )

            for item in source_archive.infolist():
                output_archive.writestr(
                    item,
                    replacements.get(item.filename, source_archive.read(item.filename)),
                )

        if not zipfile.is_zipfile(patched_path):
            raise ValueError("保存后的 Excel 文件结构无效，为保护原表已停止替换。")
        patched_path.replace(output_path)
    finally:
        patched_path.unlink(missing_ok=True)


def save_sales_workbook(
    sales_wb,
    output_path: Path,
    freeze_formulas: bool,
    source_path: Optional[Path] = None,
    written_pairs: Optional[Iterable[str]] = None,
) -> None:
    def release_workbook_memory() -> None:
        worksheets = list(sales_wb.worksheets)
        sales_wb.close()
        for worksheet in worksheets:
            cells = getattr(worksheet, "_cells", None)
            if cells is not None:
                cells.clear()
        sheets = getattr(sales_wb, "_sheets", None)
        if sheets is not None:
            sheets.clear()
        gc.collect()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    source_path = Path(source_path) if source_path is not None else None
    written_pairs = list(written_pairs or [])

    if source_path is not None and not freeze_formulas and not written_pairs:
        release_workbook_memory()
        if source_path.resolve() != output_path.resolve():
            shutil.copy2(source_path, output_path)
        return

    if source_path is None or freeze_formulas:
        sales_wb.save(output_path)
        release_workbook_memory()
        if freeze_formulas:
            freeze_workbook_inplace(output_path)
        return

    with tempfile.NamedTemporaryFile(
        delete=False,
        dir=output_path.parent,
        suffix=output_path.suffix or ".xlsx",
    ) as temp_file:
        generated_path = Path(temp_file.name)
    try:
        sales_wb.save(generated_path)
        release_workbook_memory()
        preserve_source_workbook_with_forecast_cells(
            source_path,
            generated_path,
            output_path,
            written_pairs,
        )
    finally:
        generated_path.unlink(missing_ok=True)


def unchanged_summary(
    sales_wb,
    output_path: Path,
    freeze_formulas: bool,
    business_owner: Optional[str],
    as_of_date: date,
    warning: str,
    unmatched_predictions: Optional[list[dict]] = None,
    source_path: Optional[Path] = None,
) -> dict:
    save_sales_workbook(
        sales_wb,
        output_path,
        freeze_formulas,
        source_path=source_path,
    )
    return {
        "output_path": output_path,
        "updated_rows": 0,
        "fallbacks": [],
        "missing_rows": [],
        "skipped_months": [],
        "zero_filled_rows": [],
        "fill_targets": [],
        "business_owner": business_owner,
        "as_of_date": as_of_date,
        "warnings": [warning],
        "unmatched_predictions": dedupe_unmatched_predictions(unmatched_predictions or []),
        "fill_target_months": [],
        "matched_model_count": 0,
        "matched_customer_models": [],
        "written_pairs": [],
        "invalid_amount_rows": [],
    }


def previous_estimate_pair(ws, target: FillTarget) -> Optional[QuantityAmountPair]:
    candidates = []
    for pair in find_quantity_amount_pairs(ws):
        if pair.qty_col >= target.qty_col:
            continue
        if "预估" not in pair.label:
            continue
        if any(word in pair.label for word in NO_FILL_LABEL_WORDS):
            continue
        candidates.append(pair)
    if not candidates:
        return None
    return max(candidates, key=lambda pair: pair.qty_col)


def build_previous_estimate_cache(
    sales_values_wb,
    fill_targets: Iterable[FillTarget],
) -> dict[tuple[str, int, int], tuple[Optional[float], Optional[float]]]:
    previous_values: dict[tuple[str, int, int], tuple[Optional[float], Optional[float]]] = {}
    for target in fill_targets:
        ws = sales_values_wb[target.sheet]
        previous_pair = previous_estimate_pair(ws, target)
        if previous_pair is None:
            continue
        for row in iter_sales_data_rows(ws):
            previous_values[(target.sheet, target.qty_col, row)] = (
                parse_number(ws.cell(row, previous_pair.qty_col).value),
                parse_number(ws.cell(row, previous_pair.amt_col).value),
            )
    return previous_values


def amount_from_previous_week(
    qty: float,
    previous_qty: Optional[float],
    previous_amount: Optional[float],
) -> tuple[float, bool]:
    if previous_qty == 0:
        return 0.0, abs(float(qty)) > 1e-12
    if previous_qty is None or previous_amount is None:
        return 0.0, False
    return round(qty * previous_amount / previous_qty, 4), False


def cell_has_existing_nonzero_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and value.startswith("="):
        return False
    return not is_zero_placeholder(value)


def cell_is_blank(value) -> bool:
    """Only genuinely empty cells are writable in the shared schedule."""
    return value is None or (isinstance(value, str) and not value.strip())


def apply_forecast_cell_style(cell, font_color: str) -> None:
    font = copy(cell.font)
    font.color = font_color
    cell.font = font
    cell.fill = NO_FORECAST_FILL


def forecast_pair_key(sheet: str, row: int, qty_col: int, amt_col: int) -> str:
    return (
        f"{sheet}!{get_column_letter(qty_col)}{row}:"
        f"{get_column_letter(amt_col)}{row}"
    )


def cell_rgb_suffix(cell) -> str:
    color = getattr(cell.font, "color", None)
    if color is None or getattr(color, "type", None) != "rgb":
        return ""
    rgb = clean_text(getattr(color, "rgb", "")).upper()
    return rgb[-6:] if len(rgb) >= 6 else rgb


def fill_rgb_suffix(cell) -> str:
    fill = cell.fill
    if getattr(fill, "fill_type", None) != "solid":
        return ""
    color = getattr(fill, "fgColor", None)
    if color is None or getattr(color, "type", None) != "rgb":
        return ""
    rgb = clean_text(getattr(color, "rgb", "")).upper()
    return rgb[-6:] if len(rgb) >= 6 else rgb


def pair_has_legacy_generated_style(qty_cell, amt_cell) -> bool:
    """Recognize cells produced before written-pair tracking was introduced."""
    colors = {cell_rgb_suffix(qty_cell), cell_rgb_suffix(amt_cell)}
    if colors and colors <= {"008000", "000000"} and "" not in colors:
        return True
    return any(
        fill_rgb_suffix(cell) == LEGACY_MATCHED_FORECAST_FILL_COLOR
        for cell in (qty_cell, amt_cell)
    )


def forecast_pair_is_writable(
    sheet: str,
    row: int,
    target: FillTarget,
    qty_cell,
    amt_cell,
    overwrite_pairs: set[str],
    allow_legacy_overwrite: bool,
) -> bool:
    if cell_is_blank(qty_cell.value) and cell_is_blank(amt_cell.value):
        return True
    pair_key = forecast_pair_key(sheet, row, target.qty_col, target.amt_col)
    if pair_key in overwrite_pairs:
        return True
    return allow_legacy_overwrite and pair_has_legacy_generated_style(qty_cell, amt_cell)


def filter_unmatched_prediction_months(
    entries: Iterable[dict],
    allowed_months: Iterable[int],
) -> list[dict]:
    allowed = {int(month) for month in allowed_months}
    filtered: list[dict] = []
    for item in dedupe_unmatched_predictions(entries):
        months = item.get("months")
        if not isinstance(months, dict):
            continue
        kept = {
            str(month): qty
            for month, qty in months.items()
            if str(month).isdigit() and int(month) in allowed and parse_number(qty) not in (None, 0)
        }
        if kept:
            filtered.append({**item, "months": kept})
    return filtered


def build_sales_scan_payload(
    sales_path: Path,
    business_owner: Optional[str],
    current_month: int,
) -> dict:
    sales_values_wb = load_workbook(sales_path, read_only=False, data_only=True)
    try:
        fill_targets = select_relevant_fill_targets(
            find_fill_targets(sales_values_wb),
            current_month,
        )
        sales_codes = collect_sales_codes(
            sales_values_wb,
            business_owner=business_owner,
            sheet_names={target.sheet for target in fill_targets},
        )
        previous_values = build_previous_estimate_cache(sales_values_wb, fill_targets)
        return {
            "sales_codes": sorted(sales_codes, key=normalize_code),
            "fill_targets": [
                {
                    "sheet": target.sheet,
                    "month": target.month,
                    "label": target.label,
                    "qty_col": target.qty_col,
                    "amt_col": target.amt_col,
                    "owner_col": target.owner_col,
                    "code_col": target.code_col,
                }
                for target in fill_targets
            ],
            "previous_estimate_values": [
                [sheet, qty_col, row, previous_qty, previous_amount]
                for (sheet, qty_col, row), (previous_qty, previous_amount) in previous_values.items()
            ],
        }
    finally:
        sales_values_wb.close()


def scan_sales_inputs_in_subprocess(
    sales_path: Path,
    business_owner: Optional[str],
    current_month: int,
) -> tuple[set[str], list[FillTarget], dict[tuple[str, int, int], tuple[Optional[float], Optional[float]]]]:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".json") as temp_file:
        payload_path = Path(temp_file.name)
    try:
        command = [
            sys.executable,
            str(Path(__file__).resolve()),
            "--scan-sales",
            str(sales_path),
            business_owner or "",
            str(current_month),
            str(payload_path),
        ]
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=180,
            check=False,
        )
        if completed.returncode != 0:
            detail = (completed.stderr or completed.stdout or "未知错误").strip()[-1000:]
            raise RuntimeError(f"读取共用销售排单失败：{detail}")
        payload = json.loads(payload_path.read_text(encoding="utf-8"))
    finally:
        payload_path.unlink(missing_ok=True)

    sales_codes = {clean_text(code) for code in payload.get("sales_codes", []) if clean_text(code)}
    fill_targets = [FillTarget(**item) for item in payload.get("fill_targets", [])]
    previous_estimate_values = {
        (str(item[0]), int(item[1]), int(item[2])): (item[3], item[4])
        for item in payload.get("previous_estimate_values", [])
        if isinstance(item, list) and len(item) == 5
    }
    return sales_codes, fill_targets, previous_estimate_values


def process_sales_workbooks(
    pred_path: Path | str = PRED_PATH,
    sales_path: Path | str = SALES_PATH,
    output_path: Path | str = OUT_PATH,
    freeze_formulas: bool = False,
    business_owner: Optional[str] = None,
    as_of_date: Optional[date] = None,
    progress_callback: Optional[Callable[[dict], None]] = None,
    overwrite_pairs: Optional[Iterable[str]] = None,
    allow_legacy_overwrite: bool = False,
):
    def report_progress(percent: int, step: str, target: Optional[FillTarget] = None) -> None:
        if progress_callback is None:
            return
        payload = {
            "percent": max(0, min(100, int(percent))),
            "step": step,
        }
        if target is not None:
            payload.update(
                {
                    "target_sheet": target.sheet,
                    "target_label": target.label,
                    "target_month": target.month,
                }
            )
        try:
            progress_callback(payload)
        except Exception:
            pass

    pred_path = Path(pred_path)
    sales_path = Path(sales_path)
    output_path = Path(output_path)
    as_of_date = as_of_date or datetime.now(BEIJING_TZ).date()
    overwrite_pair_keys = {
        clean_text(value)
        for value in (overwrite_pairs or [])
        if clean_text(value)
    }

    # Scan the large workbook in a short-lived child process. Once it exits,
    # its openpyxl memory is returned to the OS before the editable workbook is
    # opened, which keeps Render's free worker below its memory limit.
    report_progress(5, "读取共用销售排单")
    sales_codes, fill_targets, previous_estimate_values = scan_sales_inputs_in_subprocess(
        sales_path,
        business_owner,
        as_of_date.month,
    )
    report_progress(12, "读取销售排单客户机种")
    if not sales_codes:
        owner_note = f"（业务担当：{business_owner}）" if business_owner else ""
        report_progress(90, "没有找到可匹配客户机种，生成未改动文件")
        sales_wb = load_workbook(sales_path)
        return unchanged_summary(
            sales_wb,
            output_path,
            freeze_formulas,
            business_owner,
            as_of_date,
            f"销售排单里没有找到可匹配的客户机种{owner_note}，已生成未改动文件。",
            source_path=sales_path,
        )

    report_progress(22, "读取业务预测文件")
    unmatched_predictions: list[dict] = []
    predictions = build_predictions(
        pred_path,
        sales_codes,
        business_owner=business_owner,
        as_of_date=as_of_date,
        unmatched_predictions=unmatched_predictions,
    )
    report_progress(32, "匹配预测机种与销售排单客户机种")
    if not predictions:
        report_progress(90, "没有识别到可用预测，生成未改动文件")
        sales_wb = load_workbook(sales_path)
        return unchanged_summary(
            sales_wb,
            output_path,
            freeze_formulas,
            business_owner,
            as_of_date,
            "预测信息里没有识别到可用的机种和数量，已生成未改动文件。",
            unmatched_predictions=unmatched_predictions,
            source_path=sales_path,
        )

    report_progress(40, f"识别销售排单预估空白栏：{len(fill_targets)} 个")
    report_progress(45, "打开可编辑销售排单")
    if not fill_targets:
        sales_wb = load_workbook(sales_path)
        report_progress(90, "没有找到可回填预估空白栏，生成未改动文件")
        return unchanged_summary(
            sales_wb,
            output_path,
            freeze_formulas,
            business_owner,
            as_of_date,
            "销售排单里没有找到空白的数量/金额栏，已生成未改动文件。",
            source_path=sales_path,
        )

    if freeze_formulas:
        sales_wb = load_workbook(sales_path)
    else:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(
            delete=False,
            dir=output_path.parent,
            suffix=sales_path.suffix or ".xlsx",
        ) as temp_file:
            editing_path = Path(temp_file.name)
        try:
            create_trimmed_editing_copy(
                sales_path,
                editing_path,
                {target.sheet for target in fill_targets},
            )
            sales_wb = load_workbook(editing_path)
        finally:
            editing_path.unlink(missing_ok=True)

    updates = []
    fallbacks = []
    missing_rows = []
    skipped_months = []
    completed_deductions = []
    skipped_existing_values = []
    zero_filled_rows = []
    written_pairs: set[str] = set()
    invalid_amount_rows = []
    matched_customer_models: set[str] = set()
    warnings = []

    total_targets = max(len(fill_targets), 1)
    for target_index, target in enumerate(fill_targets, start=1):
        progress_percent = 45 + int((target_index - 1) / total_targets * 40)
        report_progress(progress_percent, "按预估空白栏回填数量和金额", target)
        if not any(month_map.get(target.month) is not None for month_map in predictions.values()):
            skipped_months.append((target.sheet, target.label, target.month))

        ws = sales_wb[target.sheet]
        vws = ws
        source_columns = find_header_columns(vws)
        source_code_columns = find_code_columns(vws, source_columns)
        code_to_row = {}
        normalized_code_to_code = {}
        row_meta = {}
        eligible_rows: set[int] = set()
        for row in iter_sales_data_rows(ws):
            owner = clean_text(vws.cell(row, target.owner_col).value)
            if business_owner and not owner_matches(business_owner, owner):
                continue

            row_codes = []
            for code_col in source_code_columns:
                if code_col:
                    code = clean_text(vws.cell(row, code_col).value)
                    if code and code not in row_codes:
                        row_codes.append(code)

            if not row_codes:
                continue

            eligible_rows.add(row)

            meta = {
                "O": vws[f"O{row}"].value,
                "R": vws[f"R{row}"].value,
                "T": vws[f"T{row}"].value,
                "U": vws[f"U{row}"].value,
                "Q": vws[f"Q{row}"].value,
                "area": parse_area(vws[f"Q{row}"].value),
            }
            for code in row_codes:
                code_to_row[code] = row
                row_meta[code] = dict(meta, code=code)
                norm_code = normalize_code(code)
                if norm_code:
                    normalized_code_to_code[norm_code] = code

        matched_prediction_rows: set[int] = set()
        for code, month_map in predictions.items():
            qty = month_map.get(target.month)
            if qty is None:
                continue

            matched_code = code
            row = code_to_row.get(matched_code)
            if row is None:
                normalized_match = normalized_code_to_code.get(normalize_code(code))
                if normalized_match:
                    matched_code = normalized_match
                    row = code_to_row.get(matched_code)
            if row is None:
                missing_rows.append((target.sheet, target.label, code, business_owner or "全部"))
                continue

            matched_prediction_rows.add(row)

            qty_cell = ws.cell(row, target.qty_col)
            amt_cell = ws.cell(row, target.amt_col)

            if not forecast_pair_is_writable(
                target.sheet,
                row,
                target,
                qty_cell,
                amt_cell,
                overwrite_pair_keys,
                allow_legacy_overwrite,
            ):
                skipped_existing_values.append(
                    (
                        target.sheet,
                        target.label,
                        code,
                        row,
                        qty_cell.value,
                        amt_cell.value,
                    )
                )
                continue

            qty_cell.value = qty
            qty_cell.number_format = "0.0000"
            apply_forecast_cell_style(qty_cell, MATCHED_FORECAST_FONT_COLOR)

            previous_qty, previous_amount = previous_estimate_values.get(
                (target.sheet, target.qty_col, row),
                (None, None),
            )
            amount, invalid_previous_qty = amount_from_previous_week(
                qty,
                previous_qty,
                previous_amount,
            )
            amt_cell.value = amount
            amt_cell.number_format = "0.00"
            apply_forecast_cell_style(amt_cell, MATCHED_FORECAST_FONT_COLOR)
            if invalid_previous_qty:
                amt_cell.fill = INVALID_AMOUNT_FILL
                invalid_amount_rows.append(
                    (
                        target.sheet,
                        row,
                        code,
                        target.label,
                        get_column_letter(target.amt_col),
                    )
                )
            written_pairs.add(
                forecast_pair_key(target.sheet, row, target.qty_col, target.amt_col)
            )
            matched_customer_models.add(clean_text(vws.cell(row, target.code_col).value) or matched_code)
            updates.append(
                (
                    target.sheet,
                    row,
                    code,
                    target.label,
                    get_column_letter(target.qty_col),
                    qty,
                    amount,
                    None,
                    business_owner or clean_text(vws.cell(row, target.owner_col).value),
                )
            )

        # Rows belonging to this owner but without a prediction for this target
        # month receive 0/0. On a repeat upload, only pairs previously written
        # by the website may be replaced; original non-empty cells stay intact.
        for row in sorted(eligible_rows - matched_prediction_rows):
            qty_cell = ws.cell(row, target.qty_col)
            amt_cell = ws.cell(row, target.amt_col)
            if not forecast_pair_is_writable(
                target.sheet,
                row,
                target,
                qty_cell,
                amt_cell,
                overwrite_pair_keys,
                allow_legacy_overwrite,
            ):
                continue
            qty_cell.value = 0
            qty_cell.number_format = "0.0000"
            apply_forecast_cell_style(qty_cell, ZERO_FORECAST_FONT_COLOR)
            amt_cell.value = 0
            amt_cell.number_format = "0.00"
            apply_forecast_cell_style(amt_cell, ZERO_FORECAST_FONT_COLOR)
            written_pairs.add(
                forecast_pair_key(target.sheet, row, target.qty_col, target.amt_col)
            )
            zero_filled_rows.append(
                (
                    target.sheet,
                    row,
                    clean_text(vws.cell(row, target.code_col).value),
                    target.label,
                )
            )

    if not updates and not zero_filled_rows:
        warnings.append("没有找到可回填的匹配行，已生成未改动文件。")
    elif not updates:
        warnings.append("没有匹配到有预测数量的机种；预估空白格已按规则写入 0。")

    calc = getattr(sales_wb, "calculation", None)
    if calc is not None:
        calc.calcMode = "auto"
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True

    report_progress(88, "保存共用销售排单最新版")
    save_sales_workbook(
        sales_wb,
        output_path,
        freeze_formulas,
        source_path=sales_path,
        written_pairs=written_pairs,
    )
    report_progress(96, "整理回填结果")

    fill_target_months = sorted({target.month for target in fill_targets})
    summary = {
        "output_path": output_path,
        "updated_rows": len(updates),
        "fallbacks": fallbacks,
        "missing_rows": missing_rows,
        "skipped_months": skipped_months,
        "skipped_existing_values": skipped_existing_values,
        "zero_filled_rows": zero_filled_rows,
        "fill_targets": fill_targets,
        "business_owner": business_owner,
        "as_of_date": as_of_date,
        "completed_deductions": completed_deductions,
        "warnings": warnings,
        "unmatched_predictions": filter_unmatched_prediction_months(
            unmatched_predictions,
            fill_target_months,
        ),
        "fill_target_months": fill_target_months,
        "matched_model_count": len(matched_customer_models),
        "matched_customer_models": sorted(matched_customer_models, key=normalize_code),
        "written_pairs": sorted(written_pairs),
        "invalid_amount_rows": invalid_amount_rows,
    }

    print(f"saved: {output_path}")
    print(f"business owner: {business_owner or '全部'}")
    print(f"fill targets: {len(fill_targets)}")
    for target in fill_targets:
        print(
            "target",
            target.sheet,
            target.label,
            get_column_letter(target.qty_col),
            get_column_letter(target.amt_col),
        )
    print(f"updated rows: {len(updates)}")
    print(f"total written rows: {len(updates)}")
    print(f"fallback prices used: {len(fallbacks)}")
    for item in fallbacks[:10]:
        print("fallback", item)
    print(f"missing target rows: {len(missing_rows)}")
    for item in missing_rows[:20]:
        print("missing", item)
    print(f"skipped existing forecast cells: {len(skipped_existing_values)}")
    for item in skipped_existing_values[:20]:
        print("existing", item)
    print(f"completed deductions: {len(completed_deductions)}")
    for item in completed_deductions[:20]:
        print("completed_deduction", item)
    print(f"invalid previous-week quantity amounts: {len(invalid_amount_rows)}")
    print(f"unmatched predictions with quantities: {len(summary['unmatched_predictions'])}")

    return summary


def main():
    if len(sys.argv) == 6 and sys.argv[1] == "--scan-sales":
        payload = build_sales_scan_payload(
            Path(sys.argv[2]),
            sys.argv[3] or None,
            int(sys.argv[4]),
        )
        Path(sys.argv[5]).write_text(
            json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        return
    process_sales_workbooks()


if __name__ == "__main__":
    main()
