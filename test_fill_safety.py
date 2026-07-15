from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from fill_sales import (
    build_code_lookup,
    build_predictions_from_rows,
    build_predictions_from_workbook,
    build_sales_scan_payload,
    dedupe_unmatched_predictions,
    match_owner_fuzzy_code,
    match_prediction_row_code,
    patch_sheet_xml_cells,
    prediction_qty_to_wanpcs,
    process_sales_workbooks,
)


def test_owner_fuzzy_matching_is_unique_and_exact_first() -> None:
    lookup, norms = build_code_lookup(["E1A20009A-416"])
    assert match_owner_fuzzy_code("E1A20009A-426", "王永仁", lookup, norms) == "E1A20009A-416"

    lookup, norms = build_code_lookup(["E1A20009A-416", "E1A20009A-417"])
    assert match_owner_fuzzy_code("E1A20009A-426", "王永仁", lookup, norms) is None

    lookup, norms = build_code_lookup(["E1A20009A-416", "E1A20009A-426"])
    assert (
        match_prediction_row_code(
            ["E1A20009A-426"],
            [0],
            "王永仁",
            lookup,
            norms,
        )
        == "E1A20009A-426"
    )

    lookup, norms = build_code_lookup(["C5PS07 Holder+BG半成品", "M9XZ11"])
    assert (
        match_owner_fuzzy_code("C5PSO7 Holder+BG半成品", "周文龙", lookup, norms)
        == "C5PS07 Holder+BG半成品"
    )


def test_hongming_bare_quantities_are_already_wanpcs() -> None:
    assert prediction_qty_to_wanpcs(12.5, "洪鸣") == 12.5
    assert prediction_qty_to_wanpcs("12.5", "洪鸣") == 12.5
    assert prediction_qty_to_wanpcs("12500pcs", "洪鸣") == 1.25
    assert prediction_qty_to_wanpcs("2万", "洪鸣") == 2
    assert prediction_qty_to_wanpcs(12500, "王永仁") == 1.25


def test_unmatched_rows_become_zero_and_only_blank_forecast_cells_change(tmp_path: Path) -> None:
    sales_path = tmp_path / "sales.xlsx"
    prediction_path = tmp_path / "prediction.xlsx"
    output_path = tmp_path / "output.xlsx"

    sales_wb = Workbook()
    sales_ws = sales_wb.active
    sales_ws.title = "26年7月"
    sales_ws["E3"] = "业务担当"
    sales_ws["N3"] = "客户机种"
    sales_ws["T3"] = "6/29预估（7月）"
    sales_ws["T4"] = "数量"
    sales_ws["U4"] = "金额"
    sales_ws["R3"] = "7/1-7/13已完成（7月）"
    sales_ws["R4"] = "数量"
    sales_ws["S4"] = "金额"
    sales_ws["V3"] = "7/13预估（7月）"
    sales_ws["V4"] = "数量"
    sales_ws["W4"] = "金额"
    sales_ws["E5"] = "王永仁"
    sales_ws["N5"] = "E1A20009A-416"
    sales_ws["T5"] = 10
    sales_ws["U5"] = 20
    sales_ws["R5"] = 1.5
    sales_ws["S5"] = 3
    sales_ws["E6"] = "王永仁"
    sales_ws["N6"] = "UNMATCHED-MODEL"
    sales_ws["T6"] = 0
    sales_ws["U6"] = 5
    sales_ws["E7"] = "王永仁"
    sales_ws["N7"] = "EXISTING-ZERO"
    sales_ws["V7"] = 0
    sales_ws["W7"] = 0
    sales_ws["A5"] = "必须保持不变"
    sales_wb.save(sales_path)

    prediction_wb = Workbook()
    prediction_ws = prediction_wb.active
    prediction_ws["A1"] = "子件描述"
    prediction_ws["B1"] = "7月"
    prediction_ws["A2"] = "E1A20009A-426"
    prediction_ws["B2"] = 20000
    prediction_ws["A3"] = "NOT-IN-SALES"
    prediction_ws["B3"] = 5000
    prediction_wb.save(prediction_path)

    before_wb = load_workbook(sales_path, data_only=False)
    before = {
        (ws.title, cell.coordinate): (cell.value, cell.number_format)
        for ws in before_wb.worksheets
        for row in ws.iter_rows()
        for cell in row
    }
    before_wb.close()

    summary = process_sales_workbooks(
        pred_path=prediction_path,
        sales_path=sales_path,
        output_path=output_path,
        freeze_formulas=False,
        business_owner="王永仁",
        as_of_date=date(2026, 7, 13),
    )

    result_wb = load_workbook(output_path, data_only=False)
    result_ws = result_wb["26年7月"]
    assert summary["updated_rows"] == 1
    assert len(summary["zero_filled_rows"]) == 1
    assert summary["matched_model_count"] == 1
    assert summary["fill_target_months"] == [7]
    assert summary["completed_deductions"] == []
    assert summary["invalid_amount_rows"] == []
    assert summary["written_pairs"] == ["26年7月!V5:W5", "26年7月!V6:W6"]
    assert len(summary["unmatched_predictions"]) == 1
    assert summary["unmatched_predictions"][0]["model"] == "NOT-IN-SALES"
    assert result_ws["V5"].value == 2
    assert result_ws["W5"].value == 4
    assert result_ws["V5"].font.color.rgb == "FF008000"
    assert result_ws["W5"].font.color.rgb == "FF008000"
    assert result_ws["V5"].fill.fill_type is None
    assert result_ws["W5"].fill.fill_type is None
    assert result_ws["V6"].value == 0
    assert result_ws["W6"].value == 0
    assert result_ws["V6"].fill.fill_type is None
    assert result_ws["W6"].fill.fill_type is None
    assert result_ws["V6"].font.color.rgb == "FF000000"
    assert result_ws["W6"].font.color.rgb == "FF000000"
    assert result_ws["V7"].value == 0
    assert result_ws["W7"].value == 0

    allowed_changes = {
        ("26年7月", "V5"),
        ("26年7月", "W5"),
        ("26年7月", "V6"),
        ("26年7月", "W6"),
    }
    for ws in result_wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                key = (ws.title, cell.coordinate)
                if key in allowed_changes:
                    continue
                assert (cell.value, cell.number_format) == before.get(key, (None, "General"))
    result_wb.close()

    revised_prediction_path = tmp_path / "prediction_revised.xlsx"
    revised_output_path = tmp_path / "output_revised.xlsx"
    revised_prediction_wb = Workbook()
    revised_prediction_ws = revised_prediction_wb.active
    revised_prediction_ws["A1"] = "子件描述"
    revised_prediction_ws["B1"] = "7月"
    revised_prediction_ws["A2"] = "E1A20009A-426"
    revised_prediction_ws["B2"] = 30000
    revised_prediction_ws["A3"] = "UNMATCHED-MODEL"
    revised_prediction_ws["B3"] = 10000
    revised_prediction_wb.save(revised_prediction_path)

    revised_summary = process_sales_workbooks(
        pred_path=revised_prediction_path,
        sales_path=output_path,
        output_path=revised_output_path,
        freeze_formulas=False,
        business_owner="王永仁",
        as_of_date=date(2026, 7, 13),
        overwrite_pairs=summary["written_pairs"],
    )

    revised_wb = load_workbook(revised_output_path, data_only=False)
    revised_ws = revised_wb["26年7月"]
    assert revised_summary["updated_rows"] == 2
    assert revised_summary["completed_deductions"] == []
    assert revised_summary["invalid_amount_rows"] == [
        ("26年7月", 6, "UNMATCHED-MODEL", "7/13预估（7月）", "W")
    ]
    assert revised_ws["V5"].value == 3
    assert revised_ws["W5"].value == 6
    assert revised_ws["V6"].value == 1
    assert revised_ws["W6"].value == 0
    for coordinate in ("V5", "W5", "V6", "W6"):
        assert revised_ws[coordinate].font.color.rgb == "FF008000"
    for coordinate in ("V5", "W5", "V6"):
        assert revised_ws[coordinate].fill.fill_type is None
    assert revised_ws["W6"].fill.fill_type == "solid"
    assert revised_ws["W6"].fill.fgColor.rgb == "FF8B0000"
    assert revised_ws["V7"].value == 0
    assert revised_ws["W7"].value == 0
    assert "26年7月!V7:W7" not in revised_summary["written_pairs"]
    revised_wb.close()

    legacy_output_path = tmp_path / "output_legacy_revised.xlsx"
    legacy_summary = process_sales_workbooks(
        pred_path=revised_prediction_path,
        sales_path=output_path,
        output_path=legacy_output_path,
        freeze_formulas=False,
        business_owner="王永仁",
        as_of_date=date(2026, 7, 13),
        allow_legacy_overwrite=True,
    )
    legacy_wb = load_workbook(legacy_output_path, data_only=False)
    legacy_ws = legacy_wb["26年7月"]
    assert legacy_summary["updated_rows"] == 2
    assert legacy_ws["V5"].value == 3
    assert legacy_ws["V6"].value == 1
    assert legacy_ws["V7"].value == 0
    assert legacy_ws["W7"].value == 0
    legacy_wb.close()


def test_unmatched_list_only_contains_forecasts_with_quantities() -> None:
    rows = [
        ["子件描述", "7月", "8月"],
        ["E1A20009A-426", 20000, 30000],
        ["NO-SALES-MODEL", 12000, None],
        ["NO-QUANTITY-MODEL", None, None],
        ["合计", 32000, 30000],
    ]
    unmatched: list[dict] = []
    predictions = build_predictions_from_rows(
        rows,
        ["E1A20009A-416"],
        "王永仁",
        current_month=7,
        unmatched_predictions=unmatched,
        source_name="预测汇总",
    )

    assert predictions["E1A20009A-416"][7] == 2
    assert predictions["E1A20009A-416"][8] == 3
    assert len(unmatched) == 1
    assert unmatched[0]["model"] == "NO-SALES-MODEL"
    assert unmatched[0]["months"] == {"7": 1.2}
    assert unmatched[0]["source"] == "预测汇总"

    monthly_totals = dedupe_unmatched_predictions(
        [
            unmatched[0],
            {
                "model": "NO-SALES-MODEL",
                "months": {"7": 0.8, "8": 2.0},
                "source": "预测汇总",
            },
        ]
    )
    assert monthly_totals[0]["months"] == {"7": 2.0, "8": 2.0}


def test_prediction_reader_ignores_far_formatting_only_columns(tmp_path: Path) -> None:
    prediction_path = tmp_path / "wide_hongming.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "New part NO"
    ws["B1"] = "7月"
    ws["A2"] = "HM-001"
    ws["B2"] = 2.5
    ws.cell(1, 16000).fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    wb.save(prediction_path)

    loaded = load_workbook(prediction_path, data_only=True)
    assert loaded.active.max_column == 16000
    loaded.close()

    predictions = build_predictions_from_workbook(
        prediction_path,
        ["HM-001"],
        "洪鸣",
        current_month=7,
    )

    assert predictions["HM-001"][7] == 2.5


def test_xml_patch_preserves_adjacent_formula_cells() -> None:
    source_xml = b'''<?xml version="1.0" encoding="utf-8"?>
    <worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
      <sheetData><row r="5">
        <c r="BU5" s="60"/><c r="BV5" s="63"/><c r="BW5" s="57"/>
        <c r="BY5" s="57"><f>BU5-BQ5</f><v>0</v></c>
      </row></sheetData>
    </worksheet>'''
    generated_xml = b'''<?xml version="1.0" encoding="utf-8"?>
    <worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
      <sheetData><row r="5">
        <c r="BU5" s="100" t="n"><v>5</v></c>
        <c r="BV5" s="101" t="n"><v>0</v></c>
      </row></sheetData>
    </worksheet>'''

    patched = patch_sheet_xml_cells(source_xml, generated_xml, {"BU5", "BV5"})
    patched_text = patched.decode("utf-8")

    assert '<c r="BW5" s="57"/>' in patched_text
    assert "<f>BU5-BQ5</f>" in patched_text
    assert '<c r="BU5" s="100" t="n"><v>5</v></c>' in patched_text
    assert '<c r="BV5" s="101" t="n"><v>0</v></c>' in patched_text


def test_sales_scan_ignores_historical_month_sheets(tmp_path: Path) -> None:
    sales_path = tmp_path / "sales.xlsx"
    wb = Workbook()
    historical = wb.active
    historical.title = "26年6月"
    future = wb.create_sheet("26年7-12月")
    for ws, label, code in (
        (historical, "6/29预估（6月）", "HISTORICAL-ONLY"),
        (future, "7/14预估（7月）", "FUTURE-001"),
    ):
        ws["E3"] = "业务担当"
        ws["N3"] = "客户机种"
        ws["T3"] = label
        ws["T4"] = "数量"
        ws["U4"] = "金额"
        ws["E5"] = "洪鸣"
        ws["N5"] = code
    wb.save(sales_path)

    payload = build_sales_scan_payload(sales_path, "洪鸣", 7)

    assert payload["sales_codes"] == ["FUTURE-001"]
    assert [item["month"] for item in payload["fill_targets"]] == [7]
    assert [item["sheet"] for item in payload["fill_targets"]] == ["26年7-12月"]
