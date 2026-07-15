from __future__ import annotations

import cgi
import base64
import hashlib
import hmac
import html
import json
from functools import lru_cache
import os
import re
import shutil
import socket
import subprocess
import tempfile
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from datetime import datetime
from http.cookies import SimpleCookie
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter

from fill_sales import (
    BUSINESS_OWNERS,
    DATA_START_ROW,
    SUPPORTED_PREDICTION_SUFFIXES,
    SUPPORTED_SALES_SUFFIXES,
    extract_fill_target_month,
    find_quantity_amount_pairs,
    process_sales_workbooks,
)


HOST = "0.0.0.0"
PORT = int(os.environ.get("PORT", "8000"))
UPLOAD_MIME = "multipart/form-data"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PUBLIC_URL_ENV_VARS = ("PUBLIC_URL", "RENDER_EXTERNAL_URL", "APP_URL", "SITE_URL")
DEFAULT_OWNER = BUSINESS_OWNERS[0]
BEIJING_TZ = ZoneInfo("Asia/Shanghai")
DATA_DIR = Path(os.environ.get("SALES_DATA_DIR", "/data/sales-upload" if Path("/data").exists() else "server_data"))
MASTER_SALES_PATH = DATA_DIR / "current_sales.xlsx"
LATEST_OUTPUT_PATH = DATA_DIR / "latest_generated.xlsx"
METADATA_PATH = DATA_DIR / "metadata.json"
BACKUP_DIR = DATA_DIR / "backups"
UPLOAD_DIR = DATA_DIR / "uploads"
REMOTE_STORAGE_ENDPOINT = os.environ.get("NETLIFY_BLOBS_ENDPOINT", "").strip().rstrip("/")
REMOTE_STORAGE_SECRET = os.environ.get("SALES_STORAGE_SECRET", "").strip()
REMOTE_MASTER_KEY = "current_sales.xlsx"
REMOTE_LATEST_KEY = "latest_generated.xlsx"
REMOTE_METADATA_KEY = "metadata.json"
REMOTE_CHUNK_SIZE = 2 * 1024 * 1024
REMOTE_REQUEST_ATTEMPTS = 3
REMOTE_RETRYABLE_HTTP_CODES = {408, 425, 429, 500, 502, 503, 504}
LOGIN_USERNAME = os.environ.get("SALES_LOGIN_USERNAME", "admin").strip()
LOGIN_PASSWORD = os.environ.get("SALES_LOGIN_PASSWORD", "")
LOGIN_SECRET = os.environ.get("SALES_LOGIN_SECRET", "").strip() or REMOTE_STORAGE_SECRET
LOGIN_COOKIE_NAME = "sales_upload_session"
LOGIN_SESSION_SECONDS = 12 * 60 * 60
STATE_LOCK = threading.Lock()
JOB_LOCK = threading.Lock()
JOB_STATUS = {
    "state": "idle",
    "owner": "",
    "message": "",
    "started_at": "",
    "finished_at": "",
    "updated_rows": "",
    "error": "",
    "progress": "0",
    "step": "",
    "target": "",
}
OWNER_GUIDES = {
    "洪鸣": "读取预测文件中的 New part NO 匹配销售排单“客户机种”；上传数量单位按万 pcs 直接使用。",
    "李玎玲": "优先识别截图里的“机种名”和 6-10 月预测数量，自动匹配销售排单“客户机种”；数量单位 K 会自动换算成万 pcs。",
    "周文龙": "读取预测文件中的“品名”匹配销售排单“客户机种”；允许唯一、可信的缩写或近似机种匹配。",
    "王永仁": "读取预测文件中的“子件描述”前缀匹配销售排单“客户机种”；允许唯一版本升级和可信近似机种匹配。",
    "叶振华": "读取预测文件中的“模组型号”或“料号”匹配销售排单“客户机种”。",
    "李海鹰": "读取预测文件中的“料号”匹配销售排单“客户机种”。",
}
OWNER_STATUS_LABELS = {
    "pending": "未上传",
    "running": "处理中",
    "done": "已完成",
    "error": "失败",
}
OWNER_STATUS_LIST_FIELDS = {
    "unmatched_predictions",
    "fill_target_months",
    "written_pairs",
}


@lru_cache(maxsize=1)
def get_lan_ip() -> str:
    for iface in ("en0", "en1", "en2"):
        try:
            output = subprocess.check_output(["ifconfig", iface], text=True)
            for line in output.splitlines():
                line = line.strip()
                if line.startswith("inet ") and not line.startswith("inet 127."):
                    parts = line.split()
                    if len(parts) >= 2:
                        return parts[1]
        except Exception:
            pass

    backup_path: Optional[Path] = None
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
            sock.connect(("8.8.8.8", 80))
            ip = sock.getsockname()[0]
            if ip and not ip.startswith("127."):
                return ip
    except OSError:
        pass

    return "127.0.0.1"


def normalize_base_url(value: str) -> str:
    value = value.strip()
    if not value:
        return ""
    if "://" not in value:
        value = f"https://{value.lstrip('/')}"
    return value.rstrip("/")


def get_public_share_url(handler: Optional[BaseHTTPRequestHandler] = None) -> Optional[str]:
    for env_name in PUBLIC_URL_ENV_VARS:
        raw_value = os.environ.get(env_name)
        if raw_value:
            return normalize_base_url(raw_value)

    if handler is not None:
        forwarded_host = handler.headers.get("X-Forwarded-Host")
        if forwarded_host:
            forwarded_proto = handler.headers.get("X-Forwarded-Proto", "https")
            host = forwarded_host.split(",")[0].strip()
            proto = forwarded_proto.split(",")[0].strip() or "https"
            if host:
                return normalize_base_url(f"{proto}://{host}")

        host_header = handler.headers.get("Host")
        if host_header:
            host = host_header.split(",")[0].strip()
            host_lc = host.lower()
            host_name = host_lc.split(":", 1)[0]
            is_local_host = host_lc.startswith("[::1]") or host_name in {
                "127.0.0.1",
                "localhost",
                "0.0.0.0",
                "::1",
            }
            if not is_local_host:
                scheme = handler.headers.get("X-Forwarded-Proto", "https").split(",")[0].strip() or "https"
                return normalize_base_url(f"{scheme}://{host}")

    return None


def uploaded_suffix(filename: str) -> str:
    return Path(str(filename).replace("\\", "/")).suffix.lower()


def content_disposition(filename: str) -> str:
    quoted = urllib.parse.quote(filename)
    return f'attachment; filename="result.xlsx"; filename*=UTF-8\'\'{quoted}'


def safe_owner(value: Optional[str]) -> str:
    value = urllib.parse.unquote(value or "").strip()
    return value if value in BUSINESS_OWNERS else DEFAULT_OWNER


def owner_link(owner: str) -> str:
    return f"/owner/{urllib.parse.quote(owner)}"


def login_is_configured() -> bool:
    return bool(LOGIN_USERNAME and LOGIN_PASSWORD and LOGIN_SECRET)


def session_signature(timestamp: str) -> str:
    message = f"{LOGIN_USERNAME}|{timestamp}".encode("utf-8")
    return hmac.new(LOGIN_SECRET.encode("utf-8"), message, hashlib.sha256).hexdigest()


def authenticated(handler: BaseHTTPRequestHandler) -> bool:
    if not login_is_configured():
        return False
    cookie = SimpleCookie()
    try:
        cookie.load(handler.headers.get("Cookie", ""))
        morsel = cookie.get(LOGIN_COOKIE_NAME)
        if morsel is None:
            return False
        timestamp, signature = morsel.value.split(".", 1)
        issued_at = int(timestamp)
    except (ValueError, TypeError):
        return False
    if issued_at > int(time.time()) + 60 or int(time.time()) - issued_at > LOGIN_SESSION_SECONDS:
        return False
    return hmac.compare_digest(signature, session_signature(timestamp))


def render_login_page(error: str = "") -> str:
    if not login_is_configured():
        message = (
            "网站登录尚未完成安全配置。请在 Render 的 Environment 中设置 "
            "SALES_LOGIN_PASSWORD 和 SALES_LOGIN_SECRET，保存后重新部署。"
        )
        form_html = ""
    else:
        message = error
        form_html = """
        <form method="post" action="/login">
          <label for="username">账号</label>
          <input id="username" name="username" autocomplete="username" required autofocus>
          <label for="password">密码</label>
          <input id="password" name="password" type="password" autocomplete="current-password" required>
          <button type="submit">登录工作台</button>
        </form>
        """
    notice = f'<div class="notice">{html.escape(message)}</div>' if message else ""
    return f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>登录销售排单工作台</title><style>
:root{{color-scheme:dark}}*{{box-sizing:border-box}}body{{margin:0;min-height:100vh;display:grid;place-items:center;padding:24px;color:#edf8f7;font-family:"PingFang SC","Microsoft YaHei",sans-serif;background:radial-gradient(circle at 75% 10%,rgba(22,199,186,.16),transparent 30rem),#020914}}
.login{{width:min(440px,100%);padding:30px;border:1px solid #20384a;border-radius:16px;background:linear-gradient(145deg,#0a1b2a,#06111d);box-shadow:0 24px 70px rgba(0,0,0,.5)}}
h1{{margin:0 0 8px;font-size:26px}}p{{margin:0 0 22px;color:#91a6b4;line-height:1.7}}form{{display:grid;gap:10px}}label{{font-weight:800;margin-top:6px}}input{{width:100%;min-height:46px;padding:10px 12px;border:1px solid #2b4d62;border-radius:9px;color:#edf8f7;background:#020d18;font:inherit}}button{{min-height:46px;margin-top:10px;border:0;border-radius:9px;background:linear-gradient(135deg,#35ded0,#12b8ac);color:#001713;font:inherit;font-weight:900;cursor:pointer}}.notice{{margin-bottom:16px;padding:12px;border:1px solid rgba(255,116,108,.4);border-radius:9px;background:rgba(255,116,108,.1);color:#ff918a;line-height:1.6}}
</style></head><body><main class="login"><h1>销售排单回填工作台</h1><p>此网站包含内部销售数据，请先登录后使用。</p>{notice}{form_html}</main></body></html>"""


def redirect(handler: BaseHTTPRequestHandler, location: str, cookie: str = "") -> None:
    handler.send_response(303)
    handler.send_header("Location", location)
    handler.send_header("Cache-Control", "no-store")
    if cookie:
        handler.send_header("Set-Cookie", cookie)
    handler.end_headers()


def make_session_cookie(handler: BaseHTTPRequestHandler, clear: bool = False) -> str:
    secure = handler.headers.get("X-Forwarded-Proto", "").split(",")[0].strip().lower() == "https"
    if clear:
        value = ""
        max_age = 0
    else:
        timestamp = str(int(time.time()))
        value = f"{timestamp}.{session_signature(timestamp)}"
        max_age = LOGIN_SESSION_SECONDS
    parts = [
        f"{LOGIN_COOKIE_NAME}={value}",
        "Path=/",
        f"Max-Age={max_age}",
        "HttpOnly",
        "SameSite=Strict",
    ]
    if secure:
        parts.append("Secure")
    return "; ".join(parts)


def handle_login(handler: BaseHTTPRequestHandler) -> None:
    if not login_is_configured():
        write_html(handler, 503, render_login_page())
        return
    try:
        length = min(int(handler.headers.get("Content-Length", "0") or 0), 65536)
    except ValueError:
        length = 0
    values = urllib.parse.parse_qs(handler.rfile.read(length).decode("utf-8", errors="replace"))
    username = (values.get("username") or [""])[0]
    password = (values.get("password") or [""])[0]
    if hmac.compare_digest(username, LOGIN_USERNAME) and hmac.compare_digest(password, LOGIN_PASSWORD):
        redirect(handler, "/", make_session_cookie(handler))
        return
    write_html(handler, 401, render_login_page("账号或密码不正确，请重新输入。"))


def resolve_sheet_name(requested: str, sheet_names: list[str]) -> str:
    if not sheet_names:
        raise ValueError("工作簿里没有可用的 Sheet。")
    if requested in sheet_names:
        return requested
    normalized = requested.strip()
    if not normalized:
        return sheet_names[0]
    matches = [name for name in sheet_names if name.strip() == normalized]
    if len(matches) == 1:
        return matches[0]
    return requested


def ensure_data_dir() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def remote_storage_enabled() -> bool:
    return bool(REMOTE_STORAGE_ENDPOINT and REMOTE_STORAGE_SECRET)


def remote_storage_url(key: str) -> str:
    return f"{REMOTE_STORAGE_ENDPOINT}?key={urllib.parse.quote(key)}"


def remote_request(
    method: str,
    key: str,
    data: bytes | None = None,
    content_type: str = "application/octet-stream",
) -> Optional[bytes]:
    if not remote_storage_enabled():
        return None

    headers = {"x-sales-storage-secret": REMOTE_STORAGE_SECRET}
    if data is not None:
        headers["Content-Type"] = content_type
    last_error: Optional[Exception] = None
    for attempt in range(1, REMOTE_REQUEST_ATTEMPTS + 1):
        request = urllib.request.Request(
            remote_storage_url(key),
            data=data,
            headers=headers,
            method=method,
        )
        try:
            with urllib.request.urlopen(request, timeout=60) as response:
                return response.read()
        except urllib.error.HTTPError as exc:
            if exc.code == 404 and method == "GET":
                return None
            error_body = exc.read().decode("utf-8", errors="replace")[:500]
            last_error = RuntimeError(
                f"Netlify Blobs 请求失败：HTTP {exc.code} {error_body}"
            )
            if exc.code not in REMOTE_RETRYABLE_HTTP_CODES:
                raise last_error from exc
        except (urllib.error.URLError, TimeoutError, OSError) as exc:
            reason = getattr(exc, "reason", exc)
            last_error = RuntimeError(f"连接 Netlify Blobs 失败：{reason}")

        if attempt < REMOTE_REQUEST_ATTEMPTS:
            time.sleep(attempt)

    if last_error is not None:
        raise last_error
    return None


def remote_put_file(key: str, path: Path, content_type: str = XLSX_MIME) -> None:
    if not remote_storage_enabled() or not path.exists():
        return
    file_size = path.stat().st_size
    if file_size <= REMOTE_CHUNK_SIZE:
        remote_request("PUT", key, path.read_bytes(), content_type=content_type)
        return

    chunk_keys = []
    with path.open("rb") as source:
        index = 0
        while True:
            chunk = source.read(REMOTE_CHUNK_SIZE)
            if not chunk:
                break
            chunk_key = f"chunks/{key}/{index:05d}"
            remote_request("PUT", chunk_key, chunk, content_type="application/octet-stream")
            chunk_keys.append(chunk_key)
            index += 1

    remote_put_json(
        f"manifests/{key}.json",
        {
            "key": key,
            "content_type": content_type,
            "size": file_size,
            "chunk_size": REMOTE_CHUNK_SIZE,
            "chunks": chunk_keys,
            "updated_at": now_label(),
        },
    )


def remote_get_file(key: str, path: Path) -> bool:
    if not remote_storage_enabled():
        return False

    ensure_data_dir()
    temp_path = path.with_name(f".{path.name}.download")
    expected_size: Optional[int] = None
    manifest_data = remote_request("GET", f"manifests/{key}.json")
    if manifest_data is not None:
        manifest = json.loads(manifest_data.decode("utf-8"))
        chunk_keys = manifest.get("chunks") or []
        if not chunk_keys:
            return False
        expected_size = int(manifest.get("size") or 0) or None
        with temp_path.open("wb") as out:
            for chunk_key in chunk_keys:
                chunk = remote_request("GET", str(chunk_key))
                if chunk is None:
                    temp_path.unlink(missing_ok=True)
                    return False
                out.write(chunk)
        if expected_size is not None and temp_path.stat().st_size != expected_size:
            temp_path.unlink(missing_ok=True)
            raise RuntimeError("Netlify Blobs 里的 Excel 分片不完整，请重新上传共用销售排单。")
        if key.lower().endswith((".xlsx", ".xlsm")) and not zipfile.is_zipfile(temp_path):
            temp_path.unlink(missing_ok=True)
            raise RuntimeError("Netlify Blobs 里的 Excel 文件不完整，请重新上传共用销售排单。")
        temp_path.replace(path)
        return True

    data = remote_request("GET", key)
    if data is None:
        return False
    temp_path.write_bytes(data)
    if key.lower().endswith((".xlsx", ".xlsm")) and not zipfile.is_zipfile(temp_path):
        temp_path.unlink(missing_ok=True)
        raise RuntimeError("Netlify Blobs 里的 Excel 文件不完整，请重新上传共用销售排单。")
    temp_path.replace(path)
    return True


def remote_put_json(key: str, payload: dict) -> None:
    if not remote_storage_enabled():
        return
    remote_request(
        "PUT",
        key,
        json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"),
        content_type="application/json; charset=utf-8",
    )


def hydrate_remote_state() -> None:
    if not remote_storage_enabled():
        return
    ensure_data_dir()
    if not METADATA_PATH.exists():
        remote_get_file(REMOTE_METADATA_KEY, METADATA_PATH)
    if not MASTER_SALES_PATH.exists():
        remote_get_file(REMOTE_MASTER_KEY, MASTER_SALES_PATH)
    if MASTER_SALES_PATH.exists() and not LATEST_OUTPUT_PATH.exists():
        shutil.copy2(MASTER_SALES_PATH, LATEST_OUTPUT_PATH)


def sync_master_files_to_remote() -> None:
    if not remote_storage_enabled():
        return
    # Both local paths contain the same latest workbook. Uploading it once
    # halves the remote requests and avoids unnecessary free-instance delays.
    remote_put_file(REMOTE_MASTER_KEY, MASTER_SALES_PATH)


def load_metadata() -> dict:
    if remote_storage_enabled() and not METADATA_PATH.exists():
        hydrate_remote_state()
    if not METADATA_PATH.exists():
        return {}
    try:
        return json.loads(METADATA_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def default_owner_statuses() -> dict[str, dict]:
    return {
        owner: {
            "state": "pending",
            "updated_rows": "",
            "progress": "",
            "updated_at": "",
            "uploaded_at": "",
            "prediction_name": "",
            "error": "",
            "unmatched_predictions": [],
            "fill_target_months": [],
            "matched_model_count": "",
            "written_pairs": [],
            "storage_warning": "",
        }
        for owner in BUSINESS_OWNERS
    }


def ensure_metadata_schema(metadata: dict) -> dict:
    statuses = metadata.get("owner_statuses")
    if not isinstance(statuses, dict):
        statuses = {}

    normalized = default_owner_statuses()
    for owner in BUSINESS_OWNERS:
        raw = statuses.get(owner)
        if isinstance(raw, dict):
            normalized[owner].update(
                {
                    key: str(value)
                    for key, value in raw.items()
                    if value is not None and key not in OWNER_STATUS_LIST_FIELDS
                }
            )
            raw_unmatched = raw.get("unmatched_predictions")
            if isinstance(raw_unmatched, list):
                normalized[owner]["unmatched_predictions"] = [
                    item for item in raw_unmatched if isinstance(item, dict)
                ]
            raw_months = raw.get("fill_target_months")
            if isinstance(raw_months, list):
                normalized[owner]["fill_target_months"] = [
                    int(month)
                    for month in raw_months
                    if str(month).isdigit() and 1 <= int(month) <= 12
                ]
            raw_written_pairs = raw.get("written_pairs")
            if isinstance(raw_written_pairs, list):
                normalized[owner]["written_pairs"] = [
                    str(value)
                    for value in raw_written_pairs
                    if str(value).strip()
                ]
            if normalized[owner].get("state") not in OWNER_STATUS_LABELS:
                normalized[owner]["state"] = "pending"

    metadata["owner_statuses"] = normalized
    return metadata


def save_metadata(metadata: dict) -> str:
    ensure_data_dir()
    metadata = ensure_metadata_schema(metadata)
    METADATA_PATH.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    try:
        remote_put_json(REMOTE_METADATA_KEY, metadata)
    except Exception as exc:  # noqa: BLE001
        warning = f"状态远程备份暂时失败：{exc}"
        metadata["storage_warning"] = warning
        metadata["storage_warning_at"] = now_label()
        METADATA_PATH.write_text(
            json.dumps(metadata, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(warning)
        return warning
    return ""


def update_owner_status(owner: str, **values) -> None:
    with STATE_LOCK:
        metadata = ensure_metadata_schema(load_metadata())
        status = metadata["owner_statuses"].setdefault(owner, default_owner_statuses().get(owner, {}))
        for key, value in values.items():
            if key in OWNER_STATUS_LIST_FIELDS:
                status[key] = value if isinstance(value, list) else []
            else:
                status[key] = "" if value is None else str(value)
        metadata["owner_statuses"][owner] = status
        save_metadata(metadata)


def owner_overwrite_context(owner: str) -> tuple[list[str], bool]:
    """Return the cells this owner may replace on a repeated upload."""
    with STATE_LOCK:
        metadata = ensure_metadata_schema(load_metadata())
        status = metadata["owner_statuses"].get(owner, {})
        written_pairs = status.get("written_pairs")
        if not isinstance(written_pairs, list):
            written_pairs = []
        has_previous_result = (
            status.get("state") == "done"
            or bool(str(status.get("updated_at") or "").strip())
        )
        normalized_pairs = [str(value) for value in written_pairs if str(value).strip()]
        return normalized_pairs, has_previous_result and not normalized_pairs


def get_job_status() -> dict:
    with JOB_LOCK:
        return dict(JOB_STATUS)


def set_job_status(**values) -> None:
    with JOB_LOCK:
        JOB_STATUS.update(values)


def job_is_running() -> bool:
    return get_job_status().get("state") == "running"


def format_progress(value) -> str:
    try:
        percent = int(float(value))
    except (TypeError, ValueError):
        percent = 0
    return str(max(0, min(100, percent)))


def beijing_now() -> datetime:
    return datetime.now(BEIJING_TZ)


def now_label() -> str:
    return beijing_now().strftime("%Y-%m-%d %H:%M:%S 北京时间")


def safe_filename(value: str, default: str = "销售排单.xlsx") -> str:
    name = os.path.basename(str(value).replace("\\", "/")).strip()
    return name or default


def master_sales_exists() -> bool:
    if not MASTER_SALES_PATH.exists() and remote_storage_enabled():
        hydrate_remote_state()
    if not MASTER_SALES_PATH.exists() or MASTER_SALES_PATH.stat().st_size <= 0:
        return False
    return zipfile.is_zipfile(MASTER_SALES_PATH)


def storage_description() -> str:
    if remote_storage_enabled():
        return "当前已启用 Netlify Blobs 远程保存；Render 重启后会自动恢复本周排单和业务状态。"
    if str(DATA_DIR).startswith("/data"):
        return "当前使用 Render 持久磁盘目录 /data，网站重启后仍会保留本周数据。"
    return "当前使用临时目录；若 Render 服务重启或重新部署，本周数据可能丢失，请在 Render 增加持久磁盘并挂载到 /data。"


def format_file_size(path: Path) -> str:
    if not path.exists():
        return "0 KB"
    size = path.stat().st_size
    if size >= 1024 * 1024:
        return f"{size / 1024 / 1024:.1f} MB"
    return f"{max(size / 1024, 1):.0f} KB"


def selected_owner_from_path(path: str, query: dict[str, list[str]]) -> str:
    if path.startswith("/owner/"):
        return safe_owner(path.removeprefix("/owner/"))
    if "owner" in query and query["owner"]:
        return safe_owner(query["owner"][0])
    return DEFAULT_OWNER


def render_page(
    message: str = "",
    error: str = "",
    handler: Optional[BaseHTTPRequestHandler] = None,
    selected_owner: Optional[str] = None,
) -> bytes:
    selected_owner = safe_owner(selected_owner)
    selected_owner_guide = OWNER_GUIDES.get(selected_owner, "按预测文件中的机种信息匹配销售排单“客户机种”。")
    if selected_owner == "李玎玲":
        upload_hint = "李玎玲页面建议上传完整截图，保留左侧“机种名”列和右侧 6-10 月数量列，系统会按表格结构去识别。"
    else:
        upload_hint = "预测信息优先使用表格文件，也可以上传清晰截图。"
    local_url = f"http://127.0.0.1:{PORT}/"
    public_url = get_public_share_url(handler)
    lan_ip = get_lan_ip() if public_url is None else ""
    lan_url = f"http://{lan_ip}:{PORT}/" if lan_ip and lan_ip != "127.0.0.1" else ""

    if public_url:
        access_note = f"公网地址：{public_url}/"
    elif lan_url:
        access_note = f"本机：{local_url} ｜ 同一 Wi-Fi：{lan_url}"
    else:
        access_note = f"本机：{local_url}"

    owner_nav = "\n".join(
        f'<a class="owner-link {"active" if owner == selected_owner else ""}" href="{owner_link(owner)}">{html.escape(owner)}</a>'
        for owner in BUSINESS_OWNERS
    )
    owner_options = "\n".join(
        f'<option value="{html.escape(owner)}" {"selected" if owner == selected_owner else ""}>{html.escape(owner)}</option>'
        for owner in BUSINESS_OWNERS
    )
    message_html = (
        f'<div class="notice success" role="status">{html.escape(message)}</div>' if message else ""
    )
    error_html = f'<div class="notice error" role="alert">{html.escape(error)}</div>' if error else ""
    metadata = ensure_metadata_schema(load_metadata())
    has_master = master_sales_exists()
    master_name = metadata.get("master_name", "未上传共用销售排单")
    master_time = metadata.get("master_uploaded_at", "")
    latest_owner = metadata.get("last_owner", "")
    latest_time = metadata.get("last_generated_at", "")
    latest_rows = metadata.get("last_updated_rows", "")
    latest_models = metadata.get("last_matched_model_count", "")
    latest_file = metadata.get("latest_name", metadata.get("master_name", "最新销售排单.xlsx"))
    storage_warning = str(metadata.get("storage_warning") or "").strip()
    storage_warning_html = (
        f'<div class="notice working" role="status">{html.escape(storage_warning)}。'
        '当前最新版仍可下载和继续回填，后续保存会自动再次尝试远程备份。</div>'
        if storage_warning
        else ""
    )
    master_status = "已启用" if has_master else "待上传"
    master_status_class = "ready" if has_master else "empty"
    download_latest_html = (
        '<a class="secondary button" href="/download/latest">下载当前最新版</a>' if has_master else ""
    )
    preview_latest_html = ""
    state_text = (
        f"当前共用排单：{master_name}（{format_file_size(MASTER_SALES_PATH)}）"
        if has_master
        else "请先上传本周共用销售排单，然后各业务只上传自己的预测。"
    )
    detail_parts = []
    if master_time:
        detail_parts.append(f"初始化/替换时间：{master_time}")
    if latest_owner:
        result_text = f"已回填 {latest_models} 个客户机种" if str(latest_models) else f"更新 {latest_rows} 行"
        detail_parts.append(f"最近回填：{latest_owner}，{result_text}，{latest_time}")
    if latest_file and has_master:
        detail_parts.append(f"下载文件名：{latest_file}")
    detail_parts.append(storage_description())
    detail_text = " ｜ ".join(detail_parts) if detail_parts else "共用排单会在每位业务上传预测后自动保存为最新版。"
    owner_statuses = metadata.get("owner_statuses", default_owner_statuses())
    owner_status_card_items = []
    for owner in BUSINESS_OWNERS:
        status = owner_statuses.get(owner, {})
        state = status.get("state", "pending")
        progress_text = format_progress(status.get("progress")) if status.get("progress") else ""
        if state == "running" and progress_text:
            row_text = f"处理中 {html.escape(progress_text)}%"
        else:
            matched_count = status.get("matched_model_count", "")
            row_text = (
                f'已回填 {html.escape(str(matched_count))} 个客户机种'
                if state == "done" and str(matched_count) != ""
                else "等待预测"
            )
        prediction_name = status.get("prediction_name") or "暂无上传记录"
        uploaded_at = status.get("uploaded_at") or status.get("updated_at") or ""
        done_at = status.get("updated_at") or ""
        error_text = status.get("error") or ""
        owner_storage_warning = status.get("storage_warning") or ""
        unmatched_items = status.get("unmatched_predictions")
        unmatched_count = len(unmatched_items) if isinstance(unmatched_items, list) else 0
        owner_status_card_items.append(
            f'<div class="owner-status {html.escape(state)}">'
            f'<strong>{html.escape(owner)}</strong>'
            f'<span>{html.escape(OWNER_STATUS_LABELS.get(state, "未上传"))}</span>'
            f'<small>最近文件：{html.escape(prediction_name)}</small>'
            f'<small>上传时间：{html.escape(uploaded_at or "未上传")}</small>'
            f'<small>回填结果：{row_text}{("，" + html.escape(done_at)) if done_at else ""}</small>'
            f'<small>未匹配：{unmatched_count} 项</small>'
            f'{f"<small>错误：{html.escape(error_text)}</small>" if error_text else ""}'
            f'{f"<small>备份提示：{html.escape(owner_storage_warning)}</small>" if owner_storage_warning else ""}'
            f'</div>'
        )
    owner_status_cards = "\n".join(owner_status_card_items)
    selected_status = owner_statuses.get(selected_owner, {})
    selected_unmatched = selected_status.get("unmatched_predictions")
    if not isinstance(selected_unmatched, list):
        selected_unmatched = []
    selected_target_months = selected_status.get("fill_target_months")
    if not isinstance(selected_target_months, list):
        selected_target_months = []
    selected_target_months = sorted(
        {int(month) for month in selected_target_months if str(month).isdigit() and 1 <= int(month) <= 12}
    )
    target_month_set = set(selected_target_months)
    unmatched_row_items = []
    unmatched_month_totals: dict[int, float] = {month: 0.0 for month in selected_target_months}
    for item in selected_unmatched:
        if not isinstance(item, dict):
            continue
        model = html.escape(str(item.get("model") or "未命名机种"))
        month_parts = []
        months = item.get("months")
        if isinstance(months, dict):
            def month_sort_key(pair):
                try:
                    return int(pair[0])
                except (TypeError, ValueError):
                    return 99

            for month, qty in sorted(months.items(), key=month_sort_key):
                if target_month_set and (not str(month).isdigit() or int(month) not in target_month_set):
                    continue
                try:
                    qty_number = float(qty)
                    qty_text = f"{qty_number:.4f}".rstrip("0").rstrip(".")
                except (TypeError, ValueError):
                    qty_text = str(qty)
                try:
                    month_number = int(month)
                    unmatched_month_totals[month_number] = round(
                        unmatched_month_totals.get(month_number, 0.0) + float(qty),
                        4,
                    )
                except (TypeError, ValueError):
                    pass
                month_parts.append(f"{html.escape(str(month))}月：{html.escape(qty_text)} 万pcs")
        if not month_parts:
            continue
        month_text = " ｜ ".join(month_parts)
        unmatched_row_items.append(
            f'<tr><td><strong>{model}</strong></td><td>{month_text}</td></tr>'
        )
    if unmatched_row_items:
        unmatched_body = "".join(unmatched_row_items)
        unmatched_total_items = []
        for month, qty in sorted(unmatched_month_totals.items()):
            qty_text = f"{qty:.4f}".rstrip("0").rstrip(".")
            unmatched_total_items.append(
                f'<div class="month-total"><span>{month}月合计</span><strong>{html.escape(qty_text)} 万pcs</strong></div>'
            )
        unmatched_totals_html = "".join(unmatched_total_items)
        unmatched_panel_html = (
            f'<section class="unmatched-panel" aria-label="未匹配预测列表">'
            f'<div class="unmatched-head"><div><h3>未匹配预测列表</h3>'
            f'<p>以下 {len(unmatched_row_items)} 个机种在预测文件中有数量，但没有匹配到共用销售排单，因此没有回填。</p></div>'
            f'<span class="unmatched-count">{len(unmatched_row_items)} 项</span></div>'
            f'<div class="month-totals" aria-label="未匹配预测各月合计">{unmatched_totals_html}</div>'
            f'<div class="table-scroll"><table><thead><tr><th>预测文件机种</th><th>预测文件月份数量</th>'
            f'</tr></thead><tbody>{unmatched_body}</tbody></table></div></section>'
        )
    elif selected_status.get("state") == "done":
        unmatched_panel_html = (
            '<section class="unmatched-panel empty-list" aria-label="未匹配预测列表">'
            '<div><h3>未匹配预测列表</h3><p>最近一次预测中，所有有数量的机种均已匹配。</p></div>'
            '<span class="unmatched-count ok">0 项</span></section>'
        )
    else:
        unmatched_panel_html = (
            '<section class="unmatched-panel empty-list" aria-label="未匹配预测列表">'
            '<div><h3>未匹配预测列表</h3><p>上传并处理预测后，这里会列出有数量但没有匹配成功的机种。</p></div>'
            '<span class="unmatched-count">待处理</span></section>'
        )
    job = get_job_status()
    job_state = job.get("state", "idle")
    refresh_url = owner_link(selected_owner)
    if job_state == "running":
        progress = format_progress(job.get("progress"))
        step = job.get("step") or job.get("message") or "正在处理"
        target_text = job.get("target") or ""
        job_html = (
            f'<div class="notice working" role="status">'
            f'<div>正在处理：{html.escape(job.get("owner", ""))} 的预测，进度 {html.escape(progress)}%。'
            f'当前步骤：{html.escape(step)}{(" ｜ " + html.escape(target_text)) if target_text else ""}。</div>'
            f'<div class="progress-track" aria-label="处理进度"><span style="width: {html.escape(progress)}%"></span></div>'
            f'<div class="progress-note">开始时间：{html.escape(job.get("started_at", ""))}。处理完成前请不要重复上传。</div>'
            f'</div>'
        )
        refresh_meta = f'<meta http-equiv="refresh" content="5; url={html.escape(refresh_url)}">'
    elif job_state == "done":
        completed_count = job.get("matched_model_count", "")
        completed_text = (
            f'已回填 {html.escape(str(completed_count))} 个客户机种'
            if str(completed_count) != ""
            else f'更新 {html.escape(str(job.get("updated_rows", "")))} 行'
        )
        job_storage_warning = str(job.get("storage_warning") or "").strip()
        job_storage_warning_html = (
            f'<div class="notice working" role="status">'
            f'{html.escape(job_storage_warning)}。当前最新版仍可下载。</div>'
            if job_storage_warning
            else ""
        )
        job_html = (
            f'<div class="notice success" role="status">最近处理完成：{html.escape(job.get("owner", ""))}，'
            f'{completed_text}，'
            f'{html.escape(job.get("finished_at", ""))}。可以下载当前最新版。</div>'
            f'{job_storage_warning_html}'
        )
        refresh_meta = ""
    elif job_state == "error":
        job_html = (
            f'<div class="notice error" role="alert">最近处理失败：{html.escape(job.get("owner", ""))}，'
            f'{html.escape(job.get("error", ""))}</div>'
        )
        refresh_meta = ""
    else:
        job_html = ""
        refresh_meta = ""

    page = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  {refresh_meta}
  <title>销售排单回填工作台</title>
  <style>
    :root {{
      color-scheme: dark;
      --paper: #020914;
      --panel: #071522;
      --panel-raised: #0a1b2a;
      --line: #20384a;
      --line-bright: #2b4d62;
      --text: #edf8f7;
      --muted: #91a6b4;
      --accent: #16c7ba;
      --accent-dark: #6cf0e5;
      --warn: #f5a524;
      --danger: #ff746c;
      --success: #35d48d;
      --shadow: 0 20px 50px rgba(0, 0, 0, 0.36);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      min-height: 100vh;
      color: var(--text);
      font-family: "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", -apple-system, BlinkMacSystemFont, sans-serif;
      background:
        radial-gradient(circle at 84% 0%, rgba(22, 199, 186, 0.12), transparent 30rem),
        radial-gradient(circle at 8% 42%, rgba(20, 102, 161, 0.10), transparent 34rem),
        linear-gradient(90deg, rgba(67, 103, 126, 0.08) 1px, transparent 1px),
        linear-gradient(180deg, rgba(67, 103, 126, 0.08) 1px, transparent 1px),
        var(--paper);
      background-size: auto, auto, 32px 32px, 32px 32px, auto;
    }}
    a {{ color: inherit; }}
    .shell {{
      width: min(1180px, calc(100% - 28px));
      margin: 0 auto;
      padding: 22px 0 40px;
    }}
    .topline {{
      display: flex;
      justify-content: space-between;
      gap: 16px;
      align-items: flex-start;
      padding: 14px 0 18px;
      border-bottom: 1px solid var(--line);
      position: relative;
    }}
    .topline::after {{
      content: "";
      position: absolute;
      left: 0;
      bottom: -1px;
      width: 128px;
      height: 2px;
      background: linear-gradient(90deg, var(--accent), transparent);
    }}
    h1 {{
      margin: 0 0 6px;
      font-size: 26px;
      line-height: 1.25;
      letter-spacing: 0;
    }}
    .subtext {{
      margin: 0;
      color: var(--muted);
      line-height: 1.65;
      font-size: 14px;
    }}
    .access {{
      max-width: 430px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.6;
      text-align: right;
    }}
    .layout {{
      display: grid;
      grid-template-columns: 220px minmax(0, 1fr);
      gap: 18px;
      margin-top: 18px;
      align-items: start;
    }}
    .rail,
    .panel,
    .summary-item {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 12px;
      box-shadow: var(--shadow);
    }}
    .rail {{
      padding: 12px;
      position: sticky;
      top: 14px;
    }}
    .rail-title {{
      padding: 4px 6px 10px;
      color: var(--muted);
      font-size: 13px;
      font-weight: 700;
    }}
    .owner-list {{
      display: grid;
      gap: 6px;
    }}
    .owner-link {{
      display: flex;
      align-items: center;
      min-height: 38px;
      padding: 9px 10px;
      border-radius: 8px;
      text-decoration: none;
      color: var(--text);
      border: 1px solid transparent;
      font-weight: 700;
    }}
    .owner-link.active {{
      color: #dffffb;
      background: linear-gradient(90deg, rgba(22, 199, 186, 0.22), rgba(22, 199, 186, 0.06));
      border-color: rgba(22, 199, 186, 0.56);
      box-shadow: inset 3px 0 0 var(--accent);
    }}
    .panel {{
      padding: 20px;
    }}
    .panel-head {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: flex-start;
      margin-bottom: 16px;
    }}
    h2 {{
      margin: 0 0 6px;
      font-size: 20px;
      line-height: 1.3;
      letter-spacing: 0;
    }}
    .badge {{
      display: inline-flex;
      align-items: center;
      min-height: 32px;
      padding: 6px 10px;
      border-radius: 999px;
      background: rgba(245, 165, 36, 0.13);
      border: 1px solid rgba(245, 165, 36, 0.32);
      color: var(--warn);
      font-weight: 800;
      white-space: nowrap;
    }}
    .badge.ready {{
      background: rgba(53, 212, 141, 0.12);
      border-color: rgba(53, 212, 141, 0.3);
      color: var(--success);
    }}
    .badge.empty {{
      background: rgba(255, 116, 108, 0.12);
      border-color: rgba(255, 116, 108, 0.30);
      color: var(--danger);
    }}
    .notice {{
      margin: 0 0 14px;
      padding: 12px 14px;
      border-radius: 10px;
      border: 1px solid transparent;
      line-height: 1.6;
      font-weight: 700;
    }}
    .notice.success {{
      background: rgba(53, 212, 141, 0.10);
      border-color: rgba(53, 212, 141, 0.32);
      color: var(--success);
    }}
    .notice.working {{
      background: rgba(245, 165, 36, 0.10);
      border-color: rgba(245, 165, 36, 0.34);
      color: var(--warn);
    }}
    .progress-track {{
      height: 10px;
      margin-top: 10px;
      overflow: hidden;
      border-radius: 999px;
      background: rgba(255, 255, 255, 0.08);
    }}
    .progress-track span {{
      display: block;
      height: 100%;
      border-radius: inherit;
      background: linear-gradient(90deg, #10bfb1, #42e3d6);
      box-shadow: 0 0 18px rgba(22, 199, 186, 0.45);
      transition: width 0.25s ease;
    }}
    .progress-note {{
      margin-top: 8px;
      color: var(--muted);
      font-size: 13px;
      font-weight: 600;
    }}
    .notice.error {{
      background: rgba(255, 116, 108, 0.10);
      border-color: rgba(255, 116, 108, 0.34);
      color: var(--danger);
    }}
    form {{
      display: grid;
      gap: 16px;
    }}
    .field-grid {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 14px;
    }}
    label {{
      display: block;
      margin-bottom: 7px;
      font-weight: 800;
    }}
    select,
    input[type="file"] {{
      width: 100%;
      min-height: 44px;
      border: 1px solid var(--line);
      border-radius: 9px;
      background: #06121e;
      padding: 10px 12px;
      font: inherit;
      color: var(--text);
    }}
    input[type="file"]::file-selector-button {{
      margin-right: 12px;
      padding: 8px 12px;
      border: 1px solid rgba(22, 199, 186, 0.34);
      border-radius: 7px;
      color: #b8f8f2;
      background: rgba(22, 199, 186, 0.10);
      font: inherit;
      font-weight: 800;
      cursor: pointer;
    }}
    .hint {{
      margin-top: 7px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.55;
    }}
    .rule-panel {{
      display: grid;
      gap: 8px;
      padding: 12px 14px;
      border: 1px solid rgba(22, 199, 186, 0.30);
      border-radius: 10px;
      background: linear-gradient(135deg, rgba(22, 199, 186, 0.10), rgba(7, 21, 34, 0.75));
      color: var(--accent-dark);
      line-height: 1.55;
      font-size: 14px;
      font-weight: 700;
    }}
    .rule-panel span {{
      color: var(--muted);
      font-size: 13px;
      font-weight: 600;
    }}
    .master-panel {{
      margin-top: 18px;
      padding: 18px;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 12px;
      box-shadow: var(--shadow);
    }}
    .master-card {{
      display: grid;
      gap: 10px;
      padding: 14px;
      border: 1px solid rgba(22, 199, 186, 0.28);
      border-radius: 10px;
      background: linear-gradient(145deg, rgba(10, 27, 42, 0.98), rgba(5, 17, 28, 0.98));
    }}
    .master-top {{
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 12px;
    }}
    .master-title {{
      margin: 0;
      font-size: 17px;
      font-weight: 900;
    }}
    .master-actions {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      align-items: end;
      margin-top: 12px;
    }}
    .master-actions .upload-field {{
      flex: 1 1 320px;
    }}
    .status-board {{
      display: grid;
      grid-template-columns: repeat(6, minmax(0, 1fr));
      gap: 10px;
      margin-top: 12px;
    }}
    .owner-status {{
      display: grid;
      gap: 4px;
      min-height: 92px;
      padding: 12px;
      border-radius: 10px;
      border: 1px solid var(--line);
      background: rgba(8, 24, 38, 0.86);
    }}
    .owner-status strong {{
      font-size: 15px;
    }}
    .owner-status span {{
      display: inline-flex;
      width: fit-content;
      padding: 3px 7px;
      border-radius: 999px;
      font-size: 12px;
      font-weight: 900;
      color: var(--muted);
      background: rgba(145, 166, 180, 0.13);
      border: 1px solid rgba(145, 166, 180, 0.20);
    }}
    .owner-status small {{
      color: var(--muted);
      line-height: 1.35;
    }}
    .owner-status.running span {{
      color: var(--warn);
      background: rgba(245, 165, 36, 0.13);
    }}
    .owner-status.done span {{
      color: var(--success);
      background: rgba(53, 212, 141, 0.13);
    }}
    .owner-status.error span {{
      color: var(--danger);
      background: rgba(255, 116, 108, 0.13);
    }}
    .actions {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      align-items: center;
      padding-top: 4px;
    }}
    button,
    .button {{
      min-height: 44px;
      border: 0;
      border-radius: 9px;
      padding: 11px 16px;
      font: inherit;
      font-weight: 800;
      cursor: pointer;
    }}
    .primary {{
      color: #001713;
      background: linear-gradient(135deg, #35ded0, #12b8ac);
      box-shadow: 0 8px 24px rgba(22, 199, 186, 0.22);
    }}
    .secondary {{
      color: #b8f8f2;
      background: rgba(22, 199, 186, 0.08);
      border: 1px solid rgba(22, 199, 186, 0.36);
      text-decoration: none;
      display: inline-flex;
      align-items: center;
    }}
    .summary {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 12px;
      margin-top: 14px;
    }}
    .summary-item {{
      padding: 14px;
      box-shadow: none;
      background: linear-gradient(145deg, rgba(9, 25, 39, 0.96), rgba(5, 17, 28, 0.96));
    }}
    .summary-item strong {{
      display: block;
      margin-bottom: 5px;
      font-size: 15px;
    }}
    .summary-item span {{
      color: var(--muted);
      line-height: 1.55;
      font-size: 13px;
    }}
    .unmatched-panel {{
      margin-top: 18px;
      padding: 16px;
      border: 1px solid rgba(245, 165, 36, 0.34);
      border-radius: 10px;
      background: linear-gradient(145deg, rgba(245, 165, 36, 0.08), rgba(6, 18, 30, 0.94));
    }}
    .unmatched-head,
    .unmatched-panel.empty-list {{
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 16px;
    }}
    .unmatched-panel h3 {{
      margin: 0 0 5px;
      font-size: 17px;
    }}
    .unmatched-panel p {{
      margin: 0;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.55;
    }}
    .unmatched-count {{
      display: inline-flex;
      flex: 0 0 auto;
      padding: 5px 9px;
      border: 1px solid rgba(245, 165, 36, 0.36);
      border-radius: 999px;
      color: var(--warn);
      background: rgba(245, 165, 36, 0.12);
      font-size: 12px;
      font-weight: 900;
    }}
    .unmatched-count.ok {{
      color: var(--success);
      border-color: rgba(53, 212, 141, 0.34);
      background: rgba(53, 212, 141, 0.11);
    }}
    .month-totals {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(130px, 1fr));
      gap: 8px;
      margin-top: 14px;
    }}
    .month-total {{
      display: grid;
      gap: 4px;
      padding: 10px 12px;
      border: 1px solid rgba(22, 199, 186, 0.24);
      border-radius: 8px;
      background: rgba(22, 199, 186, 0.07);
    }}
    .month-total span {{
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
    }}
    .month-total strong {{
      color: #b8f8f2;
      font-size: 15px;
    }}
    .table-scroll {{
      max-height: 360px;
      margin-top: 14px;
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
    }}
    .unmatched-panel table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 520px;
      font-size: 13px;
    }}
    .unmatched-panel th,
    .unmatched-panel td {{
      padding: 10px 12px;
      border-bottom: 1px solid var(--line);
      text-align: left;
      vertical-align: top;
      line-height: 1.5;
    }}
    .unmatched-panel th {{
      position: sticky;
      top: 0;
      z-index: 1;
      color: #b8f8f2;
      background: #0a1b2a;
      font-weight: 900;
    }}
    .unmatched-panel tbody tr:last-child td {{
      border-bottom: 0;
    }}
    .unmatched-panel tbody tr:hover {{
      background: rgba(22, 199, 186, 0.05);
    }}
    @media (max-width: 820px) {{
      .topline,
      .panel-head {{
        display: block;
      }}
      .access {{
        margin-top: 10px;
        text-align: left;
      }}
      .layout,
      .field-grid,
      .summary,
      .status-board,
      .master-top {{
        grid-template-columns: 1fr;
        display: block;
      }}
      .rail {{
        position: static;
      }}
      .owner-list {{
        grid-template-columns: repeat(2, minmax(0, 1fr));
      }}
      .unmatched-head,
      .unmatched-panel.empty-list {{
        display: flex;
      }}
    }}
  </style>
</head>
<body>
  <main class="shell">
    <section class="topline">
      <div>
        <h1>销售排单回填工作台</h1>
        <p class="subtext">网站维护一份共用销售排单。每周先上传/替换本周排单，随后每位业务只上传自己的预测；首次上传只写入当前空白的“预估”数量/金额栏，同一业务重复上传时只覆盖该业务上一次由网站写入的预估格。匹配数据使用绿色字体，未取得预测数据写入黑色 0，所有回填格均无填充颜色，其他原始数据不修改。</p>
      </div>
      <div class="access">{html.escape(access_note)}<br><a href="/logout">安全退出</a></div>
    </section>

    <section class="master-panel" aria-label="共用销售排单">
      <div class="master-card">
        <div class="master-top">
          <div>
            <p class="master-title">共用销售排单</p>
            <p class="subtext">{html.escape(state_text)}</p>
            <p class="hint">{html.escape(detail_text)}</p>
          </div>
          <div class="badge {master_status_class}">{html.escape(master_status)}</div>
        </div>
        <form class="master-actions" method="post" action="/sales-master" enctype="multipart/form-data">
          <div class="upload-field">
            <label for="master_sales_file">上传/替换本周共用销售排单</label>
            <input id="master_sales_file" name="sales_file" type="file" accept=".xlsx,.xlsm" required>
            <div class="hint">只有共用排单需要上传一次；各业务后续只上传预测。替换排单会从新文件重新开始。</div>
          </div>
          <button class="primary" type="submit">保存共用排单</button>
          {download_latest_html}
          {preview_latest_html}
        </form>
        <div class="status-board" aria-label="业务回填状态">
          {owner_status_cards}
        </div>
      </div>
    </section>

    <section class="layout">
      <aside class="rail" aria-label="业务担当">
        <div class="rail-title">业务担当页面</div>
        <nav class="owner-list">
          {owner_nav}
        </nav>
      </aside>

      <section class="panel">
        <div class="panel-head">
          <div>
            <h2>{html.escape(selected_owner)} 的预测上传</h2>
            <p class="subtext">{html.escape(upload_hint)} 预测会写入上方共用销售排单，销售排单必须保留“业务担当”和“客户机种”。</p>
          </div>
          <div class="badge ready">登录保护已启用</div>
        </div>
        {message_html}
        {error_html}
        {storage_warning_html}
        {job_html}
        <form method="post" action="/generate" enctype="multipart/form-data">
          <div class="rule-panel">
            {html.escape(selected_owner_guide)}
            <span>系统以销售排单实际存在的预估栏月份为准；北京时间当月直接采用业务上传的剩余预估，不再扣减已完成数量。首次上传只写空白预估格；同一业务再次上传可覆盖其上一次由网站写入的预估数据，其他已有值和原始数据绝不覆盖。若上周数量为 0，本周金额写 0 并使用深红色背景提示无法计算单价。</span>
          </div>
          <div>
            <label for="business_owner">业务担当</label>
            <select id="business_owner" name="business_owner">
              {owner_options}
            </select>
            <div class="hint">只会回填销售排单中属于所选业务担当的行。</div>
          </div>
          <div>
            <label for="prediction_file">预测信息</label>
            <input id="prediction_file" name="prediction_file" type="file" accept=".xlsx,.jpg,.jpeg,.png" required>
            <div class="hint">支持 .xlsx、.jpg、.png；建议优先上传 .xlsx，清晰截图会按已适配规则识别。无需再上传销售排单。</div>
          </div>
          <div class="actions">
            <button class="primary" type="submit">开始回填</button>
            <a class="secondary button" href="{owner_link(selected_owner)}">刷新处理状态</a>
            {download_latest_html}
            {preview_latest_html}
          </div>
        </form>
        {unmatched_panel_html}
      </section>
    </section>

    <section class="summary" aria-label="处理规则">
      <div class="summary-item">
        <strong>分人匹配</strong>
        <span>按业务担当读取 New part NO、机种名、品名、子件描述前缀、模组型号或料号。</span>
      </div>
      <div class="summary-item">
        <strong>担当隔离</strong>
        <span>每次只处理所选业务担当的销售排单行；无预测数据的预估空白格写入 0。</span>
      </div>
      <div class="summary-item">
        <strong>当月直接采用</strong>
        <span>以北京时间判断当月；业务上传数量就是当月剩余预估，不再扣除之前已完成数量。</span>
      </div>
      <div class="summary-item">
        <strong>金额异常提示</strong>
        <span>上周数量为 0 时，本周金额写 0，并用深红色背景标记无法计算单价。</span>
      </div>
    </section>
  </main>
</body>
</html>
"""
    return page.encode("utf-8")


def write_html(handler: BaseHTTPRequestHandler, status: int, body: str, head: bool = False) -> None:
    encoded = body.encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "text/html; charset=utf-8")
    handler.send_header("Content-Length", str(len(encoded)))
    handler.send_header("Cache-Control", "no-store")
    handler.end_headers()
    if not head:
        handler.wfile.write(encoded)


def write_text(
    handler: BaseHTTPRequestHandler,
    status: int,
    body: str,
    content_type: str = "text/plain; charset=utf-8",
    head: bool = False,
) -> None:
    encoded = body.encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", content_type)
    handler.send_header("Content-Length", str(len(encoded)))
    handler.send_header("Cache-Control", "no-store")
    handler.end_headers()
    if not head:
        handler.wfile.write(encoded)


def write_json(handler: BaseHTTPRequestHandler, status: int, payload: dict, head: bool = False) -> None:
    encoded = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(encoded)))
    handler.send_header("Cache-Control", "no-store")
    handler.end_headers()
    if not head:
        handler.wfile.write(encoded)


def save_upload(field: cgi.FieldStorage, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    with dest.open("wb") as out:
        shutil.copyfileobj(field.file, out)


def validate_excel_file(path: Path, label: str) -> None:
    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        return
    if not path.exists() or path.stat().st_size <= 0:
        raise ValueError(f"{label}上传后为空文件，请重新选择原始 Excel 文件上传。")
    if not zipfile.is_zipfile(path):
        with path.open("rb") as f:
            head = f.read(16).hex(" ")
        raise ValueError(
            f"{label}不是标准 Excel 文件，网站无法读取。"
            f"请重新上传原始 .xlsx/.xlsm 文件（当前大小 {format_file_size(path)}，文件头 {head or '空'}）。"
        )
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        wb.close()
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"{label}无法被 Excel 解析：{exc}") from exc


def validate_prediction_file(path: Path, label: str) -> None:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        validate_excel_file(path, label)
        return

    if suffix in {".jpg", ".jpeg", ".png"}:
        if not path.exists() or path.stat().st_size <= 0:
            raise ValueError(f"{label}上传后为空文件，请重新选择原始图片上传。")
        with path.open("rb") as f:
            head = f.read(12)
        is_png = suffix == ".png" and head.startswith(b"\x89PNG\r\n\x1a\n")
        is_jpg = suffix in {".jpg", ".jpeg"} and head.startswith(b"\xff\xd8\xff")
        if not (is_png or is_jpg):
            raise ValueError(f"{label}不是标准 {suffix} 图片，请重新上传原始 jpg/png 截图。")
        return

    raise ValueError("预测信息请上传 .xlsx、.jpg 或 .png 文件。")


def first_upload(form: cgi.FieldStorage, name: str):
    field = form[name]
    if isinstance(field, list):
        return field[0]
    return field


def send_xlsx(handler: BaseHTTPRequestHandler, path: Path, download_name: str, head: bool = False) -> None:
    file_size = path.stat().st_size
    handler.send_response(200)
    handler.send_header("Content-Type", XLSX_MIME)
    handler.send_header("Content-Disposition", content_disposition(download_name))
    handler.send_header("Content-Length", str(file_size))
    handler.send_header("Cache-Control", "no-store")
    handler.end_headers()
    if not head:
        with path.open("rb") as f:
            shutil.copyfileobj(f, handler.wfile)


def parse_editable_forecast_value(raw_value: str):
    value = raw_value.strip()
    if value == "":
        return None
    if value.startswith("="):
        raise ValueError("预估栏只能填写数字，不能填写公式。")
    normalized = value.replace(",", "").replace("，", "")
    if re.fullmatch(r"-?\d+", normalized) and not re.match(r"-?0\d+", normalized):
        return int(normalized)
    if re.fullmatch(r"-?(?:\d+\.\d+|\d+\.|\.\d+)", normalized):
        return float(normalized)
    raise ValueError("预估栏只能填写数字，留空则清空该格。")


def parse_workbook_cell_value(raw_value: str):
    value = raw_value.replace("\r\n", "\n").replace("\r", "\n")
    if value == "":
        return None
    if value.startswith("'"):
        return value[1:]
    if value.startswith("="):
        return value

    normalized = value.replace(",", "").replace("，", "")
    if re.fullmatch(r"-?\d+", normalized) and not re.match(r"-?0\d+", normalized):
        return int(normalized)
    if re.fullmatch(r"-?(?:\d+\.\d+|\d+\.|\.\d+)", normalized):
        return float(normalized)

    upper = value.strip().upper()
    if upper == "TRUE":
        return True
    if upper == "FALSE":
        return False
    return value


def parse_column_ref(value: str, default: int = 1) -> int:
    raw = str(value or "").strip().upper()
    if not raw:
        return default
    if re.fullmatch(r"[A-Z]{1,3}", raw):
        return column_index_from_string(raw)
    if re.fullmatch(r"[1-9][0-9]{0,3}", raw):
        return int(raw)
    return default


def cell_fill_rgb(cell) -> Optional[tuple[int, int, int]]:
    color = getattr(getattr(cell, "fill", None), "fgColor", None)
    rgb = getattr(color, "rgb", None)
    if not rgb or not isinstance(rgb, str):
        return None
    value = rgb[-6:]
    if not re.fullmatch(r"[0-9A-Fa-f]{6}", value):
        return None
    return int(value[0:2], 16), int(value[2:4], 16), int(value[4:6], 16)


def is_green_header_cell(cell) -> bool:
    color = getattr(getattr(cell, "fill", None), "fgColor", None)
    if getattr(color, "type", None) == "theme":
        try:
            return int(color.theme) == 6
        except (TypeError, ValueError):
            return False
    rgb = cell_fill_rgb(cell)
    if rgb is None:
        return False
    red, green, blue = rgb
    return green >= 120 and green > red + 15 and green > blue + 15


def cell_has_visible_fill(cell) -> bool:
    color = getattr(getattr(cell, "fill", None), "fgColor", None)
    color_type = getattr(color, "type", None)
    if color_type == "theme":
        return True
    if color_type == "indexed":
        return True
    rgb = getattr(color, "rgb", None)
    return bool(isinstance(rgb, str) and rgb[-6:] not in {"000000", "00000000"})


def editable_pair_by_label(label: str) -> bool:
    text = str(label or "")
    if not any(word in text for word in ("预估", "预计")):
        return False
    if any(word in text for word in ("差异", "已完成")):
        return False
    return True


def pair_has_green_header(ws, pair) -> bool:
    cells = (
        ws.cell(3, pair.qty_col),
        ws.cell(3, pair.amt_col),
        ws.cell(4, pair.qty_col),
        ws.cell(4, pair.amt_col),
    )
    return any(is_green_header_cell(cell) for cell in cells)


def pair_has_visible_header_fill(ws, pair) -> bool:
    cells = (
        ws.cell(3, pair.qty_col),
        ws.cell(3, pair.amt_col),
        ws.cell(4, pair.qty_col),
        ws.cell(4, pair.amt_col),
    )
    return any(cell_has_visible_fill(cell) for cell in cells)


def editable_forecast_columns(wb, sheet_name: str) -> dict[int, str]:
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]

    # Read-only worksheets are very slow when ws.cell() is called repeatedly because
    # each random lookup can rescan worksheet XML. Read both header rows in one pass.
    scan_max_col = min(max(int(ws.max_column or 1), 1), 260)
    header_rows = list(ws.iter_rows(min_row=3, max_row=4, min_col=1, max_col=scan_max_col))
    if len(header_rows) < 2:
        return {}
    label_row, field_row = header_rows
    candidate_pairs = []
    for offset in range(scan_max_col - 1):
        qty_text = str(field_row[offset].value or "").strip()
        amt_text = str(field_row[offset + 1].value or "").strip()
        if "数量" not in qty_text or "金额" not in amt_text:
            continue
        label = str(label_row[offset].value or "").strip()
        if not editable_pair_by_label(label):
            continue
        qty_col = offset + 1
        candidate_pairs.append(
            {
                "label": label,
                "qty_col": qty_col,
                "amt_col": qty_col + 1,
                "cells": (
                    label_row[offset],
                    label_row[offset + 1],
                    field_row[offset],
                    field_row[offset + 1],
                ),
            }
        )

    green_pairs = [
        pair for pair in candidate_pairs if any(is_green_header_cell(cell) for cell in pair["cells"])
    ]
    selected_pairs = green_pairs

    if not selected_pairs:
        if any(
            any(cell_has_visible_fill(cell) for cell in pair["cells"])
            for pair in candidate_pairs
        ):
            return {}
        latest_by_month = {}
        for pair in candidate_pairs:
            month = extract_fill_target_month(sheet_name, pair["label"])
            if month is None:
                continue
            current = latest_by_month.get(month)
            if current is None or pair["qty_col"] > current["qty_col"]:
                latest_by_month[month] = pair
        selected_pairs = list(latest_by_month.values())

    allowed: dict[int, str] = {}
    for pair in selected_pairs:
        allowed[pair["qty_col"]] = f'{pair["label"]} 数量'
        allowed[pair["amt_col"]] = f'{pair["label"]} 金额'
    return allowed


def merged_follower_coords(ws, start_row: int, max_row: int, start_col: int, max_col: int) -> set[str]:
    blocked: set[str] = set()
    try:
        ranges = list(ws.merged_cells.ranges)
    except Exception:
        ranges = []
    for merged_range in ranges:
        min_col, min_row, end_col, end_row = merged_range.bounds
        if end_row < start_row or min_row > max_row or end_col < start_col or min_col > max_col:
            continue
        for row_idx in range(max(min_row, start_row), min(end_row, max_row) + 1):
            for col_idx in range(max(min_col, start_col), min(end_col, max_col) + 1):
                if row_idx == min_row and col_idx == min_col:
                    continue
                blocked.add(f"{get_column_letter(col_idx)}{row_idx}")
    return blocked


def render_preview_page(
    message: str = "",
    error: str = "",
    selected_sheet: str = "",
    start_row: int = 1,
    start_col: int = 1,
    rows_limit: int = 80,
    cols_limit: int = 140,
    full_view: bool = False,
) -> bytes:
    FULL_VIEW_ROW_CAP = 120
    FULL_VIEW_COL_CAP = 80
    if not master_sales_exists():
        return render_page(error="还没有共用销售排单可预览，请先上传本周排单。")

    metadata = ensure_metadata_schema(load_metadata())
    message_html = f'<div class="notice success">{html.escape(message)}</div>' if message else ""
    error_html = f'<div class="notice error">{html.escape(error)}</div>' if error else ""
    start_row = max(1, start_row)
    start_col = max(1, start_col)
    rows_limit = max(20, min(rows_limit, 200))
    cols_limit = max(40, min(cols_limit, 260))

    try:
        wb = load_workbook(MASTER_SALES_PATH, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return render_page(error=f"在线预览失败：当前共用销售排单无法读取（{exc}）。请先下载当前最新版确认文件，或重新上传本周共用排单。")

    try:
        sheet_names = wb.sheetnames
        selected_sheet = resolve_sheet_name(selected_sheet, sheet_names)
        if selected_sheet not in sheet_names:
            selected_sheet = sheet_names[0]
        ws = wb[selected_sheet]
        sheet_max_row = max(int(ws.max_row or 1), 1)
        sheet_max_col = max(int(ws.max_column or 1), 1)
        if full_view:
            start_row = min(start_row, sheet_max_row)
            start_col = min(start_col, sheet_max_col)
            rows_limit = min(max(rows_limit, 20), FULL_VIEW_ROW_CAP)
            cols_limit = min(max(cols_limit, 40), FULL_VIEW_COL_CAP)
            max_row = min(sheet_max_row, start_row + rows_limit - 1)
            max_col = min(sheet_max_col, start_col + cols_limit - 1)
        else:
            start_row = min(start_row, sheet_max_row)
            start_col = min(start_col, sheet_max_col)
            max_row = min(sheet_max_row, start_row + rows_limit - 1)
            max_col = min(sheet_max_col, start_col + cols_limit - 1)
        sheet_tab_html = "\n".join(
            (
                f'<a class="sheet-tab {"active" if name == selected_sheet else ""}" '
                f'href="/table-preview?sheet={urllib.parse.quote(name)}&start_row={start_row}&start_col={get_column_letter(start_col)}'
                f'&rows={rows_limit}&cols={cols_limit}{"&full=1" if full_view else ""}">'
                f'{html.escape(name)}</a>'
            )
            for name in sheet_names
        )
        sheet_options = "\n".join(
            f'<option value="{html.escape(name)}" {"selected" if name == selected_sheet else ""}>{html.escape(name)}</option>'
            for name in sheet_names
        )
        header_cells = "".join(f"<th>{get_column_letter(col)}</th>" for col in range(start_col, max_col + 1))
        body_rows = []
        for row_idx, row_values in enumerate(
            ws.iter_rows(min_row=start_row, max_row=max_row, min_col=start_col, max_col=max_col, values_only=False),
            start=start_row,
        ):
            cells = [f"<th>{row_idx}</th>"]
            for col_idx, cell in enumerate(row_values, start=start_col):
                value = cell.value
                display = "" if value is None else str(value)
                if len(display) > 80:
                    display = display[:77] + "..."
                coord = f"{get_column_letter(col_idx)}{row_idx}"
                fill_css = fill_css_from_cell(cell)
                style_attr = f' style="background:{fill_css}"' if fill_css else ""
                cells.append(f'<td{style_attr} title="{html.escape(coord)}">{html.escape(display)}</td>')
            body_rows.append("<tr>" + "".join(cells) + "</tr>")
        table_html = "\n".join(body_rows)
    finally:
        wb.close()

    latest_name = metadata.get("latest_name") or metadata.get("master_name") or "当前最新版"
    page = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>网站内打开表格</title>
  <style>
    :root {{
      --paper: #f6f7f4;
      --panel: #ffffff;
      --line: #d8ded5;
      --text: #17211b;
      --muted: #5f6b63;
      --accent: #136f63;
      --danger: #b42318;
      --success: #126b45;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      color: var(--text);
      font-family: "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", -apple-system, BlinkMacSystemFont, sans-serif;
      background: var(--paper);
    }}
    .shell {{ width: min(1380px, calc(100% - 28px)); margin: 0 auto; padding: 22px 0 40px; }}
    .top {{ display: flex; justify-content: space-between; gap: 12px; align-items: flex-start; margin-bottom: 16px; }}
    h1 {{ margin: 0 0 6px; font-size: 24px; }}
    .subtext {{ margin: 0; color: var(--muted); line-height: 1.6; }}
    .panel {{ background: var(--panel); border: 1px solid var(--line); border-radius: 8px; padding: 16px; margin-bottom: 16px; }}
    .notice {{ margin: 0 0 14px; padding: 12px 14px; border-radius: 8px; line-height: 1.6; font-weight: 800; }}
    .notice.success {{ color: var(--success); background: #e8f5ee; border: 1px solid #bbdec9; }}
    .notice.error {{ color: var(--danger); background: #fff0ee; border: 1px solid #f3c3bd; }}
    form {{ display: flex; flex-wrap: wrap; gap: 10px; align-items: end; }}
    label {{ display: grid; gap: 6px; font-weight: 800; }}
    input, select {{ min-height: 40px; border: 1px solid var(--line); border-radius: 7px; padding: 8px 10px; font: inherit; }}
    button, .button {{ min-height: 40px; border: 0; border-radius: 7px; padding: 9px 14px; font: inherit; font-weight: 800; text-decoration: none; cursor: pointer; }}
    .primary {{ color: #fff; background: var(--accent); }}
    .secondary {{ color: #0c4f47; background: #e9f3ee; border: 1px solid #bdd9cd; display: inline-flex; align-items: center; }}
    .table-wrap {{ overflow: auto; max-height: 72vh; border: 1px solid var(--line); border-radius: 8px; background: #fff; }}
    table {{ border-collapse: collapse; min-width: 100%; font-size: 12px; }}
    th, td {{ border: 1px solid #d7ddd4; padding: 5px 7px; white-space: nowrap; max-width: 220px; overflow: hidden; text-overflow: ellipsis; }}
    th {{ position: sticky; top: 0; background: #e8efe9; z-index: 1; }}
    tr th:first-child {{ position: sticky; left: 0; z-index: 2; background: #e8efe9; }}
    .sheet-tabs {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-bottom: 10px;
    }}
    .sheet-tab {{
      display: inline-flex;
      align-items: center;
      min-height: 32px;
      padding: 6px 10px;
      border-radius: 999px;
      border: 1px solid #bdd9cd;
      background: #eef6f1;
      color: #0c4f47;
      text-decoration: none;
      font-weight: 800;
    }}
    .sheet-tab.active {{
      background: #136f63;
      color: #fff;
      border-color: #136f63;
    }}
  </style>
</head>
<body>
  <main class="shell">
    <div class="top">
      <div>
        <h1>网站内打开当前表格</h1>
        <p class="subtext">当前文件：{html.escape(latest_name)}。这里直接在网站里打开你上传的共用销售排单，按 Sheet 查看，不进入 xlsx 编辑器。</p>
      </div>
      <a class="secondary button" href="/">返回上传页</a>
    </div>
    {message_html}
    {error_html}
    <section class="panel">
      <form method="get" action="/table-preview">
        <label>选择 Sheet
          <select name="sheet">{sheet_options}</select>
        </label>
        <label>起始行
          <input name="start_row" value="{start_row}" inputmode="numeric">
        </label>
        <label>起始列
          <input name="start_col" value="{html.escape(get_column_letter(start_col))}" placeholder="例如 BQ">
        </label>
        <label>预览行数
          <input name="rows" value="{rows_limit}" inputmode="numeric">
        </label>
        <label>预览列数
          <input name="cols" value="{cols_limit}" inputmode="numeric">
        </label>
        <input type="hidden" name="full" value="{1 if full_view else 0}">
        <button class="primary" type="submit">刷新预览</button>
        <a class="secondary button" href="/download/latest">下载当前最新版</a>
        <a class="secondary button" href="/table-preview?sheet={urllib.parse.quote(selected_sheet)}&full=1">整表模式</a>
        <a class="secondary button" href="/table-preview?sheet={urllib.parse.quote(selected_sheet)}&start_row=1&start_col=A&rows=80&cols=140">窗口模式</a>
      </form>
    </section>
    <section class="panel">
      <div class="sheet-tabs">{sheet_tab_html}</div>
      <p class="subtext">{f"当前为整表浏览模式，已打开 {min(rows_limit, FULL_VIEW_ROW_CAP)} 行 × {min(cols_limit, FULL_VIEW_COL_CAP)} 列的窗口，可继续用起始行/列向后翻。" if full_view else "当前为窗口浏览模式；如果要像整张工作簿一样查看，请点击“整表模式”。"} 颜色和内容都按上传表格直接展示。</p>
    </section>
    <div class="table-wrap">
      <table>
        <thead><tr><th>#</th>{header_cells}</tr></thead>
        <tbody>{table_html}</tbody>
      </table>
    </div>
  </main>
</body>
</html>
"""
    return page.encode("utf-8")


def fill_css_from_cell(cell) -> str:
    rgb = cell_fill_rgb(cell)
    if rgb is None:
        return ""
    red, green, blue = rgb
    if red == green == blue == 255:
        return ""
    return f"#{red:02x}{green:02x}{blue:02x}"


def onlyoffice_document_server_url() -> str:
    # OnlyOffice must point to a real, publicly reachable Document Server.
    # The public ONLYOFFICE website is not an editor endpoint.
    return normalize_base_url(os.environ.get("ONLYOFFICE_DOCUMENT_SERVER_URL", ""))


def onlyoffice_jwt_secret() -> str:
    return os.environ.get("ONLYOFFICE_JWT_SECRET", "").strip()


def onlyoffice_config_token(config: dict, secret: str) -> str:
    def encode(value: bytes) -> bytes:
        return base64.urlsafe_b64encode(value).rstrip(b"=")

    header = encode(json.dumps({"alg": "HS256", "typ": "JWT"}, separators=(",", ":")).encode("utf-8"))
    payload = encode(json.dumps(config, ensure_ascii=False, separators=(",", ":")).encode("utf-8"))
    signing_input = header + b"." + payload
    signature = encode(hmac.new(secret.encode("utf-8"), signing_input, hashlib.sha256).digest())
    return b".".join((header, payload, signature)).decode("ascii")


def onlyoffice_public_base_url(handler: Optional[BaseHTTPRequestHandler] = None) -> Optional[str]:
    return get_public_share_url(handler)


def onlyoffice_document_key() -> str:
    metadata = ensure_metadata_schema(load_metadata()) if METADATA_PATH.exists() else {}
    parts = [MASTER_SALES_PATH.name]
    if MASTER_SALES_PATH.exists():
        stat = MASTER_SALES_PATH.stat()
        parts.extend([str(stat.st_size), str(stat.st_mtime_ns)])
    parts.extend(
        [
            str(metadata.get("latest_name", "")),
            str(metadata.get("last_generated_at", "")),
            str(metadata.get("last_owner", "")),
        ]
    )
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


def onlyoffice_document_url(handler: Optional[BaseHTTPRequestHandler] = None) -> str:
    public_base = onlyoffice_public_base_url(handler)
    return f"{public_base}/download/latest" if public_base else ""


def onlyoffice_callback_url(handler: Optional[BaseHTTPRequestHandler] = None) -> str:
    public_base = onlyoffice_public_base_url(handler)
    return f"{public_base}/api/onlyoffice/callback" if public_base else ""


def render_onlyoffice_page(
    message: str = "",
    error: str = "",
    handler: Optional[BaseHTTPRequestHandler] = None,
) -> bytes:
    if not master_sales_exists():
        return render_page(error="还没有共用销售排单可打开，请先上传本周排单。", handler=handler)

    public_base = onlyoffice_public_base_url(handler)
    if not public_base:
        return render_page(
            error="OnlyOffice 需要一个公网可访问的网址才能打开和保存。请用 Render 的正式域名访问该页面。",
            handler=handler,
        )

    docs_server = onlyoffice_document_server_url()
    if not docs_server:
        return render_page(error="未配置 OnlyOffice 文档服务器地址。", handler=handler)

    metadata = ensure_metadata_schema(load_metadata())
    latest_name = metadata.get("latest_name") or metadata.get("master_name") or "当前最新版.xlsx"
    doc_url = onlyoffice_document_url(handler)
    callback_url = onlyoffice_callback_url(handler)
    document_key = onlyoffice_document_key()
    config = {
        "documentType": "cell",
        "document": {
            "fileType": "xlsx",
            "key": document_key,
            "title": latest_name,
            "url": doc_url,
        },
        "editorConfig": {
            "mode": "edit",
            "lang": "zh-CN",
            "callbackUrl": callback_url,
            "user": {
                "id": "shared-sales-workbook",
                "name": "共享排单",
            },
        },
        "permissions": {
            "comment": False,
            "download": True,
            "edit": True,
            "print": True,
            "review": False,
            "fillForms": True,
            "copy": True,
        },
        "customization": {
            "autosave": True,
            "forcesave": True,
            "toolbarNoTabs": False,
            "compactToolbar": False,
        },
    }
    jwt_secret = onlyoffice_jwt_secret()
    if jwt_secret:
        config["token"] = onlyoffice_config_token(config, jwt_secret)
    editor_config_json = json.dumps(config, ensure_ascii=False)
    docs_js_url = f"{docs_server}/web-apps/apps/api/documents/api.js"
    page = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>网站内直接打开 xlsx</title>
  <style>
    :root {{
      --paper: #f3f6f1;
      --panel: #ffffff;
      --line: #d7e0d8;
      --text: #17211b;
      --muted: #5e6b63;
      --accent: #136f63;
      --accent-dark: #0c4f47;
      --danger: #b42318;
      --shadow: 0 16px 34px rgba(28, 43, 34, 0.08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      color: var(--text);
      font-family: "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", -apple-system, BlinkMacSystemFont, sans-serif;
      background:
        linear-gradient(90deg, rgba(19, 111, 99, 0.06) 1px, transparent 1px),
        linear-gradient(180deg, rgba(138, 88, 0, 0.05) 1px, transparent 1px),
        var(--paper);
      background-size: 30px 30px;
    }}
    .shell {{
      width: min(1600px, calc(100% - 28px));
      margin: 0 auto;
      padding: 18px 0 24px;
    }}
    .top {{
      display: flex;
      justify-content: space-between;
      gap: 16px;
      align-items: flex-start;
      margin-bottom: 14px;
      padding-bottom: 12px;
      border-bottom: 1px solid var(--line);
    }}
    h1 {{
      margin: 0 0 6px;
      font-size: 26px;
      line-height: 1.2;
    }}
    .subtext {{
      margin: 0;
      color: var(--muted);
      line-height: 1.6;
      font-size: 14px;
    }}
    .meta {{
      max-width: 420px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.6;
      text-align: right;
    }}
    .notice {{
      margin: 0 0 12px;
      padding: 12px 14px;
      border-radius: 8px;
      line-height: 1.6;
      font-weight: 800;
    }}
    .notice.success {{
      color: var(--accent);
      background: #e7f4ef;
      border: 1px solid #bcdccd;
    }}
    .notice.error {{
      color: var(--danger);
      background: #fff0ee;
      border: 1px solid #f3c3bd;
    }}
    .editor-shell {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 12px;
      box-shadow: var(--shadow);
      overflow: hidden;
    }}
    #onlyoffice-editor {{
      width: 100%;
      height: calc(100vh - 220px);
      min-height: 760px;
      background: #fff;
    }}
    .hint {{
      margin-top: 12px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.6;
    }}
  </style>
</head>
<body>
  <main class="shell">
    <section class="top">
      <div>
        <h1>网站内直接打开 xlsx</h1>
        <p class="subtext">这里直接打开并编辑原始 xlsx 文件。页面里看到的就是在线表格，保存后会自动回写到当前共用排单，不会转换成别的格式。</p>
      </div>
      <div class="meta">
        <div>当前文件：{html.escape(latest_name)}</div>
        <div>文件地址：{html.escape(doc_url)}</div>
        <div>保存回调：{html.escape(callback_url)}</div>
      </div>
    </section>

    {f'<div class="notice success">{html.escape(message)}</div>' if message else ""}
    {f'<div class="notice error">{html.escape(error)}</div>' if error else ""}

    <section class="editor-shell">
      <div id="onlyoffice-editor"></div>
    </section>

    <div class="hint">如果在线表格暂时没有显示，请确认当前网站是通过 Render 的公网地址打开，而不是本机 127.0.0.1。也可以先用 <a href="/table-preview">网页浏览模式</a> 查看当前表格。</div>
  </main>
  <script>
    const config = {editor_config_json};
    const docsScriptUrl = {json.dumps(docs_js_url, ensure_ascii=False)};
    const editorContainerId = "onlyoffice-editor";

    function showError(text) {{
      const host = document.getElementById(editorContainerId);
      if (!host) return;
      const box = document.createElement("div");
      box.style.padding = "24px";
      box.style.color = "#b42318";
      box.style.fontWeight = "700";
      box.style.lineHeight = "1.6";
      box.textContent = String(text);
      host.replaceChildren(box);
    }}

    const script = document.createElement("script");
    script.src = docsScriptUrl;
    script.onload = () => {{
      if (!window.DocsAPI || !window.DocsAPI.DocEditor) {{
        showError("OnlyOffice 编辑器脚本已加载，但未找到 DocEditor。");
        return;
      }}
      try {{
        new window.DocsAPI.DocEditor(editorContainerId, config);
      }} catch (err) {{
        console.error(err);
        showError(err && err.message ? err.message : "OnlyOffice 编辑器初始化失败。");
      }}
    }};
    script.onerror = () => {{
      showError("无法加载 OnlyOffice 编辑器脚本，请检查文档服务器地址是否正确。");
    }};
    document.head.appendChild(script);
  </script>
</body>
</html>
"""
    return page.encode("utf-8")


def editor_window_payload(
    wb,
    sheet_name: str,
    start_row: int,
    start_col: int,
    rows_limit: int,
    cols_limit: int,
) -> dict:
    sheet_name = resolve_sheet_name(sheet_name, list(wb.sheetnames))
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"没有找到 Sheet：{sheet_name}")

    ws = wb[sheet_name]
    sheet_max_row = max(int(ws.max_row or 1), 1)
    sheet_max_col = max(int(ws.max_column or 1), 1)
    start_row = max(1, min(start_row, sheet_max_row))
    start_col = max(1, min(start_col, sheet_max_col))
    rows_limit = max(1, min(rows_limit, 200))
    cols_limit = max(1, min(cols_limit, 120))
    max_row = min(sheet_max_row, start_row + rows_limit - 1)
    max_col = min(sheet_max_col, start_col + cols_limit - 1)
    col_headers = [get_column_letter(col_idx) for col_idx in range(start_col, max_col + 1)]
    blocked_coords = merged_follower_coords(ws, start_row, max_row, start_col, max_col)

    rows: list[list[dict]] = []
    for row_idx, row_values in enumerate(
        ws.iter_rows(min_row=start_row, max_row=max_row, min_col=start_col, max_col=max_col, values_only=False),
        start=start_row,
    ):
        row_cells: list[dict] = []
        for col_idx, cell in enumerate(row_values, start=start_col):
            value = cell.value
            row_cells.append(
                {
                    "coord": f"{get_column_letter(col_idx)}{row_idx}",
                    "row": row_idx,
                    "col": col_idx,
                    "text": "" if value is None else str(value),
                    "editable": f"{get_column_letter(col_idx)}{row_idx}" not in blocked_coords,
                    "fill": fill_css_from_cell(cell),
                    "label": "",
                }
            )
        rows.append(row_cells)

    return {
        "sheet": sheet_name,
        "start_row": start_row,
        "start_col": start_col,
        "start_col_letter": get_column_letter(start_col),
        "rows_limit": rows_limit,
        "cols_limit": cols_limit,
        "sheet_max_row": sheet_max_row,
        "sheet_max_col": sheet_max_col,
        "max_row": max_row,
        "max_col": max_col,
        "col_headers": col_headers,
        "rows": rows,
    }


def render_workbook_editor_page(
    message: str = "",
    error: str = "",
    selected_sheet: str = "",
    start_row: int = 1,
    start_col: int = 1,
    rows_limit: int = 80,
    cols_limit: int = 40,
    full_view: bool = False,
) -> bytes:
    if not master_sales_exists():
        return render_page(error="还没有共用销售排单可预览，请先上传本周排单。")

    metadata = ensure_metadata_schema(load_metadata())
    message_html = f'<div class="notice success">{html.escape(message)}</div>' if message else ""
    error_html = f'<div class="notice error">{html.escape(error)}</div>' if error else ""
    start_row = max(1, start_row)
    start_col = max(1, start_col)
    if full_view:
        rows_limit = max(20, min(rows_limit, 120))
        cols_limit = max(20, min(cols_limit, 80))
    else:
        rows_limit = max(20, min(rows_limit, 80))
        cols_limit = max(20, min(cols_limit, 40))

    try:
        wb = load_workbook(MASTER_SALES_PATH, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        selected_sheet = resolve_sheet_name(selected_sheet, sheet_names)
        if selected_sheet not in sheet_names:
            selected_sheet = sheet_names[0]
        selected_ws = wb[selected_sheet]
        sheet_max_row = max(int(selected_ws.max_row or 1), 1)
        sheet_max_col = max(int(selected_ws.max_column or 1), 1)
        start_row = min(start_row, sheet_max_row)
        start_col = min(start_col, sheet_max_col)
        wb.close()
    except Exception as exc:  # noqa: BLE001
        return render_page(error=f"在线编辑器无法读取当前共用销售排单（{exc}）。请先下载当前最新版确认文件，或重新上传本周共用排单。")

    latest_name = metadata.get("latest_name") or metadata.get("master_name") or "当前最新版"
    sheet_names_json = json.dumps(sheet_names, ensure_ascii=False)
    current_sheet_json = json.dumps(selected_sheet, ensure_ascii=False)
    initial_state_json = json.dumps(
        {
            "sheet": selected_sheet,
            "start_row": start_row,
            "start_col": start_col,
            "rows": rows_limit,
            "cols": cols_limit,
            "full": bool(full_view),
        },
        ensure_ascii=False,
    )
    page = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>在线编辑工作簿</title>
  <style>
    :root {{
      --paper: #f6f7f4;
      --panel: #ffffff;
      --line: #d8ded5;
      --text: #17211b;
      --muted: #5f6b63;
      --accent: #136f63;
      --accent-dark: #0c4f47;
      --danger: #b42318;
      --success: #126b45;
      --warn: #8a5800;
      --shadow: 0 16px 34px rgba(28, 43, 34, 0.08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      color: var(--text);
      font-family: "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", -apple-system, BlinkMacSystemFont, sans-serif;
      background:
        linear-gradient(90deg, rgba(19, 111, 99, 0.06) 1px, transparent 1px),
        linear-gradient(180deg, rgba(138, 88, 0, 0.05) 1px, transparent 1px),
        var(--paper);
      background-size: 28px 28px;
    }}
    .shell {{
      width: min(1500px, calc(100% - 24px));
      margin: 0 auto;
      padding: 18px 0 28px;
    }}
    .top {{
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
      gap: 12px;
      margin-bottom: 14px;
      padding-bottom: 12px;
      border-bottom: 1px solid var(--line);
    }}
    h1 {{
      margin: 0 0 6px;
      font-size: 25px;
      line-height: 1.25;
    }}
    .subtext {{
      margin: 0;
      color: var(--muted);
      line-height: 1.6;
      font-size: 14px;
    }}
    .access {{
      max-width: 380px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.6;
      text-align: right;
    }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      padding: 14px;
      margin-bottom: 14px;
    }}
    .notice {{
      margin: 0 0 12px;
      padding: 11px 13px;
      border-radius: 8px;
      line-height: 1.6;
      font-weight: 800;
    }}
    .notice.success {{
      color: var(--success);
      background: #e8f5ee;
      border: 1px solid #bbdec9;
    }}
    .notice.error {{
      color: var(--danger);
      background: #fff0ee;
      border: 1px solid #f3c3bd;
    }}
    .toolbar {{
      display: grid;
      grid-template-columns: 1fr auto;
      gap: 12px;
      align-items: start;
    }}
    .sheet-tabs {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
    }}
    .sheet-tab {{
      display: inline-flex;
      align-items: center;
      min-height: 32px;
      padding: 6px 10px;
      border-radius: 999px;
      border: 1px solid #bdd9cd;
      background: #eef6f1;
      color: #0c4f47;
      text-decoration: none;
      font-weight: 800;
      cursor: pointer;
    }}
    .sheet-tab.active {{
      background: #136f63;
      color: #fff;
      border-color: #136f63;
    }}
    .controls {{
      display: grid;
      grid-template-columns: repeat(6, auto);
      gap: 8px;
      align-items: end;
      justify-content: end;
    }}
    .wide-control {{
      min-width: 132px;
    }}
    label {{
      display: grid;
      gap: 5px;
      font-weight: 800;
      font-size: 13px;
    }}
    input, select, button, .button {{
      min-height: 38px;
      border-radius: 7px;
      border: 1px solid var(--line);
      padding: 8px 10px;
      font: inherit;
    }}
    button, .button {{
      border: 0;
      cursor: pointer;
      font-weight: 800;
      text-decoration: none;
      display: inline-flex;
      align-items: center;
      justify-content: center;
    }}
    .primary {{
      background: var(--accent);
      color: #fff;
    }}
    .secondary {{
      background: #e9f3ee;
      color: var(--accent-dark);
      border: 1px solid #bdd9cd;
    }}
    .meta {{
      display: flex;
      flex-wrap: wrap;
      gap: 12px;
      margin-top: 10px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.5;
    }}
    .meta strong {{
      color: var(--text);
    }}
    .grid-wrap {{
      overflow: auto;
      max-height: 74vh;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #fff;
    }}
    table {{
      border-collapse: collapse;
      min-width: 100%;
      font-size: 12px;
    }}
    th, td {{
      border: 1px solid #d7ddd4;
      padding: 0;
      white-space: nowrap;
      min-width: 78px;
      max-width: 240px;
      vertical-align: middle;
    }}
    thead th {{
      position: sticky;
      top: 0;
      background: #e8efe9;
      z-index: 2;
      padding: 6px 8px;
      text-align: center;
    }}
    tbody th {{
      position: sticky;
      left: 0;
      z-index: 1;
      background: #e8efe9;
      padding: 6px 8px;
      text-align: center;
      min-width: 56px;
    }}
    .corner {{
      position: sticky;
      top: 0;
      left: 0;
      z-index: 3;
    }}
    .cell {{
      display: block;
      width: 100%;
      min-height: 30px;
      padding: 6px 8px;
      border: 0;
      background: transparent;
      font: inherit;
      color: inherit;
      outline: none;
    }}
    .editable {{
      background: transparent;
    }}
    .locked {{
      background: rgba(23, 33, 27, 0.03);
    }}
    .dirty {{
      box-shadow: inset 0 0 0 2px rgba(19, 111, 99, 0.35);
    }}
    .selected {{
      outline: 2px solid #136f63;
      outline-offset: -2px;
    }}
    .cell-text {{
      display: block;
      min-height: 30px;
      padding: 6px 8px;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .hint {{
      margin-top: 8px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.55;
    }}
    .formula-bar {{
      display: grid;
      grid-template-columns: 120px 1fr auto;
      gap: 8px;
      align-items: end;
      margin-top: 12px;
    }}
    .formula-bar input {{
      width: 100%;
    }}
    .statusline {{
      margin-top: 10px;
      color: var(--muted);
      font-size: 13px;
    }}
    @media (max-width: 1024px) {{
      .toolbar, .controls {{
        grid-template-columns: 1fr;
      }}
      .access {{
        text-align: left;
      }}
    }}
  </style>
</head>
<body>
  <main class="shell">
    <section class="top">
      <div>
        <h1>在线编辑工作簿</h1>
        <p class="subtext">当前文件：{html.escape(latest_name)}。这个页面直接读取并回写同一个 xlsx 工作簿，你可以像打开表格一样切 Sheet、跳行列、定位单元格、整本编辑并保存，不会转换成其他文件格式。</p>
      </div>
      <div class="access">{html.escape(storage_description())}</div>
    </section>

    {message_html}
    {error_html}

    <section class="panel">
      <div class="toolbar">
        <div>
          <div id="sheet-tabs" class="sheet-tabs"></div>
          <div class="meta">
            <div>文件：<strong>{html.escape(latest_name)}</strong></div>
            <div>当前 Sheet：<strong id="sheet-name">{html.escape(selected_sheet)}</strong></div>
            <div>状态：<strong id="save-state">未修改</strong></div>
          </div>
        </div>
        <div class="controls">
          <label>起始行 <input id="start-row" type="number" min="1" value="{start_row}"></label>
          <label>起始列 <input id="start-col" type="text" value="{html.escape(get_column_letter(start_col))}"></label>
          <label>行数 <input id="rows-limit" type="number" min="1" max="200" value="{rows_limit}"></label>
          <label>列数 <input id="cols-limit" type="number" min="1" max="120" value="{cols_limit}"></label>
          <label class="wide-control">定位单元格 <input id="goto-cell" type="text" placeholder="例如 BX27"></label>
          <div style="display:flex; gap:8px; align-items:end;">
            <button id="goto-btn" class="secondary" type="button">定位</button>
            <button id="refresh-btn" class="secondary" type="button">刷新窗口</button>
            <button id="save-btn" class="primary" type="button">保存修改</button>
          </div>
        </div>
      </div>
      <div class="formula-bar">
        <label>当前单元格 <input id="selected-cell" type="text" readonly></label>
        <label>编辑栏 <input id="formula-bar" type="text" placeholder="这里可直接编辑当前单元格，公式请以 = 开头"></label>
        <button id="apply-formula-btn" class="secondary" type="button">应用到当前格</button>
      </div>
      <div class="hint">整个工作簿都会以 xlsx 原样读取并写回；公式请以 <code>=</code> 开头，若要强制输入文本可在最前面加英文单引号 <code>'</code>。合并单元格中除左上角外的从属格会保持只读。</div>
    </section>

    <section class="panel">
      <div id="grid" class="grid-wrap"></div>
      <div id="statusline" class="statusline">正在加载工作簿窗口…</div>
    </section>
  </main>
  <script>
    const SHEET_NAMES = {sheet_names_json};
    const INITIAL = {initial_state_json};
    const DEFAULT_SHEET = {current_sheet_json};
    const dirtyCells = new Map();
    let currentSheet = INITIAL.sheet || DEFAULT_SHEET;
    let currentWindow = {{
      sheet: currentSheet,
      start_row: INITIAL.start_row,
      start_col: INITIAL.start_col,
      rows: INITIAL.rows,
      cols: INITIAL.cols,
    }};
    let currentPayload = null;
    let selectedCoord = "";

    const grid = document.getElementById("grid");
    const tabsEl = document.getElementById("sheet-tabs");
    const statusEl = document.getElementById("statusline");
    const saveStateEl = document.getElementById("save-state");
    const sheetNameEl = document.getElementById("sheet-name");
    const startRowEl = document.getElementById("start-row");
    const startColEl = document.getElementById("start-col");
    const rowsLimitEl = document.getElementById("rows-limit");
    const colsLimitEl = document.getElementById("cols-limit");
    const gotoCellEl = document.getElementById("goto-cell");
    const gotoBtn = document.getElementById("goto-btn");
    const refreshBtn = document.getElementById("refresh-btn");
    const saveBtn = document.getElementById("save-btn");
    const selectedCellEl = document.getElementById("selected-cell");
    const formulaBarEl = document.getElementById("formula-bar");
    const applyFormulaBtn = document.getElementById("apply-formula-btn");

    function setStatus(text, kind = "") {{
      statusEl.textContent = text;
      statusEl.dataset.kind = kind;
    }}

    function updateSaveState() {{
      const count = dirtyCells.size;
      saveStateEl.textContent = count ? `有 ${{count}} 个单元格未保存` : "未修改";
      saveBtn.disabled = !count;
    }}

    function makeTab(sheet) {{
      const a = document.createElement("button");
      a.type = "button";
      a.className = "sheet-tab" + (sheet === currentSheet ? " active" : "");
      a.textContent = sheet;
      a.addEventListener("click", () => {{
        currentSheet = sheet;
        loadWindow({{ resetSelection: true }});
      }});
      return a;
    }}

    function renderTabs() {{
      tabsEl.innerHTML = "";
      SHEET_NAMES.forEach((sheet) => tabsEl.appendChild(makeTab(sheet)));
    }}

    function currentParams() {{
      return {{
        sheet: currentSheet,
        start_row: Number(startRowEl.value || currentWindow.start_row || 1),
        start_col: startColEl.value || "A",
        rows: Number(rowsLimitEl.value || currentWindow.rows || 80),
        cols: Number(colsLimitEl.value || currentWindow.cols || 40),
      }};
    }}

    function escapeHtml(text) {{
      return String(text ?? "")
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#39;");
    }}

    function columnToNumber(value) {{
      const raw = String(value || "").trim().toUpperCase();
      if (!/^[A-Z]{{1,3}}$/.test(raw)) return 1;
      let total = 0;
      for (const ch of raw) {{
        total = total * 26 + (ch.charCodeAt(0) - 64);
      }}
      return total;
    }}

    function numberToColumn(value) {{
      let num = Math.max(1, Number(value || 1));
      let text = "";
      while (num > 0) {{
        const remain = (num - 1) % 26;
        text = String.fromCharCode(65 + remain) + text;
        num = Math.floor((num - 1) / 26);
      }}
      return text;
    }}

    function normalizeCoord(value) {{
      const raw = String(value || "").trim().toUpperCase().replace(/\\s+/g, "");
      const match = raw.match(/^([A-Z]{{1,3}})([1-9][0-9]*)$/);
      if (!match) return null;
      return {{
        coord: raw,
        colLetters: match[1],
        row: Number(match[2]),
        col: columnToNumber(match[1]),
      }};
    }}

    function getInputByCoord(coord) {{
      return grid.querySelector(`input[data-coord="${{coord}}"]`);
    }}

    function getCellNodeByCoord(coord) {{
      return grid.querySelector(`[data-coord="${{coord}}"]`);
    }}

    function syncFormulaBar() {{
      selectedCellEl.value = selectedCoord || "";
      if (!selectedCoord) {{
        formulaBarEl.value = "";
        formulaBarEl.disabled = true;
        applyFormulaBtn.disabled = true;
        return;
      }}
      const input = getInputByCoord(selectedCoord);
      if (input) {{
        formulaBarEl.value = input.value;
        formulaBarEl.disabled = false;
        applyFormulaBtn.disabled = false;
        return;
      }}
      const cellNode = getCellNodeByCoord(selectedCoord);
      formulaBarEl.value = cellNode ? (cellNode.textContent || "") : "";
      formulaBarEl.disabled = true;
      applyFormulaBtn.disabled = true;
    }}

    function selectCoord(coord) {{
      selectedCoord = coord || "";
      highlightSelection();
      syncFormulaBar();
    }}

    function registerDirty(input, sheetName, coord) {{
      const key = `${{sheetName}}!${{coord}}`;
      if (input.value === input.dataset.original) {{
        dirtyCells.delete(key);
        input.classList.remove("dirty");
      }} else {{
        dirtyCells.set(key, {{ sheet: sheetName, cell: coord, value: input.value }});
        input.classList.add("dirty");
      }}
      if (selectedCoord === coord) {{
        formulaBarEl.value = input.value;
      }}
      updateSaveState();
    }}

    function moveSelection(rowOffset, colOffset) {{
      const current = normalizeCoord(selectedCoord);
      if (!current) return;
      const targetRow = Math.max(1, current.row + rowOffset);
      const targetCol = Math.max(1, current.col + colOffset);
      const nextCoord = `${{numberToColumn(targetCol)}}${{targetRow}}`;
      const nextInput = getInputByCoord(nextCoord);
      if (nextInput) {{
        nextInput.focus();
        nextInput.select();
        return;
      }}
      startRowEl.value = Math.max(1, targetRow - 8);
      startColEl.value = numberToColumn(Math.max(1, targetCol - 3));
      selectedCoord = nextCoord;
      loadWindow({{ resetSelection: false }}).then(() => {{
        const loadedInput = getInputByCoord(nextCoord);
        if (loadedInput) {{
          loadedInput.focus();
          loadedInput.select();
        }} else {{
          selectCoord(nextCoord);
        }}
      }}).catch((err) => {{
        console.error(err);
        setStatus(err.message || "移动失败", "error");
      }});
    }}

    function applyFormulaBarValue() {{
      if (!selectedCoord) return;
      const input = getInputByCoord(selectedCoord);
      if (!input) return;
      input.value = formulaBarEl.value;
      registerDirty(input, input.dataset.sheet || currentSheet, selectedCoord);
      input.focus();
    }}

    function buildGrid(payload) {{
      currentPayload = payload;
      currentWindow = {{
        sheet: payload.sheet,
        start_row: payload.start_row,
        start_col: payload.start_col,
        rows: payload.rows_limit,
        cols: payload.cols_limit,
      }};
      sheetNameEl.textContent = payload.sheet;
      startRowEl.value = payload.start_row;
      startColEl.value = payload.start_col_letter;
      rowsLimitEl.value = payload.rows_limit;
      colsLimitEl.value = payload.cols_limit;

      const table = document.createElement("table");
      const thead = document.createElement("thead");
      const headRow = document.createElement("tr");
      const corner = document.createElement("th");
      corner.className = "corner";
      corner.textContent = "#";
      headRow.appendChild(corner);
      payload.col_headers.forEach((col) => {{
        const th = document.createElement("th");
        th.textContent = col;
        headRow.appendChild(th);
      }});
      thead.appendChild(headRow);
      table.appendChild(thead);

      const tbody = document.createElement("tbody");
      payload.rows.forEach((rowCells) => {{
        const tr = document.createElement("tr");
        const rowHeader = document.createElement("th");
        rowHeader.textContent = rowCells.length ? rowCells[0].row : "";
        tr.appendChild(rowHeader);
        rowCells.forEach((cell) => {{
          const td = document.createElement("td");
          if (cell.fill) td.style.background = cell.fill;
          td.title = cell.coord;
          td.dataset.coord = cell.coord;
          if (cell.editable) {{
            const input = document.createElement("input");
            const dirtyKey = `${{payload.sheet}}!${{cell.coord}}`;
            const pending = dirtyCells.get(dirtyKey);
            input.className = "cell editable";
            input.value = pending ? String(pending.value ?? "") : (cell.text || "");
            input.dataset.coord = cell.coord;
            input.dataset.original = cell.text || "";
            input.dataset.sheet = payload.sheet;
            if (pending) {{
              input.classList.add("dirty");
            }}
            input.addEventListener("focus", () => {{
              selectCoord(cell.coord);
            }});
            input.addEventListener("input", () => {{
              registerDirty(input, payload.sheet, cell.coord);
            }});
            input.addEventListener("keydown", (ev) => {{
              if (ev.key === "Enter" && !ev.shiftKey) {{
                ev.preventDefault();
                moveSelection(1, 0);
              }} else if (ev.key === "ArrowUp") {{
                ev.preventDefault();
                moveSelection(-1, 0);
              }} else if (ev.key === "ArrowDown") {{
                ev.preventDefault();
                moveSelection(1, 0);
              }} else if (ev.key === "ArrowLeft") {{
                ev.preventDefault();
                moveSelection(0, -1);
              }} else if (ev.key === "ArrowRight") {{
                ev.preventDefault();
                moveSelection(0, 1);
              }}
            }});
            td.appendChild(input);
          }} else {{
            const span = document.createElement("span");
            span.className = "cell-text locked";
            span.textContent = cell.text || "";
            td.appendChild(span);
            td.addEventListener("click", () => selectCoord(cell.coord));
          }}
          tr.appendChild(td);
        }});
        tbody.appendChild(tr);
      }});
      table.appendChild(tbody);
      grid.innerHTML = "";
      grid.appendChild(table);
      if (!selectedCoord && payload.rows.length && payload.rows[0].length) {{
        const firstEditable = grid.querySelector("input[data-coord]");
        if (firstEditable) {{
          selectedCoord = firstEditable.dataset.coord || "";
        }}
      }}
      highlightSelection();
      syncFormulaBar();
      setStatus(`已加载 ${{payload.sheet}} 的窗口：第 ${{payload.start_row}} 行起，第 ${{payload.start_col_letter}} 列起，共 ${{payload.rows_limit}} 行 × ${{payload.cols_limit}} 列。`, "ready");
    }}

    function highlightSelection() {{
      grid.querySelectorAll(".selected").forEach((el) => el.classList.remove("selected"));
      if (!selectedCoord) return;
      const el = grid.querySelector(`[data-coord="${{selectedCoord}}"]`);
      if (el) el.classList.add("selected");
    }}

    async function loadWindow({{ resetSelection = false }}) {{
      if (resetSelection) selectedCoord = "";
      renderTabs();
      const params = currentParams();
      setStatus("正在读取工作簿窗口…");
      const url = new URL("/api/editor/window", location.origin);
      Object.entries(params).forEach(([key, value]) => url.searchParams.set(key, value));
      const resp = await fetch(url);
      const data = await resp.json();
      if (!resp.ok) {{
        throw new Error(data.error || "读取工作簿失败");
      }}
      buildGrid(data);
      renderTabs();
      updateSaveState();
    }}

    async function saveChanges() {{
      if (!dirtyCells.size) return;
      saveBtn.disabled = true;
      setStatus("正在保存修改…");
      const changes = Array.from(dirtyCells.values()).map((item) => ({{
        cell: item.cell,
        value: item.value,
      }}));
      const resp = await fetch("/api/editor/save", {{
        method: "POST",
        headers: {{ "Content-Type": "application/json" }},
        body: JSON.stringify({{
          sheet: currentSheet,
          changes,
        }}),
      }});
      const data = await resp.json();
      if (!resp.ok || !data.ok) {{
        setStatus(data.error || "保存失败", "error");
        saveBtn.disabled = false;
        return;
      }}
      dirtyCells.clear();
      updateSaveState();
      setStatus(`已保存 ${{data.saved_count || changes.length}} 个单元格。`, "success");
      await loadWindow({{ resetSelection: false }});
    }}

    gotoBtn.addEventListener("click", () => {{
      const parsed = normalizeCoord(gotoCellEl.value);
      if (!parsed) {{
        setStatus("请输入类似 BX27 的单元格坐标。", "error");
        return;
      }}
      startRowEl.value = Math.max(1, parsed.row - 8);
      startColEl.value = numberToColumn(Math.max(1, parsed.col - 3));
      selectedCoord = parsed.coord;
      loadWindow({{ resetSelection: false }}).then(() => {{
        const input = getInputByCoord(parsed.coord);
        if (input) {{
          input.focus();
          input.select();
        }} else {{
          selectCoord(parsed.coord);
        }}
      }}).catch((err) => {{
        console.error(err);
        setStatus(err.message || "定位失败", "error");
      }});
    }});

    refreshBtn.addEventListener("click", () => loadWindow({{ resetSelection: false }}));
    saveBtn.addEventListener("click", () => saveChanges());
    applyFormulaBtn.addEventListener("click", () => applyFormulaBarValue());
    formulaBarEl.addEventListener("keydown", (ev) => {{
      if (ev.key === "Enter") {{
        ev.preventDefault();
        applyFormulaBarValue();
      }}
    }});
    gotoCellEl.addEventListener("keydown", (ev) => {{
      if (ev.key === "Enter") {{
        ev.preventDefault();
        gotoBtn.click();
      }}
    }});

    document.addEventListener("keydown", (ev) => {{
      if ((ev.ctrlKey || ev.metaKey) && ev.key.toLowerCase() === "s") {{
        ev.preventDefault();
        saveChanges();
      }}
    }});

    window.addEventListener("beforeunload", (ev) => {{
      if (dirtyCells.size) {{
        ev.preventDefault();
        ev.returnValue = "";
      }}
    }});

    renderTabs();
    loadWindow({{ resetSelection: true }}).catch((err) => {{
      console.error(err);
      setStatus(err.message || "加载失败", "error");
    }});
  </script>
</body>
</html>
"""
    return page.encode("utf-8")


def read_json_body(handler: BaseHTTPRequestHandler) -> dict:
    content_length = int(handler.headers.get("Content-Length", "0") or "0")
    if content_length <= 0:
        return {}
    body = handler.rfile.read(content_length).decode("utf-8")
    if not body.strip():
        return {}
    return json.loads(body)


def handle_editor_window(handler: BaseHTTPRequestHandler, head: bool = False) -> None:
    try:
        if not master_sales_exists():
            raise ValueError("还没有共用销售排单可预览，请先上传本周排单。")
        query = urllib.parse.parse_qs(urllib.parse.urlparse(handler.path).query)
        sheet_name = query.get("sheet", [""])[0] or ""
        try:
            start_row = int(query.get("start_row", ["1"])[0])
        except ValueError:
            start_row = 1
        start_col = parse_column_ref(query.get("start_col", ["1"])[0], default=1)
        try:
            rows_limit = int(query.get("rows", ["80"])[0])
        except ValueError:
            rows_limit = 80
        try:
            cols_limit = int(query.get("cols", ["40"])[0])
        except ValueError:
            cols_limit = 40
        with STATE_LOCK:
            wb = load_workbook(MASTER_SALES_PATH, read_only=True, data_only=False)
            try:
                payload = editor_window_payload(
                    wb,
                    resolve_sheet_name(sheet_name or wb.sheetnames[0], list(wb.sheetnames)),
                    start_row,
                    start_col,
                    rows_limit,
                    cols_limit,
                )
            finally:
                wb.close()
        write_json(handler, 200, payload, head=head)
    except Exception as exc:  # noqa: BLE001
        write_json(handler, 400, {"ok": False, "error": str(exc)}, head=head)


def handle_editor_save(handler: BaseHTTPRequestHandler, head: bool = False) -> None:
    try:
        if job_is_running():
            raise ValueError("当前有预测正在后台回填，请处理完成后再编辑工作簿。")
        if not master_sales_exists():
            raise ValueError("还没有共用销售排单可编辑，请先上传本周排单。")

        payload = read_json_body(handler)
        sheet_name = payload.get("sheet") or ""
        changes = payload.get("changes") or []
        if not sheet_name:
            raise ValueError("缺少 Sheet 名称。")
        if not isinstance(changes, list):
            raise ValueError("修改数据格式不正确。")
        if not changes:
            write_json(handler, 200, {"ok": True, "saved_count": 0}, head=head)
            return

        saved_count = 0
        with STATE_LOCK:
            wb = load_workbook(MASTER_SALES_PATH)
            try:
                sheet_name = resolve_sheet_name(sheet_name, list(wb.sheetnames))
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"没有找到 Sheet：{sheet_name}")
                ws = wb[sheet_name]
                for item in changes:
                    cell = str(item.get("cell") or "").strip().upper().replace(" ", "")
                    raw_value = str(item.get("value") or "")
                    cell_match = re.fullmatch(r"([A-Z]{1,3})([1-9][0-9]{0,6})", cell)
                    if not cell_match:
                        raise ValueError(f"单元格格式不正确：{cell}")
                    cell_col = column_index_from_string(cell_match.group(1))
                    cell_row = int(cell_match.group(2))
                    target_cell = ws[cell]
                    if isinstance(target_cell, MergedCell):
                        raise ValueError(f"{cell} 是合并单元格的从属格，请修改该合并区域左上角单元格。")
                    ws[cell].value = parse_workbook_cell_value(raw_value)
                    saved_count += 1
                wb.save(MASTER_SALES_PATH)
                shutil.copy2(MASTER_SALES_PATH, LATEST_OUTPUT_PATH)
                sync_master_files_to_remote()
                metadata = ensure_metadata_schema(load_metadata())
                metadata["last_manual_edit_at"] = now_label()
                metadata["last_manual_edit_cell"] = f"{sheet_name}!{changes[-1].get('cell', '')}"
                metadata["last_updated_rows"] = str(saved_count)
                save_metadata(metadata)
            finally:
                wb.close()

        write_json(handler, 200, {"ok": True, "saved_count": saved_count}, head=head)
    except Exception as exc:  # noqa: BLE001
        write_json(handler, 400, {"ok": False, "error": str(exc)}, head=head)


def handle_edit_cell(handler: BaseHTTPRequestHandler, head: bool = False) -> None:
    selected_sheet = ""
    try:
        if job_is_running():
            raise ValueError("当前有预测正在后台回填，请处理完成后再编辑本周预估栏。")
        if not master_sales_exists():
            raise ValueError("还没有共用销售排单可编辑，请先上传本周排单。")

        content_length = int(handler.headers.get("Content-Length", "0") or "0")
        body = handler.rfile.read(content_length).decode("utf-8")
        form = urllib.parse.parse_qs(body, keep_blank_values=True)
        selected_sheet = form.get("sheet", [""])[0] or ""
        cell = (form.get("cell", [""])[0] or "").strip().upper().replace(" ", "")
        raw_value = form.get("value", [""])[0]
        try:
            start_row = int(form.get("start_row", ["1"])[0])
        except ValueError:
            start_row = 1
        start_col = parse_column_ref(form.get("start_col", ["1"])[0], default=1)
        try:
            rows_limit = int(form.get("rows", ["80"])[0])
        except ValueError:
            rows_limit = 80
        try:
            cols_limit = int(form.get("cols", ["140"])[0])
        except ValueError:
            cols_limit = 140
        full_view = (form.get("full", ["0"])[0] or "0").strip() in {"1", "true", "yes", "on"}

        cell_match = re.fullmatch(r"([A-Z]{1,3})([1-9][0-9]{0,6})", cell)
        if not cell_match:
            raise ValueError("单元格格式不正确。")
        cell_col = column_index_from_string(cell_match.group(1))
        cell_row = int(cell_match.group(2))

        with STATE_LOCK:
            wb = load_workbook(MASTER_SALES_PATH)
            try:
                selected_sheet = resolve_sheet_name(selected_sheet, list(wb.sheetnames))
                if selected_sheet not in wb.sheetnames:
                    raise ValueError(f"没有找到 Sheet：{selected_sheet}")
                editable_columns = editable_forecast_columns(wb, selected_sheet)
                if cell_row < DATA_START_ROW or cell_col not in editable_columns:
                    raise ValueError("该单元格属于原始数据、历史预估或差异栏，不能在线修改。只能修改浅绿色的本周预估数量/金额格。")

                ws = wb[selected_sheet]
                ws[cell].value = parse_editable_forecast_value(raw_value)
                wb.save(MASTER_SALES_PATH)
                shutil.copy2(MASTER_SALES_PATH, LATEST_OUTPUT_PATH)
                sync_master_files_to_remote()

                metadata = ensure_metadata_schema(load_metadata())
                metadata["last_manual_edit_at"] = now_label()
                metadata["last_manual_edit_cell"] = f"{selected_sheet}!{cell}"
                save_metadata(metadata)
            finally:
                wb.close()

        write_html(
            handler,
            200,
            render_preview_page(
                message=f"已保存 {selected_sheet}!{cell} 的本周预估修改。",
                selected_sheet=selected_sheet,
                start_row=start_row,
                start_col=start_col,
                rows_limit=rows_limit,
                cols_limit=cols_limit,
                full_view=full_view,
            ).decode("utf-8"),
            head=head,
        )
    except Exception as exc:  # noqa: BLE001
        write_html(
            handler,
            400,
            render_preview_page(error=f"保存失败：{exc}", selected_sheet=selected_sheet, full_view=full_view).decode("utf-8"),
            head=head,
        )


def handle_onlyoffice_callback(handler: BaseHTTPRequestHandler, head: bool = False) -> None:
    try:
        payload = read_json_body(handler)
        status = payload.get("status")
        try:
            status = int(status)
        except (TypeError, ValueError):
            status = None

        callback_url = onlyoffice_callback_url(handler)
        if not callback_url:
            raise ValueError("无法确定回传地址。")

        if status in {2, 6}:
            download_url = (payload.get("url") or "").strip()
            if not download_url:
                raise ValueError("OnlyOffice 回调缺少文件下载地址。")
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                tmp_path = Path(tmp_file.name)
                with urllib.request.urlopen(download_url, timeout=180) as response:
                    shutil.copyfileobj(response, tmp_file)
            try:
                if not zipfile.is_zipfile(tmp_path):
                    raise ValueError("OnlyOffice 返回的文件不是有效的 xlsx。")

                with STATE_LOCK:
                    ensure_data_dir()
                    if MASTER_SALES_PATH.exists():
                        backup_name = "onlyoffice_" + beijing_now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
                        shutil.copy2(MASTER_SALES_PATH, BACKUP_DIR / backup_name)
                    shutil.copy2(tmp_path, MASTER_SALES_PATH)
                    shutil.copy2(tmp_path, LATEST_OUTPUT_PATH)
                    sync_master_files_to_remote()
                    metadata = ensure_metadata_schema(load_metadata())
                    metadata["latest_name"] = metadata.get("latest_name") or metadata.get("master_name") or "当前最新版.xlsx"
                    metadata["last_owner"] = "在线编辑"
                    metadata["last_generated_at"] = now_label()
                    metadata["last_updated_rows"] = "整本表格"
                    save_metadata(metadata)
            finally:
                try:
                    tmp_path.unlink(missing_ok=True)
                except Exception:
                    pass

        write_json(handler, 200, {"error": 0}, head=head)
    except Exception as exc:  # noqa: BLE001
        write_json(handler, 200, {"error": 1, "message": str(exc)}, head=head)


def handle_sales_master(handler: BaseHTTPRequestHandler, head: bool = False) -> None:
    if handler.headers.get_content_type() != UPLOAD_MIME:
        write_html(
            handler,
            400,
            render_page(error="请使用网页表单上传共用销售排单。", handler=handler).decode("utf-8"),
            head=head,
        )
        return

    backup_path: Optional[Path] = None
    storage_warning = ""
    try:
        if job_is_running():
            raise ValueError("当前有预测正在后台回填，请处理完成后再替换共用销售排单。")

        form = cgi.FieldStorage(
            fp=handler.rfile,
            headers=handler.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": handler.headers.get("Content-Type", ""),
                "CONTENT_LENGTH": handler.headers.get("Content-Length", "0"),
            },
        )
        sales_field = first_upload(form, "sales_file")
        if not getattr(sales_field, "filename", None):
            raise ValueError("请先选择共用销售排单文件。")

        sales_suffix = uploaded_suffix(sales_field.filename)
        if sales_suffix not in SUPPORTED_SALES_SUFFIXES:
            raise ValueError("共用销售排单请上传 .xlsx 或 .xlsm 文件。")

        original_name = safe_filename(sales_field.filename)
        with STATE_LOCK:
            ensure_data_dir()
            if master_sales_exists():
                backup_name = beijing_now().strftime("%Y%m%d_%H%M%S") + "_" + MASTER_SALES_PATH.name
                backup_path = BACKUP_DIR / backup_name
                shutil.copy2(MASTER_SALES_PATH, backup_path)
            save_upload(sales_field, MASTER_SALES_PATH)
            validate_excel_file(MASTER_SALES_PATH, "共用销售排单")
            shutil.copy2(MASTER_SALES_PATH, LATEST_OUTPUT_PATH)
            try:
                sync_master_files_to_remote()
            except Exception as exc:  # noqa: BLE001
                storage_warning = f"共用排单已保存，但远程备份暂时失败：{exc}"
                print(storage_warning)
            metadata = {
                "master_name": original_name,
                "master_uploaded_at": now_label(),
                "latest_name": f"{Path(original_name).stem}_当前最新版.xlsx",
                "last_owner": "",
                "last_generated_at": "",
                "last_updated_rows": "",
                "owner_statuses": default_owner_statuses(),
                "storage_warning": storage_warning,
                "storage_warning_at": now_label() if storage_warning else "",
            }
            metadata_warning = save_metadata(metadata)
            if metadata_warning:
                storage_warning = "；".join(
                    value for value in (storage_warning, metadata_warning) if value
                )

        message_text = f"共用销售排单已保存：{original_name}。现在各业务只需要上传自己的预测。"
        if storage_warning:
            message_text += " 当前文件可继续使用，网站会在后续保存时再次尝试远程备份。"
        write_html(
            handler,
            200,
            render_page(
                message=message_text,
                handler=handler,
            ).decode("utf-8"),
            head=head,
        )
    except Exception as exc:  # noqa: BLE001
        if backup_path and backup_path.exists():
            try:
                shutil.copy2(backup_path, MASTER_SALES_PATH)
                shutil.copy2(backup_path, LATEST_OUTPUT_PATH)
            except Exception:
                pass
        write_html(
            handler,
            400,
            render_page(error=f"保存共用排单失败：{exc}", handler=handler).decode("utf-8"),
            head=head,
        )


def process_prediction_job(
    selected_owner: str,
    pred_path: Path,
    overwrite_pairs: list[str],
    allow_legacy_overwrite: bool,
) -> None:
    storage_warning = ""

    def progress_callback(progress: dict) -> None:
        target_parts = []
        if progress.get("target_sheet"):
            target_parts.append(str(progress["target_sheet"]))
        if progress.get("target_label"):
            target_parts.append(str(progress["target_label"]))
        if progress.get("target_month"):
            target_parts.append(f'{progress["target_month"]}月')
        set_job_status(
            progress=format_progress(progress.get("percent")),
            step=str(progress.get("step") or ""),
            target=" / ".join(target_parts),
        )

    set_job_status(
        state="running",
        owner=selected_owner,
        message="正在回填共用销售排单",
        started_at=now_label(),
        finished_at="",
        updated_rows="",
        error="",
        storage_warning="",
        progress="1",
        step="准备处理上传文件",
        target="",
    )
    try:
        with tempfile.TemporaryDirectory(prefix="sales_job_") as tmp_dir:
            tmp_path = Path(tmp_dir)
            working_sales_path = tmp_path / "current_sales.xlsx"
            output_path = tmp_path / "generated.xlsx"

            with STATE_LOCK:
                if not master_sales_exists():
                    raise ValueError("共用销售排单不存在，请先重新上传本周排单。")
                set_job_status(progress="3", step="复制当前共用销售排单", target="")
                ensure_data_dir()
                shutil.copy2(MASTER_SALES_PATH, working_sales_path)
                validate_excel_file(working_sales_path, "当前共用销售排单")
            validate_prediction_file(pred_path, f"{selected_owner} 的预测文件")

            summary = process_sales_workbooks(
                pred_path=pred_path,
                sales_path=working_sales_path,
                output_path=output_path,
                # The shared weekly schedule can contain hundreds of thousands
                # of formulas. Freezing them during a web upload makes the
                # request time out, so the website keeps formulas live.
                freeze_formulas=False,
                business_owner=selected_owner,
                as_of_date=beijing_now().date(),
                progress_callback=progress_callback,
                overwrite_pairs=overwrite_pairs,
                allow_legacy_overwrite=allow_legacy_overwrite,
            )

            with STATE_LOCK:
                if not master_sales_exists():
                    raise ValueError("共用销售排单不存在，请先重新上传本周排单。")
                set_job_status(progress="97", step="备份并替换共用销售排单", target="")
                backup_name = beijing_now().strftime("%Y%m%d_%H%M%S") + f"_{selected_owner}_before.xlsx"
                shutil.copy2(MASTER_SALES_PATH, BACKUP_DIR / backup_name)
                shutil.copy2(output_path, MASTER_SALES_PATH)
                shutil.copy2(output_path, LATEST_OUTPUT_PATH)
                set_job_status(progress="98", step="同步保存共享网站数据", target="")
                try:
                    sync_master_files_to_remote()
                except Exception as exc:  # noqa: BLE001
                    storage_warning = f"Excel 已回填完成，但远程备份暂时失败：{exc}"
                    print(storage_warning)

                metadata = load_metadata()
                master_name = metadata.get("master_name", "销售排单.xlsx")
                stem = Path(master_name).stem if master_name else "销售排单"
                download_name = f"{stem}_{selected_owner}_已回填_当前最新版.xlsx"
                metadata.update(
                    {
                        "latest_name": download_name,
                        "last_owner": selected_owner,
                        "last_generated_at": now_label(),
                        "last_updated_rows": summary["updated_rows"],
                        "last_matched_model_count": summary.get("matched_model_count", 0),
                        "storage_warning": storage_warning,
                        "storage_warning_at": now_label() if storage_warning else "",
                    }
                )
                metadata = ensure_metadata_schema(metadata)
                owner_result = {
                    "state": "done",
                    "updated_rows": str(summary["updated_rows"]),
                    "progress": "100",
                    "updated_at": now_label(),
                    "error": "",
                    "unmatched_predictions": summary.get("unmatched_predictions", []),
                    "fill_target_months": summary.get("fill_target_months", []),
                    "matched_model_count": str(summary.get("matched_model_count", 0)),
                    "storage_warning": storage_warning,
                }
                if summary.get("written_pairs"):
                    owner_result["written_pairs"] = summary["written_pairs"]
                metadata["owner_statuses"][selected_owner].update(owner_result)
                metadata_warning = save_metadata(metadata)
                if metadata_warning:
                    storage_warning = "；".join(
                        value for value in (storage_warning, metadata_warning) if value
                    )

        set_job_status(
            state="done",
            owner=selected_owner,
            message="处理完成" if not storage_warning else "回填完成，远程备份待恢复",
            finished_at=now_label(),
            updated_rows=summary["updated_rows"],
            matched_model_count=summary.get("matched_model_count", 0),
            error="",
            storage_warning=storage_warning,
            progress="100",
            step="处理完成",
            target="",
        )
    except Exception as exc:  # noqa: BLE001
        set_job_status(
            state="error",
            owner=selected_owner,
            message="处理失败",
            finished_at=now_label(),
            updated_rows="",
            error=str(exc),
            progress=format_progress(get_job_status().get("progress")),
            step="处理失败",
        )
        update_owner_status(
            selected_owner,
            state="error",
            updated_rows="",
            progress=format_progress(get_job_status().get("progress")),
            updated_at=now_label(),
            error=str(exc),
        )


def handle_generate(handler: BaseHTTPRequestHandler, head: bool = False) -> None:
    if handler.headers.get_content_type() != UPLOAD_MIME:
        write_html(
            handler,
            400,
            render_page(error="请使用网页表单上传预测信息。", handler=handler).decode("utf-8"),
            head=head,
        )
        return

    selected_owner = DEFAULT_OWNER
    try:
        form = cgi.FieldStorage(
            fp=handler.rfile,
            headers=handler.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": handler.headers.get("Content-Type", ""),
                "CONTENT_LENGTH": handler.headers.get("Content-Length", "0"),
            },
        )
        selected_owner = safe_owner(form.getfirst("business_owner", DEFAULT_OWNER))
        pred_field = first_upload(form, "prediction_file")
        if not getattr(pred_field, "filename", None):
            raise ValueError("请先选择预测信息文件。")

        pred_suffix = uploaded_suffix(pred_field.filename)
        if pred_suffix not in SUPPORTED_PREDICTION_SUFFIXES:
            raise ValueError("预测信息请上传 .xlsx、.jpg 或 .png 文件。")
        if not master_sales_exists():
            raise ValueError("请先在页面上方上传本周共用销售排单。")
        if job_is_running():
            raise ValueError("上一份预测还在处理，请稍后刷新页面，完成后再上传下一位业务预测。")

        overwrite_pairs, allow_legacy_overwrite = owner_overwrite_context(selected_owner)

        ensure_data_dir()
        pred_name = beijing_now().strftime("%Y%m%d_%H%M%S") + f"_{selected_owner}{pred_suffix}"
        pred_path = UPLOAD_DIR / pred_name
        save_upload(pred_field, pred_path)
        validate_prediction_file(pred_path, f"{selected_owner} 的预测文件")
        update_owner_status(
            selected_owner,
            state="running",
            updated_rows="",
            progress="1",
            uploaded_at=now_label(),
            updated_at="",
            prediction_name=safe_filename(pred_field.filename, "预测文件"),
            error="",
            storage_warning="",
            unmatched_predictions=[],
            fill_target_months=[],
            matched_model_count="",
        )
        set_job_status(
            state="running",
            owner=selected_owner,
            message="正在回填共用销售排单",
            started_at=now_label(),
            finished_at="",
            updated_rows="",
            error="",
            storage_warning="",
            progress="1",
            step="已收到预测文件，等待后台处理",
            target="",
        )

        worker = threading.Thread(
            target=process_prediction_job,
            args=(
                selected_owner,
                pred_path,
                overwrite_pairs,
                allow_legacy_overwrite,
            ),
            daemon=True,
        )
        worker.start()

        write_html(
            handler,
            202,
            render_page(
                message=f"{selected_owner} 的预测已上传，网站正在后台回填。请稍后刷新页面，完成后点击“下载当前最新版”。",
                handler=handler,
                selected_owner=selected_owner,
            ).decode("utf-8"),
            head=head,
        )
        return
    except Exception as exc:  # noqa: BLE001
        write_html(
            handler,
            400,
            render_page(error=f"生成失败：{exc}", handler=handler, selected_owner=selected_owner).decode("utf-8"),
            head=head,
        )


class SalesUploadHandler(BaseHTTPRequestHandler):
    server_version = "SalesUploadHTTP/4.0"

    def log_message(self, format: str, *args) -> None:  # noqa: A003
        print(f"{self.address_string()} - {format % args}")

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        query = urllib.parse.parse_qs(parsed.query)
        if path == "/healthz":
            write_text(self, 200, "ok")
            return
        if path == "/login":
            if authenticated(self):
                redirect(self, "/")
            else:
                write_html(self, 200 if login_is_configured() else 503, render_login_page())
            return
        if path == "/logout":
            redirect(self, "/login", make_session_cookie(self, clear=True))
            return
        if not authenticated(self):
            redirect(self, "/login")
            return
        if path in ("/", "/index.html") or path.startswith("/owner/"):
            owner = selected_owner_from_path(path, query)
            write_html(self, 200, render_page(handler=self, selected_owner=owner).decode("utf-8"))
            return

        if path in ("/generate", "/sales-master"):
            owner = selected_owner_from_path("/", query)
            write_html(self, 200, render_page(handler=self, selected_owner=owner).decode("utf-8"))
            return

        if path == "/preview":
            write_html(
                self,
                404,
                render_page(
                    error="在线预览和编辑功能已暂时关闭，请下载当前最新版查看。",
                    handler=self,
                ).decode("utf-8"),
            )
            return

        if path == "/table-preview":
            write_html(
                self,
                404,
                render_page(
                    error="在线预览和编辑功能已暂时关闭，请下载当前最新版查看。",
                    handler=self,
                ).decode("utf-8"),
            )
            return

        if path == "/api/editor/window":
            write_text(self, 404, "Preview and editing are temporarily disabled")
            return

        if path == "/api/onlyoffice/callback":
            write_text(self, 404, "Preview and editing are temporarily disabled")
            return

        if path == "/status":
            metadata = ensure_metadata_schema(load_metadata())
            write_json(
                self,
                200,
                {
                    "job": get_job_status(),
                    "master_exists": master_sales_exists(),
                    "master_name": metadata.get("master_name", ""),
                    "latest_name": metadata.get("latest_name", ""),
                    "last_owner": metadata.get("last_owner", ""),
                    "last_generated_at": metadata.get("last_generated_at", ""),
                    "last_updated_rows": metadata.get("last_updated_rows", ""),
                    "storage_warning": metadata.get("storage_warning", ""),
                    "owner_statuses": metadata.get("owner_statuses", default_owner_statuses()),
                    "data_dir": str(DATA_DIR),
                    "storage": storage_description(),
                },
            )
            return

        if path == "/download/latest":
            if not master_sales_exists():
                write_html(
                    self,
                    404,
                    render_page(error="还没有共用销售排单可下载，请先上传本周排单。", handler=self).decode("utf-8"),
                )
                return
            metadata = load_metadata()
            download_name = metadata.get("latest_name") or metadata.get("master_name") or "销售排单_当前最新版.xlsx"
            send_xlsx(self, MASTER_SALES_PATH, download_name)
            return

        write_text(self, 404, "Not found")

    def do_HEAD(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        query = urllib.parse.parse_qs(parsed.query)
        if path == "/healthz":
            write_text(self, 200, "ok", head=True)
            return
        if path == "/login":
            write_html(self, 200 if login_is_configured() else 503, render_login_page(), head=True)
            return
        if not authenticated(self):
            redirect(self, "/login")
            return
        if path in ("/", "/index.html") or path.startswith("/owner/"):
            owner = selected_owner_from_path(path, query)
            write_html(self, 200, render_page(handler=self, selected_owner=owner).decode("utf-8"), head=True)
            return

        if path == "/preview":
            write_text(self, 404, "Not found", head=True)
            return

        if path == "/table-preview":
            write_text(self, 404, "Not found", head=True)
            return

        if path == "/api/editor/window":
            write_text(self, 404, "Not found", head=True)
            return

        if path == "/api/onlyoffice/callback":
            write_text(self, 404, "Not found", head=True)
            return

        if path == "/status":
            write_json(self, 200, {"ok": True}, head=True)
            return

        if path == "/download/latest":
            if not master_sales_exists():
                write_text(self, 404, "Not found", head=True)
                return
            metadata = load_metadata()
            download_name = metadata.get("latest_name") or metadata.get("master_name") or "销售排单_当前最新版.xlsx"
            send_xlsx(self, MASTER_SALES_PATH, download_name, head=True)
            return

        write_text(self, 404, "Not found", head=True)

    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/login":
            handle_login(self)
            return
        if not authenticated(self):
            redirect(self, "/login")
            return
        if parsed.path == "/api/editor/save":
            write_text(self, 404, "Preview and editing are temporarily disabled")
            return
        if parsed.path == "/api/onlyoffice/callback":
            write_text(self, 404, "Preview and editing are temporarily disabled")
            return
        if parsed.path == "/sales-master":
            handle_sales_master(self)
            return

        if parsed.path == "/edit-cell":
            write_text(self, 404, "Preview and editing are temporarily disabled")
            return

        if parsed.path != "/generate":
            write_text(self, 404, "Not found")
            return

        handle_generate(self)


def main() -> None:
    server = ThreadingHTTPServer((HOST, PORT), SalesUploadHandler)
    public_url = get_public_share_url()
    print(f"销售排单网站已启动：http://0.0.0.0:{PORT}")
    print("网站已启用登录保护；登录后可按业务担当页面上传预测信息。")
    if not login_is_configured():
        print("尚未配置 SALES_LOGIN_PASSWORD / SALES_LOGIN_SECRET，数据页面将保持锁定。")
    if public_url:
        print(f"公开访问：{public_url}/")
    else:
        lan_ip = get_lan_ip()
        print(f"本机访问：http://127.0.0.1:{PORT}")
        print(f"局域网访问：http://{lan_ip}:{PORT}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n已停止。")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
